"""
excel_export.py — Excel Workbook Export Builder

Purpose:
- Convert model outputs (projections, DCF, comps) into a structured .xlsx file.
- Populate Excel template with historical financial data from structured JSON.
- This is the final client-facing artifact used by analysts and investors.
- Template-driven approach to ensure readability and familiarity.

Inputs:
- company_id
- model_version (optional)
- DB session (for historicals + assumptions)
- Structured JSON files (for template population)
- Must call:
    run_three_statement()
    run_dcf()
    run_comps()
  to gather values prior to export.

Outputs:
- In-memory .xlsx file (returned as bytes or BytesIO stream).
- Populated Excel template with historical data.

This module does NOT:
- Perform valuation logic.
- Fetch filings or prices.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
import json
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.services.modeling.three_statement import run_three_statement
from app.services.modeling.dcf import run_dcf
from app.services.modeling.comps import run_comps
from app.services.modeling.types import (
    CompanyModelInput,
    build_company_model_input,
    HistoricalSeries,
    DcfOutput,
)
from app.services.modeling.excel_dcf_scenarios import (
    build_scenario_formula_dict,
    apply_scenario_to_sheet,
)

# Use openpyxl for reading/modifying existing Excel files
# Use xlsxwriter for creating new workbooks
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Font = None

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

logger = get_logger(__name__)

# Model role to row label mappings for Excel template
# Maps row labels (from column B) to model_role values
# Order matters: more specific patterns should come first
ROW_LABEL_TO_MODEL_ROLE: Dict[str, str] = {
    # Revenue
    "revenue": "IS_REVENUE",
    "sales": "IS_REVENUE",
    "net sales": "IS_REVENUE",
    "total revenue": "IS_REVENUE",
    
    # COGS - match patterns with parentheses and dashes first, and more specific patterns first
    "(-) cogs": "IS_COGS",
    "cost of revenue": "IS_COGS",  # Must come before "revenue" to avoid false match
    "cost of goods sold": "IS_COGS",
    "cost of goods": "IS_COGS",
    "cost of sales": "IS_COGS",
    "cogs": "IS_COGS",
    
    # Gross Profit
    "gross profit": "IS_GROSS_PROFIT",
    "gross income": "IS_GROSS_PROFIT",
    
    # Operating Expenses
    "(-) op ex": "IS_OPERATING_EXPENSE",
    "(-) operating expenses": "IS_OPERATING_EXPENSE",
    "(-) opex": "IS_OPERATING_EXPENSE",
    "op ex": "IS_OPERATING_EXPENSE",
    "operating expenses": "IS_OPERATING_EXPENSE",
    "opex": "IS_OPERATING_EXPENSE",
    "operating expense": "IS_OPERATING_EXPENSE",
    "sg&a": "IS_OPERATING_EXPENSE",
    "selling, general and administrative": "IS_OPERATING_EXPENSE",
    
    # Depreciation & Amortization
    "(+) d&a": "CF_DEPRECIATION",
    "(+) depreciation": "CF_DEPRECIATION",
    "(+) depreciation & amortization": "CF_DEPRECIATION",
    "d&a": "CF_DEPRECIATION",
    "depreciation & amortization": "CF_DEPRECIATION",
    "depreciation and amortization": "CF_DEPRECIATION",
    
    # Capital Expenditures
    "(-) capex": "CF_CAPEX",
    "(-) capital expenditures": "CF_CAPEX",
    "capex": "CF_CAPEX",
    "capital expenditures": "CF_CAPEX",
    "capital expense": "CF_CAPEX",
    
    # Changes in Working Capital
    "(-) changes in wc": "CF_CHANGE_IN_WORKING_CAPITAL",
    "(-) change in wc": "CF_CHANGE_IN_WORKING_CAPITAL",
    "(-) changes in working capital": "CF_CHANGE_IN_WORKING_CAPITAL",
    "(-) change in working capital": "CF_CHANGE_IN_WORKING_CAPITAL",
    "changes in wc": "CF_CHANGE_IN_WORKING_CAPITAL",
    "change in wc": "CF_CHANGE_IN_WORKING_CAPITAL",
    "changes in working capital": "CF_CHANGE_IN_WORKING_CAPITAL",
    "change in working capital": "CF_CHANGE_IN_WORKING_CAPITAL",
    
    # Income Statement items
    "net income": "IS_NET_INCOME",
    "net interest income": "IS_NON_OPERATING_INCOME",
    "income before taxes": "IS_OPERATING_INCOME",
    "income before tax": "IS_OPERATING_INCOME",
    "income tax expense": "IS_TAX_EXPENSE",
    "operating profit": "IS_OPERATING_INCOME",
    "operating income": "IS_OPERATING_INCOME",
    "eps": "IS_NET_INCOME",
    "ebit": "IS_EBIT",
    "ebitda": "IS_EBITDA",
    "nopat": "IS_NOPAT",
    
    # Balance Sheet items
    "cash": "BS_CASH",
    "cash & equivalents": "BS_CASH",
    "cash and equivalents": "BS_CASH",
    "pp&e": "BS_PP_AND_E",
    "property, plant and equipment": "BS_PP_AND_E",
    "ppe": "BS_PP_AND_E",
    "inventory": "BS_INVENTORY",
    "accounts receivable": "BS_ACCOUNTS_RECEIVABLE",
    "ar": "BS_ACCOUNTS_RECEIVABLE",
    "accounts payable": "BS_ACCOUNTS_PAYABLE",
    "ap": "BS_ACCOUNTS_PAYABLE",
    "accrued expenses": "BS_ACCRUED_LIABILITIES",
    "accrued liabilities": "BS_ACCRUED_LIABILITIES",
    "total debt": "BS_DEBT_CURRENT",
    "debt": "BS_DEBT_CURRENT",
    
    # Cash Flow items
    "share repurchases": "CF_SHARE_REPURCHASES",
    "share buybacks": "CF_SHARE_REPURCHASES",
    "dividends paid": "CF_DIVIDENDS_PAID",
    "dividends": "CF_DIVIDENDS_PAID",
    "debt issued": "CF_DEBT_ISSUANCE",
    "debt repaid": "CF_DEBT_REPAYMENT",
}


def load_structured_json(path: str) -> Dict[str, Any]:
    """
    Load structured JSON file containing financial data.
    
    This is the standard entry point for loading financial JSON files across the backend.
    It handles multiple JSON formats and transforms them to a consistent standard format.
    
    Supported formats:
    - Standard format: {"company": {...}, "filings": [...]}
    - Array format: [{date, symbol, revenue, costOfRevenue, ...}, ...] 
      (e.g., F_income_statement_stable_raw.json - the final structured output format)
    
    The array format is automatically transformed to standard format for compatibility
    with existing code that expects the standard structure.
    
    Args:
        path: Path to JSON file
        
    Returns:
        Dictionary containing structured financial data in standard format:
        {
            "company": {"ticker": str, "company_name": str},
            "filings": [{
                "statements": {
                    "income_statement": {"line_items": [...]},
                    ...
                }
            }]
        }
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # Check if it's an array format (like F_income_statement_stable_raw.json)
    if isinstance(data, list) and len(data) > 0:
        # Transform array format to standard format
        return _transform_array_format_to_standard(data)
    
    return data


def convert_assumptions_to_legacy_format(assumptions: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert new assumptions format to legacy format for compatibility.
    
    New format (project JSON):
    {
        "projectId": "...",
        "incomeStatement": {
            "revenue": {"method": "step", "stepValue": "1"},
            "grossMargin": {"method": "stable", "stableValue": "0.4"},
            ...
        },
        "balanceSheet": {...},
        "cashFlow": {...},
        "relativeValuation": {"competitors": [...]}
    }
    
    Legacy format:
    {
        "assumptions": {
            "revenue": {"type": "step", "values": [0.05, 0.06, ...]},
            ...
        },
        "balance_sheet": {...},
        "cash_flow": {...},
        "competitors": [...]
    }
    
    Args:
        assumptions: Assumptions dictionary in either format
        
    Returns:
        Assumptions dictionary in legacy format
    """
    # Check if it's already in legacy format
    if "assumptions" in assumptions or ("competitors" in assumptions and "projectId" not in assumptions):
        # Already in legacy format, but extract DCF assumptions to top level if present
        if "dcf" in assumptions:
            dcf_data = assumptions["dcf"]
            # Extract DCF assumptions to top level for populate_wacc_and_assumptions
            if "market_risk_premium" in dcf_data:
                assumptions["market_risk_premium"] = dcf_data["market_risk_premium"]
            if "risk_free_rate" in dcf_data:
                assumptions["risk_free_rate"] = dcf_data["risk_free_rate"]
            if "cost_of_debt" in dcf_data:
                assumptions["cost_of_debt"] = dcf_data["cost_of_debt"]
            if "tax_rate" in dcf_data:
                assumptions["tax_rate"] = dcf_data["tax_rate"]
        return assumptions
    
    # Check if it's new format
    if "projectId" in assumptions or "incomeStatement" in assumptions:
        logger.info("Converting new assumptions format to legacy format")
        legacy = {}
        
        # Convert competitors
        if "relativeValuation" in assumptions:
            legacy["competitors"] = assumptions["relativeValuation"].get("competitors", [])
        elif "competitors" in assumptions:
            legacy["competitors"] = assumptions["competitors"]
        
        # Helper function to convert a single assumption value
        def convert_assumption_value(assumption_obj: Any, forecast_periods: int = 5) -> Dict[str, Any]:
            """Convert single assumption from new format to legacy format."""
            if isinstance(assumption_obj, dict):
                method = assumption_obj.get("method", "").lower()
                
                # Map method names: "stable" -> "constant", "step" -> "step", "manual" -> "custom"
                method_map = {
                    "stable": "constant",
                    "step": "step",
                    "manual": "custom",
                    "percent": "constant",  # For deferred tax
                    "revenue": "constant",   # For capex
                }
                assumption_type = method_map.get(method, method)
                
                # Extract values based on method
                values = []
                if method == "step":
                    step_value = assumption_obj.get("stepValue")
                    if step_value is not None and step_value != "":
                        try:
                            step_val = float(step_value)
                            # Generate step values: [v, v+step, v+2*step, v+3*step, v+4*step]
                            # For step method, each period increases by the step amount
                            # stepValue represents the step amount (increment per period)
                            values = [step_val * (i + 1) for i in range(forecast_periods)]
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert stepValue '{step_value}' to float")
                elif method == "stable":
                    stable_value = assumption_obj.get("stableValue")
                    if stable_value is not None and stable_value != "":
                        try:
                            stable_val = float(stable_value)
                            # Generate constant values: [v, v, v, v, v]
                            values = [stable_val] * forecast_periods
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert stableValue '{stable_value}' to float")
                elif method == "manual":
                    # Check for values array
                    if "values" in assumption_obj:
                        values = assumption_obj["values"]
                    elif "stepValue" in assumption_obj:
                        # Fallback to step if values not provided
                        try:
                            step_val = float(assumption_obj["stepValue"])
                            values = [step_val + i * step_val for i in range(forecast_periods)]
                        except (ValueError, TypeError):
                            pass
                elif method == "percent":
                    # For deferred tax - use percent value as constant
                    percent_value = assumption_obj.get("percent")
                    if percent_value is not None:
                        try:
                            percent_val = float(percent_value)
                            values = [percent_val] * forecast_periods
                        except (ValueError, TypeError):
                            pass
                elif method == "revenue":
                    # For capex - use value as constant
                    value = assumption_obj.get("value")
                    if value is not None:
                        try:
                            val = float(value)
                            values = [val] * forecast_periods
                        except (ValueError, TypeError):
                            pass
                
                return {"type": assumption_type, "values": values}
            elif isinstance(assumption_obj, (int, float, str)):
                # Direct value (like shareRepurchases: "1")
                try:
                    val = float(assumption_obj)
                    return {"type": "constant", "values": [val] * forecast_periods}
                except (ValueError, TypeError):
                    return {"type": "constant", "values": [0.0] * forecast_periods}
            else:
                return {"type": "constant", "values": [0.0] * forecast_periods}
        
        # Convert Income Statement assumptions
        if "incomeStatement" in assumptions:
            legacy["assumptions"] = {}
            is_data = assumptions["incomeStatement"]
            
            # Map field names: camelCase -> snake_case
            field_mapping = {
                "revenue": "revenue",
                "grossMargin": "gross_margin",
                "operatingMargin": "operating_margin",
                "taxRate": "tax_rate",
                "interestRateOnDebt": "interest_rate_debt",
            }
            
            for new_key, legacy_key in field_mapping.items():
                if new_key in is_data:
                    legacy["assumptions"][legacy_key] = convert_assumption_value(is_data[new_key])
        
        # Convert Balance Sheet assumptions
        if "balanceSheet" in assumptions:
            legacy["balance_sheet"] = {}
            bs_data = assumptions["balanceSheet"]
            
            # Map field names
            field_mapping = {
                "depreciation": "depreciation_ppe",
                "inventory": "inventory_sales",
                "totalDebt": "total_debt_amount",
                "debtChange": "lt_debt_change",
                "longTermDebtChange": "lt_debt_change",
            }
            
            for new_key, legacy_key in field_mapping.items():
                if new_key in bs_data:
                    legacy["balance_sheet"][legacy_key] = convert_assumption_value(bs_data[new_key])
        
        # Convert Cash Flow assumptions
        if "cashFlow" in assumptions:
            legacy["cash_flow"] = {}
            cf_data = assumptions["cashFlow"]
            
            # Map field names
            if "shareRepurchases" in cf_data:
                legacy["cash_flow"]["share_repurchases"] = convert_assumption_value(cf_data["shareRepurchases"])
            if "dividendPercentNI" in cf_data:
                legacy["cash_flow"]["dividends_ni"] = convert_assumption_value(cf_data["dividendPercentNI"])
            if "capex" in cf_data:
                legacy["cash_flow"]["capex"] = convert_assumption_value(cf_data["capex"])
            # Optional additional mappings: deferred tax and change in working capital
            if "deferredTax" in cf_data:
                legacy["cash_flow"]["deferred_tax"] = convert_assumption_value(cf_data["deferredTax"])
            if "changeInWC" in cf_data:
                legacy["cash_flow"]["change_in_working_capital"] = convert_assumption_value(cf_data["changeInWC"])

        # Scale percent-based assumptions to decimals (always assumption / 100, e.g., 5 → 0.05)
        def _scale_percent_block(block: Dict[str, Any], keys: List[str]) -> None:
            for key in keys:
                item = block.get(key)
                if isinstance(item, dict) and "values" in item and isinstance(item["values"], list):
                    scaled_values: List[float] = []
                    for v in item["values"]:
                        try:
                            num = float(v)
                            # Always treat as whole percent and convert to decimal
                            num = num / 100.0
                            scaled_values.append(num)
                        except (ValueError, TypeError):
                            scaled_values.append(v)
                    item["values"] = scaled_values

        # Income Statement percent assumptions
        if "assumptions" in legacy:
            _scale_percent_block(
                legacy["assumptions"],
                [
                    "revenue",
                    "gross_margin",
                    "operating_margin",
                    "tax_rate",
                    "interest_rate_debt",
                ],
            )

        # Balance Sheet percent assumptions
        if "balance_sheet" in legacy:
            _scale_percent_block(
                legacy["balance_sheet"],
                [
                    "depreciation_ppe",
                    "inventory_sales",
                    "lt_debt_change",
                ],
            )

        # Cash Flow percent assumptions
        if "cash_flow" in legacy:
            _scale_percent_block(
                legacy["cash_flow"],
                [
                    "dividends_ni",
                ],
            )
        
        # Extract DCF assumptions to top level for populate_wacc_and_assumptions
        if "dcf" in assumptions:
            dcf_data = assumptions["dcf"]
            
            # Extract beta (could be calculated or manual)
            if "beta" in dcf_data:
                beta_obj = dcf_data["beta"]
                if isinstance(beta_obj, dict):
                    if beta_obj.get("method") == "calculate":
                        # Use calculated value if available
                        calculated = beta_obj.get("calculated")
                        if calculated:
                            try:
                                legacy["beta"] = float(calculated)
                            except (ValueError, TypeError):
                                pass
                    else:
                        # Manual beta - might be in a different field
                        manual_beta = beta_obj.get("value") or beta_obj.get("manual")
                        if manual_beta:
                            try:
                                legacy["beta"] = float(manual_beta)
                            except (ValueError, TypeError):
                                pass
            
            # Extract other DCF assumptions
            if "marketRiskPremium" in dcf_data:
                try:
                    mrp_val = dcf_data["marketRiskPremium"]
                    # Handle string or number - if string, strip and convert
                    if isinstance(mrp_val, str):
                        mrp_val = mrp_val.strip()
                    mrp_val = float(mrp_val)
                    # If value is > 1, treat as percentage and convert to decimal (e.g., 4.33 -> 0.0433)
                    # If value is <= 1, assume already decimal (e.g., 0.0433 -> 0.0433)
                    if mrp_val > 1.0:
                        legacy["market_risk_premium"] = mrp_val / 100.0
                    else:
                        legacy["market_risk_premium"] = mrp_val
                    logger.info(f"Extracted market_risk_premium: {dcf_data['marketRiskPremium']} -> {legacy['market_risk_premium']}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Could not convert marketRiskPremium '{dcf_data.get('marketRiskPremium')}': {e}")
            
            if "riskFreeRate" in dcf_data:
                try:
                    rfr_val = dcf_data["riskFreeRate"]
                    # Handle string or number - if string, strip and convert
                    if isinstance(rfr_val, str):
                        rfr_val = rfr_val.strip()
                    rfr_val = float(rfr_val)
                    # If value is > 1, treat as percentage and convert to decimal (e.g., 4.06 -> 0.0406)
                    # If value is <= 1, assume already decimal (e.g., 0.0406 -> 0.0406)
                    if rfr_val > 1.0:
                        legacy["risk_free_rate"] = rfr_val / 100.0
                    else:
                        legacy["risk_free_rate"] = rfr_val
                    logger.info(f"Extracted risk_free_rate: {dcf_data['riskFreeRate']} -> {legacy['risk_free_rate']}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Could not convert riskFreeRate '{dcf_data.get('riskFreeRate')}': {e}")
            
            if "costOfDebt" in dcf_data:
                try:
                    legacy["cost_of_debt"] = float(dcf_data["costOfDebt"])
                except (ValueError, TypeError):
                    pass
            
            if "terminalGrowthRate" in dcf_data:
                try:
                    # Convert percentage to decimal if needed
                    tgr = float(dcf_data["terminalGrowthRate"])
                    legacy["terminal_growth_rate"] = tgr / 100.0 if tgr > 1 else tgr
                except (ValueError, TypeError):
                    pass
            
            # Extract tax_rate from DCF if available (even though DCF cells are formula-driven)
            if "taxRate" in dcf_data:
                try:
                    legacy["tax_rate"] = float(dcf_data["taxRate"])
                except (ValueError, TypeError):
                    pass
        
        # Extract and normalize scenario-specific assumptions (Bear/Bull)
        def _convert_scenario_block(
            src: Dict[str, Any],
            base_cash_flow: Optional[Dict[str, Any]] = None,
        ) -> Dict[str, Any]:
            scenario: Dict[str, Any] = {}
            if not isinstance(src, dict):
                return scenario

            # Map project JSON keys → scenario keys used by the DCF scenario engine
            field_mapping = {
                # Revenue growth (%)
                "revenue": "revenue_growth",
                "revenueGrowth": "revenue_growth",
                # Gross margin (%)
                "grossMargin": "gp_margin",
                # EBIT margin (%)
                "operatingMargin": "ebit_margin",
                # D&A as % of revenue
                "depreciationPercentRevenue": "da_pct_rev",
                "depreciationPercentOfRevenue": "da_pct_rev",
                # CapEx as % of revenue
                "capex": "capex_pct_rev",
                "capexPercentRevenue": "capex_pct_rev",
                # CapEx as % of D&A (if provided)
                "capexPercentDA": "capex_pct_da",
                "capexPctDA": "capex_pct_da",
                # Changes in working capital (% of revenue)
                "changeInWC": "wc_change",
                "changesInWC": "wc_change",
                "changeInWorkingCapital": "wc_change",
            }

            for new_key, legacy_key in field_mapping.items():
                if new_key in src:
                    scenario[legacy_key] = convert_assumption_value(src[new_key])

            # Scenario-specific terminal growth rate
            if "terminalGrowthRate" in src:
                try:
                    tgr = float(src["terminalGrowthRate"])
                    scenario["terminalGrowthRate"] = tgr / 100.0 if tgr > 1 else tgr
                except (ValueError, TypeError):
                    pass

            # Scale percent-type fields in scenario (values are in percent points)
            _scale_percent_block(
                scenario,
                [
                    "revenue_growth",
                    "gp_margin",
                    "ebit_margin",
                    "da_pct_rev",
                    "capex_pct_rev",
                    "capex_pct_da",
                    "wc_change",
                ],
            )

            # If wc_change not provided at scenario level, fall back to base cash_flow
            if "wc_change" not in scenario and base_cash_flow:
                base_wc = base_cash_flow.get("change_in_working_capital")
                if isinstance(base_wc, dict):
                    scenario["wc_change"] = base_wc

            return scenario

        if "bearScenario" in assumptions and isinstance(assumptions["bearScenario"], dict):
            legacy["bearScenario"] = _convert_scenario_block(
                assumptions["bearScenario"],
                legacy.get("cash_flow"),
            )
        
        if "bullScenario" in assumptions and isinstance(assumptions["bullScenario"], dict):
            legacy["bullScenario"] = _convert_scenario_block(
                assumptions["bullScenario"],
                legacy.get("cash_flow"),
            )

        return legacy
    
    # Unknown format, return as-is
    logger.warning("Unknown assumptions format, returning as-is")
    return assumptions


def _transform_array_format_to_standard(data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Transform array format JSON (the final structured output format) to standard format.
    
    This function handles the final structured output format used across the backend:
    Array format: [{date, symbol, revenue, costOfRevenue, operatingIncome, ...}, ...]
    
    Transforms to standard format: 
    {company: {...}, filings: [{statements: {income_statement: {line_items: [...]}}}]}
    
    Field mappings:
    - revenue → IS_REVENUE
    - costOfRevenue → IS_COGS
    - operatingIncome → IS_OPERATING_INCOME
    - interestExpense → IS_INTEREST_EXPENSE
    - incomeTaxExpense → IS_TAX_EXPENSE
    - netIncome → IS_NET_INCOME
    - weightedAverageShsOutDil → IS_DILUTED_SHARES
    - And other standard income statement fields
    
    Args:
        data: List of period dictionaries (final structured output format)
        
    Returns:
        Dictionary in standard format compatible with existing backend code
    """
    if not data:
        return {"company": {}, "filings": []}
    
    # Extract company info from first period
    first_period = data[0]
    symbol = first_period.get("symbol", "").strip()
    
    # Field name to model_role mapping
    FIELD_TO_MODEL_ROLE: Dict[str, Optional[str]] = {
        "revenue": "IS_REVENUE",
        "costOfRevenue": "IS_COGS",
        "grossProfit": None,  # Computed, skip
        "researchAndDevelopmentExpenses": "IS_OPERATING_EXPENSE",
        "generalAndAdministrativeExpenses": "IS_OPERATING_EXPENSE",
        "sellingAndMarketingExpenses": "IS_OPERATING_EXPENSE",
        "sellingGeneralAndAdministrativeExpenses": "IS_OPERATING_EXPENSE",
        "otherExpenses": "IS_OPERATING_EXPENSE",
        "operatingExpenses": "IS_OPERATING_EXPENSE",
        "costAndExpenses": None,  # Computed, skip
        "netInterestIncome": None,  # Computed, skip
        "interestIncome": None,  # Skip for now
        "interestExpense": "IS_INTEREST_EXPENSE",
        "depreciationAndAmortization": "CF_DEPRECIATION",  # Cash flow item, but may appear in IS
        "ebitda": None,  # Computed, skip
        "ebit": None,  # Computed, skip
        "nonOperatingIncomeExcludingInterest": None,  # Skip for now
        "operatingIncome": "IS_OPERATING_INCOME",
        "totalOtherIncomeExpensesNet": None,  # Computed, skip
        "incomeBeforeTax": None,  # Computed, skip
        "incomeTaxExpense": "IS_TAX_EXPENSE",
        "netIncomeFromContinuingOperations": None,  # Skip, use netIncome
        "netIncomeFromDiscontinuedOperations": None,  # Skip
        "otherAdjustmentsToNetIncome": None,  # Skip
        "netIncome": "IS_NET_INCOME",
        "netIncomeDeductions": None,  # Skip
        "bottomLineNetIncome": None,  # Skip, use netIncome
        "eps": None,  # Computed, skip
        "epsDiluted": None,  # Computed, skip
        "weightedAverageShsOut": None,  # Skip
        "weightedAverageShsOutDil": "IS_DILUTED_SHARES",  # For EPS calculation
    }
    
    # Field name to label mapping
    FIELD_TO_LABEL: Dict[str, str] = {
        "revenue": "Revenue",
        "costOfRevenue": "Cost of Revenue",
        "researchAndDevelopmentExpenses": "Research and Development Expenses",
        "generalAndAdministrativeExpenses": "General and Administrative Expenses",
        "sellingAndMarketingExpenses": "Selling and Marketing Expenses",
        "sellingGeneralAndAdministrativeExpenses": "Selling, General and Administrative Expenses",
        "otherExpenses": "Other Expenses",
        "operatingExpenses": "Operating Expenses",
        "interestExpense": "Interest Expense",
        "depreciationAndAmortization": "Depreciation and Amortization",
        "nonOperatingIncomeExcludingInterest": "Other Non Operating Adjustments",
        "operatingIncome": "Operating Income",
        "incomeTaxExpense": "Income Tax Expense",
        "netIncome": "Net Income",
        "weightedAverageShsOutDil": "Diluted Weighted Average Shares Outstanding",
    }
    
    # Build line items
    line_items: List[Dict[str, Any]] = []
    line_items_by_field: Dict[str, Dict[str, Any]] = {}
    
    for period in data:
        date_str = period.get("date", "")
        if not date_str:
            continue
        
        # Process each field
        for field_name, value in period.items():
            # Skip metadata fields
            if field_name in ["date", "symbol", "reportedCurrency", "cik", "filingDate", 
                             "acceptedDate", "fiscalYear", "period"]:
                continue
            
            # Skip null values (but include zero values as they're meaningful in financial statements)
            if value is None:
                continue
            
            # Get model_role
            model_role = FIELD_TO_MODEL_ROLE.get(field_name)
            if not model_role:
                continue
            
            # Get or create line item
            if field_name not in line_items_by_field:
                # Convert camelCase to Title Case
                import re
                label = FIELD_TO_LABEL.get(field_name, re.sub(r'([A-Z])', r' \1', field_name).strip().title())
                line_item = {
                    "tag": field_name,
                    "label": label,
                    "model_role": model_role,
                    "unit": "USD",
                    "periods": {}
                }
                line_items_by_field[field_name] = line_item
                line_items.append(line_item)
            
            # Add period value (include zero values as they're meaningful in financial statements)
            line_items_by_field[field_name]["periods"][date_str] = float(value)
    
    # Build standard format structure
    result = {
        "company": {
            "ticker": symbol,
            "company_name": symbol  # Use ticker as fallback, can be updated later
        },
        "filings": [
            {
                "filing_id": f"{symbol}_multi_year",
                "filing_type": "10-K",
                "fiscal_year": data[0].get("fiscalYear", ""),
                "fiscal_period": "FY",
                "statements": {
                    "income_statement": {
                        "statement_type": "income_statement",
                        "normalized_order": 1,
                        "line_items": line_items
                    }
                }
            }
        ]
    }
    
    logger.info(f"Transformed array format JSON: {len(data)} periods -> {len(line_items)} line items")
    return result


def _convert_fmp_array_to_line_items(
    data: List[Dict[str, Any]],
    statement_type: str
) -> List[Dict[str, Any]]:
    """
    Convert FMP array format to line items with model_role.
    
    Args:
        data: List of period dictionaries from FMP (e.g., income_statements, balance_sheets, cash_flow_statements)
        statement_type: One of "income_statement", "balance_sheet", "cash_flow_statement"
        
    Returns:
        List of line item dictionaries with model_role, label, and periods
    """
    import re
    
    # Field name to model_role mapping for each statement type
    FIELD_TO_MODEL_ROLE: Dict[str, Dict[str, Optional[str]]] = {
        "income_statement": {
            "revenue": "IS_REVENUE",
            "costOfRevenue": "IS_COGS",
            "grossProfit": "IS_GROSS_PROFIT",
            "researchAndDevelopmentExpenses": "IS_OPERATING_EXPENSE",
            "generalAndAdministrativeExpenses": "IS_OPERATING_EXPENSE",
            "sellingAndMarketingExpenses": "IS_OPERATING_EXPENSE",
            "sellingGeneralAndAdministrativeExpenses": "IS_OPERATING_EXPENSE",
            "otherExpenses": "IS_OPERATING_EXPENSE",
            "operatingExpenses": "IS_OPERATING_EXPENSE",
            "costAndExpenses": None,  # Computed, skip
            "netInterestIncome": "IS_NON_OPERATING_INCOME",  # Using IS_NON_OPERATING_INCOME for net interest income
            "interestIncome": None,  # Skip for now
            "interestExpense": "IS_INTEREST_EXPENSE",
            "depreciationAndAmortization": "CF_DEPRECIATION",  # Cash flow item, but may appear in IS
            "ebitda": None,  # Computed, skip
            "ebit": None,  # Computed, skip
            "nonOperatingIncomeExcludingInterest": "IS_NON_OPERATING_INCOME",  # Other Non Operating Adjustments
            "operatingIncome": "IS_OPERATING_INCOME",  # Maps to "Operating Profit" in Excel
            "totalOtherIncomeExpensesNet": None,  # Computed, skip
            "incomeBeforeTax": "IS_OPERATING_INCOME",  # Using IS_OPERATING_INCOME as closest match
            "incomeTaxExpense": "IS_TAX_EXPENSE",
            "netIncomeFromContinuingOperations": None,  # Skip, use netIncome
            "netIncomeFromDiscontinuedOperations": None,  # Skip
            "otherAdjustmentsToNetIncome": None,  # Skip
            "netIncome": "IS_NET_INCOME",
            "netIncomeDeductions": None,  # Skip
            "bottomLineNetIncome": None,  # Skip, use netIncome
            "eps": "IS_NET_INCOME",  # EPS - using IS_NET_INCOME as placeholder
            "epsDiluted": "IS_NET_INCOME",  # EPS Diluted - using IS_NET_INCOME as placeholder
            "weightedAverageShsOut": None,  # Skip
            "weightedAverageShsOutDil": "IS_DILUTED_SHARES",  # For EPS calculation
        },
        "balance_sheet": {
            "cashAndCashEquivalents": "BS_CASH",
            "shortTermInvestments": "BS_MARKETABLE_SECURITIES",
            "cashAndShortTermInvestments": None,  # Computed, skip
            "netReceivables": "BS_ACCOUNTS_RECEIVABLE",
            "accountsReceivables": "BS_ACCOUNTS_RECEIVABLE",
            "otherReceivables": None,  # Skip for now
            "inventory": "BS_INVENTORY",
            "prepaids": "BS_CASH",  # Prepaid Expenses - using BS_CASH as placeholder
            "otherCurrentAssets": "BS_ASSETS_CURRENT",  # Other Current Assets - using BS_ASSETS_CURRENT as placeholder
            "totalCurrentAssets": "BS_ASSETS_CURRENT",
            "propertyPlantEquipmentNet": "BS_PP_AND_E",
            "goodwill": "BS_GOODWILL",
            "intangibleAssets": "BS_INTANGIBLES",
            "goodwillAndIntangibleAssets": None,  # Computed, skip
            "longTermInvestments": "BS_MARKETABLE_SECURITIES",  # Long-Term Investments - using BS_MARKETABLE_SECURITIES as placeholder
            "taxAssets": "BS_ASSETS_NONCURRENT",  # Tax Assets - using BS_ASSETS_NONCURRENT as placeholder
            "otherNonCurrentAssets": "BS_ASSETS_NONCURRENT",  # Other Non-Current Assets
            "totalNonCurrentAssets": None,  # Skipped - row removed from template
            "otherAssets": None,  # Skipped - row removed from template
            "totalAssets": "BS_ASSETS_TOTAL",
            "accountsPayable": "BS_ACCOUNTS_PAYABLE",
            "accountPayables": "BS_ACCOUNTS_PAYABLE",  # Alternative field name
            "otherPayables": "BS_ACCOUNTS_PAYABLE",  # Other Payables - using BS_ACCOUNTS_PAYABLE as placeholder
            "accruedExpenses": "BS_ACCRUED_LIABILITIES",
            "shortTermDebt": "BS_DEBT_CURRENT",
            "capitalLeaseObligationsCurrent": "BS_DEBT_CURRENT",  # Current Capital Lease Obligations
            "currentCapitalLeaseObligations": "BS_DEBT_CURRENT",  # Alternative field name
            "taxPayables": "BS_LIABILITIES_CURRENT",  # Tax Payables - using BS_LIABILITIES_CURRENT as placeholder
            "deferredRevenue": "BS_LIABILITIES_CURRENT",  # Deferred Revenue (Current)
            "otherCurrentLiabilities": "BS_LIABILITIES_CURRENT",  # Other Current Liabilities
            "totalCurrentLiabilities": "BS_LIABILITIES_CURRENT",
            "longTermDebt": "BS_DEBT_NONCURRENT",
            "capitalLeaseObligationsNonCurrent": "BS_DEBT_NONCURRENT",  # Non-Current Capital Lease Obligations
            "nonCurrentCapitalLeaseObligations": "BS_DEBT_NONCURRENT",  # Alternative field name
            "deferredRevenueNonCurrent": "BS_LIABILITIES_NONCURRENT",  # Deferred Revenue (Non-Current)
            "deferredTaxLiabilitiesNonCurrent": "BS_LIABILITIES_NONCURRENT",  # Deferred Tax Liabilities (Non-Current)
            "otherNonCurrentLiabilities": "BS_LIABILITIES_NONCURRENT",  # Other Non-Current Liabilities
            "totalNonCurrentLiabilities": "BS_LIABILITIES_NONCURRENT",
            "totalLiabilities": "BS_LIABILITIES_TOTAL",
            "preferredStock": "BS_EQUITY_TOTAL",  # Preferred Stock - using BS_EQUITY_TOTAL as placeholder
            "treasuryStock": "BS_EQUITY_TOTAL",  # Treasury Stock - using BS_EQUITY_TOTAL as placeholder
            "commonStock": "BS_COMMON_STOCK",
            "retainedEarnings": "BS_RETAINED_EARNINGS",
            "additionalPaidInCapital": "BS_EQUITY_TOTAL",  # Additional Paid-In Capital - using BS_EQUITY_TOTAL as placeholder
            "accumulatedOtherComprehensiveIncomeLoss": "BS_EQUITY_TOTAL",  # Accumulated Other Comprehensive Income (Loss)
            "otherStockholdersEquity": "BS_EQUITY_TOTAL",  # Other Stockholders' Equity
            "otherTotalStockholdersEquity": "BS_EQUITY_TOTAL",  # Alternative field name for Other Stockholders' Equity
            "totalEquity": "BS_EQUITY_TOTAL",
            "totalStockholdersEquity": "BS_EQUITY_TOTAL",  # Alternative field name for Total Equity
            "totalLiabilitiesAndStockholdersEquity": None,  # Computed, skip
        },
        "cash_flow_statement": {
            "netIncome": "IS_NET_INCOME",  # Also appears in CF
            "depreciationAndAmortization": "CF_DEPRECIATION",
            "deferredIncomeTax": "CF_DEPRECIATION",  # Deferred Income Tax - using CF_DEPRECIATION as placeholder
            "stockBasedCompensation": "CF_DEPRECIATION",  # Stock Based Compensation - using CF_DEPRECIATION as placeholder
            "changeInWorkingCapital": "CF_CHANGE_IN_WORKING_CAPITAL",
            "accountsReceivables": None,  # Component of working capital, skip
            "inventory": None,  # Component of working capital, skip
            "accountsPayables": None,  # Component of working capital, skip
            "otherWorkingCapital": None,  # Component of working capital, skip
            "otherNonCashItems": None,  # Skip - not in template
            "netCashProvidedByOperatingActivities": "CF_CASH_FROM_OPERATIONS",
            "investmentsInPropertyPlantAndEquipment": "CF_CAPEX",
            "acquisitionsNet": None,  # Deleted from template
            "purchasesOfInvestments": None,  # Deleted from template
            "salesMaturitiesOfInvestments": None,  # Skip - not in template
            "otherInvestingActivities": None,  # Deleted from template
            "netCashProvidedByInvestingActivities": None,  # Deleted from template
            "netDebtIssuance": None,  # Computed, skip
            "longTermNetDebtIssuance": None,  # Deleted from template
            "shortTermNetDebtIssuance": None,  # Skip - not in template
            "netStockIssuance": "CF_SHARE_REPURCHASES",  # Net Stock Issuance - using CF_SHARE_REPURCHASES as placeholder
            "netCommonStockIssuance": "CF_SHARE_REPURCHASES",  # Net Common Stock Issuance
            "commonStockIssuance": "CF_SHARE_REPURCHASES",  # Common Stock Issued
            "commonStockRepurchased": "CF_SHARE_REPURCHASES",
            "netPreferredStockIssuance": "CF_SHARE_REPURCHASES",  # Net Preferred Stock Issuance
            "dividendsPaid": "CF_DIVIDENDS_PAID",
            "commonDividendsPaid": "CF_DIVIDENDS_PAID",
            "preferredDividendsPaid": None,  # Skip for now
            "otherFinancingActivities": "CF_CASH_FROM_FINANCING",  # Other Financing Activities
            "netCashProvidedByFinancingActivities": "CF_CASH_FROM_FINANCING",
            "effectOfForexOnCash": None,  # Skip for now
            "effectOfForexChangesOnCash": None,  # Skip for now
            "cashAtBeginningOfPeriod": None,  # Skip for now
            "netChangeInCash": "CF_NET_CHANGE_IN_CASH",
            "cashAtEndOfPeriod": None,  # Skip for now
            "freeCashFlow": None,  # Computed, skip
        },
    }
    
    # Field name to label mapping
    FIELD_TO_LABEL: Dict[str, Dict[str, str]] = {
        "income_statement": {
            "revenue": "Revenue",
            "costOfRevenue": "Cost of Revenue",
            "grossProfit": "Gross Profit",
            "researchAndDevelopmentExpenses": "Research & Development Expenses",
            "generalAndAdministrativeExpenses": "General and Administrative Expenses",
            "sellingAndMarketingExpenses": "Selling and Marketing Expenses",
            "sellingGeneralAndAdministrativeExpenses": "Selling, General & Administrative Expenses (SG&A)",
            "otherExpenses": "Other Expenses",
            "operatingExpenses": "Operating Expenses",
            "netInterestIncome": "Net Interest Income",
            "interestExpense": "Interest Expense",
            "depreciationAndAmortization": "Depreciation and Amortization",
            "operatingIncome": "Operating Profit",
            "incomeBeforeTax": "Income Before Taxes",
            "incomeTaxExpense": "Income Tax Expense",
            "netIncome": "Net Income",
            "eps": "EPS",
            "epsDiluted": "EPS",
            "weightedAverageShsOutDil": "Weighted Average Shares Outstanding",
        },
        "balance_sheet": {
            "cashAndCashEquivalents": "Cash & Cash Equivalents",
            "shortTermInvestments": "Short-Term Investments",
            "netReceivables": "Net Receivables",
            "accountsReceivables": "Accounts Receivables",
            "inventory": "Inventory",
            "prepaids": "Prepaid Expenses",
            "otherCurrentAssets": "Other Current Assets",
            "totalCurrentAssets": "Total Current Assets",
            "propertyPlantEquipmentNet": "Property, Plant & Equipment (Net)",
            "goodwill": "Goodwill",
            "intangibleAssets": "Intangible Assets",
            "longTermInvestments": "Long-Term Investments",
            "taxAssets": "Tax Assets",
            "otherNonCurrentAssets": "Other Non-Current Assets",
            "totalNonCurrentAssets": "Total Non-Current Assets",
            "otherAssets": "Other Assets",
            "totalAssets": "Total Assets",
            "accountsPayable": "Accounts Payable",
            "accountPayables": "Accounts Payable",
            "otherPayables": "Other Payables",
            "accruedExpenses": "Accrued Expenses",
            "shortTermDebt": "Short-Term Debt",
            "capitalLeaseObligationsCurrent": "Current Capital Lease Obligations",
            "currentCapitalLeaseObligations": "Current Capital Lease Obligations",
            "taxPayables": "Tax Payables",
            "deferredRevenue": "Deferred Revenue (Current)",
            "otherCurrentLiabilities": "Other Current Liabilities",
            "totalCurrentLiabilities": "Total Current Liabilities",
            "longTermDebt": "Long-Term Debt",
            "capitalLeaseObligationsNonCurrent": "Non-Current Capital Lease Obligations",
            "nonCurrentCapitalLeaseObligations": "Non-Current Capital Lease Obligations",
            "deferredRevenueNonCurrent": "Deferred Revenue (Non-Current)",
            "deferredTaxLiabilitiesNonCurrent": "Deferred Tax Liabilities (Non-Current)",
            "otherNonCurrentLiabilities": "Other Non-Current Liabilities",
            "totalNonCurrentLiabilities": "Total Non-Current Liabilities",
            "totalLiabilities": "Total Liabilities",
            "preferredStock": "Preferred Stock",
            "treasuryStock": "Treasury Stock",
            "commonStock": "Common Stock",
            "retainedEarnings": "Retained Earnings",
            "additionalPaidInCapital": "Additional Paid-In Capital",
            "accumulatedOtherComprehensiveIncomeLoss": "Accumulated Other Comprehensive Income (Loss)",
            "otherStockholdersEquity": "Other Stockholders' Equity",
            "otherTotalStockholdersEquity": "Other Stockholders' Equity",
            "totalEquity": "Total Equity",
            "totalStockholdersEquity": "Total Equity",
        },
        "cash_flow_statement": {
            "netIncome": "Net Income",
            "depreciationAndAmortization": "Depreciation & Amortization",
            "deferredIncomeTax": "Deferred Income Tax",
            "stockBasedCompensation": "Stock-Based Compensation",
            "changeInWorkingCapital": "Change in Working Capital",
            "netCashProvidedByOperatingActivities": "Net Cash Provided by Operating Activities",
            "investmentsInPropertyPlantAndEquipment": "Investments in Property, Plant & Equipment (CapEx)",
            "acquisitionsNet": "Net Acquisitions",
            "purchasesOfInvestments": "Purchases of Investments",
            "otherInvestingActivities": "Other Investing Activities",
            "netCashProvidedByInvestingActivities": "Net Cash Provided by Investing Activities",
            "longTermNetDebtIssuance": "Long-Term Net Debt Issuance",
            "netStockIssuance": "Net Stock Issuance",
            "netCommonStockIssuance": "Net Common Stock Issuance",
            "commonStockIssuance": "Common Stock Issued",
            "commonStockRepurchased": "Common Stock Repurchased",
            "netPreferredStockIssuance": "Net Preferred Stock Issuance",
            "dividendsPaid": "Dividends Paid",
            "commonDividendsPaid": "Common Dividends Paid",
            "otherFinancingActivities": "Other Financing Activities",
            "netCashProvidedByFinancingActivities": "Net Cash Provided by Financing Activities",
            "effectOfForexChangesOnCash": "Effect of Foreign Exchange on Cash",
            "cashAtBeginningOfPeriod": "Cash at Beginning of Period",
            "netChangeInCash": "Net Change in Cash",
            "cashAtEndOfPeriod": "Cash at End of Period",
            "freeCashFlow": "Free Cash Flow",
        },
    }
    
    field_mapping = FIELD_TO_MODEL_ROLE.get(statement_type, {})
    label_mapping = FIELD_TO_LABEL.get(statement_type, {})
    
    # Build line items - preserve order by processing fields in JSON order
    # For income statement, define the expected order based on JSON structure
    if statement_type == "income_statement":
        # Define the order fields should appear in (matching JSON order)
        field_order = [
            "revenue", "costOfRevenue", "grossProfit",
            "researchAndDevelopmentExpenses", "generalAndAdministrativeExpenses",
            "sellingAndMarketingExpenses", "sellingGeneralAndAdministrativeExpenses",
            "otherExpenses", "operatingExpenses", "costAndExpenses",
            "netInterestIncome", "interestIncome", "interestExpense",
            "depreciationAndAmortization", "ebitda", "ebit",
            "nonOperatingIncomeExcludingInterest", "operatingIncome",
            "totalOtherIncomeExpensesNet", "incomeBeforeTax", "incomeTaxExpense",
            "netIncomeFromContinuingOperations", "netIncomeFromDiscontinuedOperations",
            "otherAdjustmentsToNetIncome", "netIncome", "netIncomeDeductions",
            "bottomLineNetIncome", "eps", "epsDiluted",
            "weightedAverageShsOut", "weightedAverageShsOutDil"
        ]
    else:
        field_order = None  # For other statement types, use dict order
    
    line_items: List[Dict[str, Any]] = []
    line_items_by_field: Dict[str, Dict[str, Any]] = {}
    
    for period in data:
        date_str = period.get("date", "")
        if not date_str:
            continue
        
        # Process fields in order if specified, otherwise use dict order
        if field_order:
            # Include fields that exist in period, even if value is 0 (zero values are meaningful)
            fields_to_process = []
            for field_name in field_order:
                if field_name in period:
                    value = period.get(field_name)
                    # Include the field if value is not None (including 0)
                    if value is not None:
                        fields_to_process.append((field_name, value))
        else:
            fields_to_process = period.items()
        
        # Process each field
        for field_name, value in fields_to_process:
            # Skip metadata fields
            if field_name in ["date", "symbol", "reportedCurrency", "cik", "filingDate", 
                             "acceptedDate", "fiscalYear", "period", "companyName"]:
                continue
            
            # Skip null values (but include zero values as they're meaningful in financial statements)
            if value is None:
                continue
            
            # Get model_role
            model_role = field_mapping.get(field_name)
            if not model_role:
                continue
            
            # Get or create line item
            if field_name not in line_items_by_field:
                # Convert camelCase to Title Case if no label mapping
                label = label_mapping.get(field_name)
                if not label:
                    label = re.sub(r'([A-Z])', r' \1', field_name).strip().title()
                
                line_item = {
                    "tag": field_name,
                    "label": label,
                    "model_role": model_role,
                    "unit": "USD",
                    "periods": {}
                }
                line_items_by_field[field_name] = line_item
                line_items.append(line_item)
            
            # Add period value (include zero values as they're meaningful in financial statements)
            line_items_by_field[field_name]["periods"][date_str] = float(value)
    
    return line_items


def extract_line_items_from_json(json_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Extract all line_items from all filings and all statements.
    
    Handles multiple JSON formats:
    - FMP format: {"income_statements": [...], "balance_sheets": [...], "cash_flow_statements": [...]}
    - Multi-year: {"filings": [{"statements": {...}}]}
    - Single-year: {"statements": {...}}
    
    Args:
        json_data: Loaded JSON data
        
    Returns:
        Flattened list of all line items from all statements
    """
    line_items: List[Dict[str, Any]] = []
    
    # Check if it's FMP format (has income_statements, balance_sheets, or cash_flow_statements arrays)
    if "income_statements" in json_data or "balance_sheets" in json_data or "cash_flow_statements" in json_data:
        # Convert FMP arrays to line items
        if "income_statements" in json_data:
            income_data = json_data.get("income_statements", [])
            if isinstance(income_data, list):
                items = _convert_fmp_array_to_line_items(income_data, "income_statement")
                line_items.extend(items)
                logger.info(f"Converted {len(income_data)} income statement periods to {len(items)} line items")
        
        if "balance_sheets" in json_data:
            balance_data = json_data.get("balance_sheets", [])
            if isinstance(balance_data, list):
                items = _convert_fmp_array_to_line_items(balance_data, "balance_sheet")
                line_items.extend(items)
                logger.info(f"Converted {len(balance_data)} balance sheet periods to {len(items)} line items")
        
        if "cash_flow_statements" in json_data:
            cash_flow_data = json_data.get("cash_flow_statements", [])
            if isinstance(cash_flow_data, list):
                items = _convert_fmp_array_to_line_items(cash_flow_data, "cash_flow_statement")
                line_items.extend(items)
                logger.info(f"Converted {len(cash_flow_data)} cash flow periods to {len(items)} line items")
        
        return line_items
    
    # Check if it's multi-year format (has "filings" key)
    if "filings" in json_data:
        filings = json_data.get("filings", [])
        for filing in filings:
            statements = filing.get("statements", {})
            for stmt_type, stmt_data in statements.items():
                if isinstance(stmt_data, dict):
                    items = stmt_data.get("line_items", [])
                    line_items.extend(items)
    # Check if it's single-year format (has "statements" key at root)
    elif "statements" in json_data:
        statements = json_data.get("statements", {})
        for stmt_type, stmt_data in statements.items():
            if isinstance(stmt_data, dict):
                items = stmt_data.get("line_items", [])
                line_items.extend(items)
    
    return line_items


def build_role_year_matrix(line_items: List[Dict[str, Any]]) -> Dict[str, Dict[int, float]]:
    """
    Build a matrix aggregating values by model_role and year.
    
    Multiple items with the same model_role for a given year are summed.
    
    Args:
        line_items: List of line item dictionaries from JSON
        
    Returns:
        Dictionary mapping model_role -> year -> aggregated_value
        Example: {"IS_REVENUE": {2022: 100.0, 2023: 110.0}}
    """
    matrix: Dict[str, Dict[int, float]] = defaultdict(lambda: defaultdict(float))
    
    for item in line_items:
        role = (item.get("model_role") or "").strip()
        if not role:
            continue
        
        periods = item.get("periods") or {}
        for period_str, raw_val in periods.items():
            if period_str is None or raw_val is None:
                continue
            
            # Extract year from YYYY-MM-DD format
            try:
                year = int(str(period_str).split("-")[0])
                val = float(raw_val)
                matrix[role][year] += val
            except (ValueError, IndexError, TypeError) as e:
                logger.warning(f"Failed to parse period '{period_str}' or value '{raw_val}': {e}")
                continue
    
    return dict(matrix)


def load_excel_template(template_path: str) -> Workbook:
    """
    Load existing Excel template file.
    
    Args:
        template_path: Path to .xlsx template file
        
    Returns:
        openpyxl Workbook object
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for reading Excel templates. Install with: pip install openpyxl")
    
    return load_workbook(template_path)


def find_revenue_row(worksheet) -> Optional[int]:
    """
    Find the row containing "Revenue" label in column B.
    Searches for various revenue-related terms.
    
    Args:
        worksheet: openpyxl Worksheet object
        
    Returns:
        Row number (1-indexed) if found, None otherwise
    """
    # Search column B (index 2) for revenue label
    revenue_keywords = ["revenue", "sales", "net sales", "total revenue", "total sales"]
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_col=2, max_col=2, values_only=False), start=1):
        cell = row[0]
        if cell.value:
            cell_value = str(cell.value).strip().lower()
            # Check for any revenue keyword
            for keyword in revenue_keywords:
                if keyword in cell_value:
                    return row_idx
    return None


def detect_historical_columns(worksheet) -> List[int]:
    """
    Detect historical columns by searching for "Year" headers in row 16.
    Looks for patterns like "Year 'A'", "Year 'E'", "Year A", etc.
    These headers will be overwritten with actual years later.
    
    Args:
        worksheet: openpyxl Worksheet object
        
    Returns:
        List of column indices (1-indexed) containing "Year" headers, sorted left to right
    """
    hist_columns: List[int] = []
    header_row = 16
    
    # Search row 16 for "Year" headers (case-insensitive)
    # Look in columns D through K (4-11) for historical columns
    for col_idx in range(4, 12):  # Columns D through K
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell_value = str(cell.value or "").strip()
        cell_lower = cell_value.lower()
        
        # Check if cell contains "year" - be flexible with patterns
        # Matches: "Year 'A'", "Year 'E'", "Year A", "Year E", "Year\"A\"", etc.
        if "year" in cell_lower:
            # Check for various patterns: Year 'A', Year 'E', Year A, Year E, Year"A", etc.
            if any(pattern in cell_lower for pattern in ["'a'", "'e'", " a", " e", "\"a\"", "\"e\""]) or len(cell_lower) < 10:
                hist_columns.append(col_idx)
    
    # If no "Year" headers found, fallback to hardcoded D, E, F
    if not hist_columns:
        logger.warning("No 'Year' headers found in row 16, using default columns D, E, F")
        return [4, 5, 6]
    
    # Sort columns left to right
    hist_columns.sort()
    logger.debug(f"Detected historical columns: {hist_columns}")
    return hist_columns


def map_years_to_columns(years: List[int], hist_columns: List[int]) -> Dict[int, int]:
    """
    Map sorted years to historical columns from left to right.
    Hardcapped at 3 most recent years.
    
    Args:
        years: List of years from JSON data
        hist_columns: List of column indices containing "Year"A" headers (sorted left to right)
        
    Returns:
        Dictionary mapping year -> column_index
    """
    sorted_years = sorted(years)
    
    # Hardcap at 3 most recent years
    recent_years = sorted_years[-3:] if len(sorted_years) > 3 else sorted_years
    
    year_column_map: Dict[int, int] = {}
    
    # Map oldest year to leftmost column, newest to rightmost
    # Only use up to 3 columns (hardcap)
    max_columns = min(3, len(hist_columns))
    for idx, year in enumerate(recent_years):
        if idx < max_columns:
            year_column_map[year] = hist_columns[idx]
    
    return year_column_map


def write_year_headers(worksheet, year_column_map: Dict[int, int], header_row: int = 16) -> None:
    """
    Write actual year values to header row (row 16 by default).
    
    Args:
        worksheet: openpyxl Worksheet object
        year_column_map: Dictionary mapping year -> column_index
        header_row: Row number where years should be written (default 16)
    """
    # Sort years to write them in order (oldest to newest)
    sorted_years = sorted(year_column_map.keys())
    
    for year in sorted_years:
        column_idx = year_column_map[year]
        cell = worksheet.cell(row=header_row, column=column_idx)
        cell.value = year


def populate_dcf_year_headers(
    workbook,
    historical_years: List[int],
    forecast_periods: int = 5
) -> None:
    """
    Populate year headers for all three DCF sheets (DCF Base, DCF Bear, DCF Bull).
    
    Historical years (3 most recent) are written to columns D-F at row 16.
    Forecasted years (5 years starting from last historical year + 1) are written to:
    - Columns G-K at row 16
    - Columns G-K at row 33
    
    Args:
        workbook: openpyxl Workbook object
        historical_years: List of historical years from JSON data
        forecast_periods: Number of forecast years to generate (default 5)
    """
    if not historical_years:
        logger.error("No historical years provided, cannot populate DCF year headers")
        return
    
    # Extract 3 most recent historical years (sorted, take last 3)
    sorted_historical = sorted(historical_years)
    recent_years = sorted_historical[-3:] if len(sorted_historical) >= 3 else sorted_historical
    logger.info(f"Using {len(recent_years)} historical years: {recent_years}")
    
    # Calculate forecasted years: [last_year + 1, last_year + 2, ..., last_year + forecast_periods]
    last_historical_year = sorted_historical[-1]
    forecasted_years = [last_historical_year + i for i in range(1, forecast_periods + 1)]
    logger.info(f"Calculated {len(forecasted_years)} forecasted years: {forecasted_years}")
    
    # DCF sheet names
    dcf_sheets = ["DCF Base", "DCF Bear", "DCF Bull"]
    
    # Historical year columns: D(4), E(5), F(6)
    historical_columns = [4, 5, 6]
    historical_row = 16
    
    # Forecasted year columns: G(7), H(8), I(9), J(10), K(11)
    forecasted_columns = [7, 8, 9, 10, 11]
    forecasted_rows = [16, 33]
    
    # Process each DCF sheet
    for sheet_name in dcf_sheets:
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"DCF sheet '{sheet_name}' not found, skipping")
            continue
        
        worksheet = workbook[sheet_name]
        logger.info(f"Populating year headers for sheet '{sheet_name}'")
        
        # Write historical years to columns D-F at row 16 (with "A" suffix)
        for idx, year in enumerate(recent_years):
            if idx < len(historical_columns):
                col_idx = historical_columns[idx]
                cell = worksheet.cell(row=historical_row, column=col_idx)
                # Check if cell contains "Year"A"" placeholder (case-insensitive)
                cell_value = str(cell.value or "").strip()
                if "year" in cell_value.lower() and '"a"' in cell_value.lower():
                    logger.debug(f"Replacing '{cell_value}' with {year} A at {sheet_name} row {historical_row}, column {col_idx}")
                cell.value = f"{year} A"
        
        # Write forecasted years to columns G-K at rows 16 and 33 (with "E" suffix)
        for row_idx in forecasted_rows:
            for idx, year in enumerate(forecasted_years):
                if idx < len(forecasted_columns):
                    col_idx = forecasted_columns[idx]
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    # Check if cell contains "Year"E"" placeholder (case-insensitive)
                    cell_value = str(cell.value or "").strip()
                    if "year" in cell_value.lower() and '"e"' in cell_value.lower():
                        logger.debug(f"Replacing '{cell_value}' with {year} E at {sheet_name} row {row_idx}, column {col_idx}")
                    cell.value = f"{year} E"
        
        logger.info(f"Populated year headers for '{sheet_name}': Historical={recent_years}, Forecasted={forecasted_years}")


def populate_three_statement_year_headers(
    worksheet,
    historical_years: List[int],
    forecast_periods: int = 5,
    income_statement_end_row: Optional[int] = None
) -> None:
    """
    Populate year headers for the 3-Statement model sheet ("3 Statement").
    
    Historical years (3 most recent) are written to columns D-F.
    Forecasted years (5 years starting from last historical year + 1) are written to columns G-K.
    
    Year header locations (fixed rows):
    - Income Statement: Row 4
    - Balance Sheet: Row 24
    - Cash Flow: Row 69
    
    Args:
        worksheet: openpyxl Worksheet object for "3 Statement" sheet
        historical_years: List of historical years from JSON data
        forecast_periods: Number of forecast years to generate (default 5)
        income_statement_end_row: Optional (kept for backwards compatibility, not used)
    """
    if not historical_years:
        logger.error("No historical years provided, cannot populate 3-statement year headers")
        return
    
    # Extract 3 most recent historical years (sorted, take last 3)
    sorted_historical = sorted(historical_years)
    recent_years = sorted_historical[-3:] if len(sorted_historical) >= 3 else sorted_historical
    logger.info(f"Using {len(recent_years)} historical years for 3-statement: {recent_years}")
    
    # Calculate forecasted years: [last_year + 1, last_year + 2, ..., last_year + forecast_periods]
    last_historical_year = sorted_historical[-1]
    forecasted_years = [last_historical_year + i for i in range(1, forecast_periods + 1)]
    logger.info(f"Calculated {len(forecasted_years)} forecasted years for 3-statement: {forecasted_years}")
    
    # Historical year columns: D(4), E(5), F(6)
    historical_columns = [4, 5, 6]
    # Forecasted year columns: G(7), H(8), I(9), J(10), K(11)
    forecasted_columns = [7, 8, 9, 10, 11]
    
    # Helper function to write years to a specific row
    def write_years_to_row(row_num: int, statement_name: str) -> None:
        """Write historical and forecasted years to a specific row."""
        logger.info(f"Populating {statement_name} year headers at row {row_num}")
        
        # First, scan the row for any "Year A" or "Year E" placeholders and replace them
        # This ensures we catch placeholders even if they're in unexpected columns
        for col_idx in range(4, 12):  # Columns D through K
            cell = worksheet.cell(row=row_num, column=col_idx)
            cell_value = str(cell.value or "").strip().lower()
            
            # Check for "Year A" or "Year"A"" patterns
            if "year" in cell_value and (" a" in cell_value or '"a"' in cell_value or "'a'" in cell_value):
                # Determine which historical year this should be based on column position
                if col_idx in historical_columns:
                    year_idx = historical_columns.index(col_idx)
                    if year_idx < len(recent_years):
                        year = recent_years[year_idx]
                        cell.value = f"{year} A"
                        logger.debug(f"Replaced '{cell.value}' with {year} A at {statement_name} row {row_num}, column {col_idx}")
            
            # Check for "Year E" or "Year"E"" patterns
            elif "year" in cell_value and (" e" in cell_value or '"e"' in cell_value or "'e'" in cell_value):
                # Determine which forecasted year this should be based on column position
                if col_idx in forecasted_columns:
                    year_idx = forecasted_columns.index(col_idx)
                    if year_idx < len(forecasted_years):
                        year = forecasted_years[year_idx]
                        cell.value = f"{year} E"
                        logger.debug(f"Replaced '{cell.value}' with {year} E at {statement_name} row {row_num}, column {col_idx}")
        
        # Write historical years to columns D-F (with "A" suffix) - overwrite regardless of existing content
        for idx, year in enumerate(recent_years):
            if idx < len(historical_columns):
                col_idx = historical_columns[idx]
                cell = worksheet.cell(row=row_num, column=col_idx)
                cell.value = f"{year} A"
        
        # Write forecasted years to columns G-K (with "E" suffix) - overwrite regardless of existing content
        for idx, year in enumerate(forecasted_years):
            if idx < len(forecasted_columns):
                col_idx = forecasted_columns[idx]
                cell = worksheet.cell(row=row_num, column=col_idx)
                cell.value = f"{year} E"
    
    # 1. Income Statement: Row 4 (fixed)
    write_years_to_row(4, "Income Statement")
    
    # 2. Balance Sheet: Row 24 (fixed)
    write_years_to_row(24, "Balance Sheet")
    
    # 3. Cash Flow: Row 69 (fixed)
    write_years_to_row(69, "Cash Flow")
    
    logger.info(f"Populated 3-statement year headers: Historical={recent_years}, Forecasted={forecasted_years}")


def overwrite_year_headers_in_statement_section(
    worksheet,
    statement_name: str,
    year_column_map: Dict[int, int],
    section_start_row: Optional[int] = None
) -> None:
    """
    Find and overwrite "Year 'A'" headers in a statement section header row.
    Works for Income Statement, Balance Sheet, and Cash Flow Statement.
    
    This function:
    1. Finds the statement section start row (if not provided)
    2. Searches for "Year" headers in the header row(s) of that section
    3. Overwrites them with actual years from year_column_map
    
    For Cash Flow Statement, only overwrites year headers in the "Cash Flows from Operating Activities" row.
    The "Cash Flow Statement" title row is NOT touched.
    For Income Statement and Balance Sheet, the years are typically 1-2 rows after the section start.
    
    Args:
        worksheet: openpyxl Worksheet object
        statement_name: Name of the statement ("Income Statement", "Balance Sheet", or "Cash Flow Statement")
        year_column_map: Dictionary mapping year -> column_index
        section_start_row: Optional starting row of the statement section (if None, will search for it)
    """
    # Find section start if not provided
    if section_start_row is None:
        section_start_row = find_statement_section_start(worksheet, statement_name)
    
    if not section_start_row:
        logger.warning(f"Could not find '{statement_name}' section start, skipping year header overwrite")
        return
    
    # For Cash Flow Statement, ONLY check the "Cash Flows from Operating Activities" row for year headers
    # Do NOT touch the "Cash Flow Statement" title row (section_start_row)
    # For Income Statement and Balance Sheet, check rows after the section start
    if "cash flow" in statement_name.lower():
        # Cash Flow Statement: ONLY check "Cash Flows from Operating Activities" header row for year headers
        # Skip the title row (section_start_row) - it should not be touched
        check_rows = []
        operating_header = find_statement_section_start(worksheet, "cash flows from operating activities", section_start_row)
        if operating_header:
            check_rows.append(operating_header)
    else:
        # Income Statement and Balance Sheet: years are typically 1-3 rows after section start
        check_rows = [section_start_row + offset for offset in range(1, 4)]
    
    # Search for "Year" headers in the identified rows
    for check_row in check_rows:
        if check_row > worksheet.max_row:
            continue
        
        # Check all columns that are in year_column_map for "Year" headers
        # This covers columns D-F and potentially more columns if year headers extend further
        year_header_found = False
        # Get all columns that are mapped to years
        mapped_columns = set(year_column_map.values())
        # Also check a reasonable range (D through J) to catch any year headers
        for col_idx in range(4, 11):  # Columns D through J
            cell = worksheet.cell(row=check_row, column=col_idx)
            cell_value = str(cell.value or "").strip()
            
            if "year" in cell_value.lower():
                year_header_found = True
                # Overwrite with actual year if this column is mapped
                if col_idx in mapped_columns:
                    # Find which year maps to this column
                    for year, mapped_col in year_column_map.items():
                        if mapped_col == col_idx:
                            cell.value = year
                            logger.info(f"Overwrote '{cell_value}' with {year} in {statement_name} row {check_row}, column {col_idx}")
                            break
        
        # If we found year headers in this row, we're done (don't need to check further rows)
        if year_header_found:
            break


def add_model_role_column(worksheet, role_column: str = "ZZ") -> None:
    """
    Add hidden model_role column to worksheet.
    
    Maps row labels to model_role values and writes them in the specified column.
    Hides the column for clean appearance.
    
    Args:
        worksheet: openpyxl Worksheet object
        role_column: Column letter (default "ZZ") where model_role will be stored
    """
    col_idx = column_index_from_string(role_column)
    
    # Write header
    header_cell = worksheet.cell(row=1, column=col_idx)
    header_cell.value = "model_role"
    
    # Find rows with known labels and map to model roles
    # Search column B (index 2) for row labels
    for row_idx, row in enumerate(worksheet.iter_rows(min_col=2, max_col=2, values_only=False), start=1):
        cell = row[0]
        if cell.value:
            cell_value = str(cell.value).strip().lower()
            
            # Check if this label maps to a model role
            # Sort patterns by length (longest first) to match more specific patterns first
            # This ensures "cost of revenue" matches before "revenue"
            sorted_patterns = sorted(ROW_LABEL_TO_MODEL_ROLE.items(), key=lambda x: len(x[0]), reverse=True)
            for label_pattern, model_role in sorted_patterns:
                if label_pattern in cell_value:
                    role_cell = worksheet.cell(row=row_idx, column=col_idx)
                    role_cell.value = model_role
                    logger.debug(f"Matched row {row_idx} label '{cell_value}' to model_role '{model_role}' using pattern '{label_pattern}'")
                    break
    
    # Hide the column
    worksheet.column_dimensions[role_column].hidden = True


def populate_historicals(
    worksheet,
    role_column: str,
    matrix: Dict[str, Dict[int, float]],
    year_column_map: Dict[int, int]
) -> None:
    """
    Populate historical data in worksheet using model_role column for row matching.
    Skips rows with "%" in the label (these are calculated, not filled).
    
    Args:
        worksheet: openpyxl Worksheet object
        role_column: Column letter containing model_role values
        matrix: Role/year matrix from build_role_year_matrix()
        year_column_map: Dictionary mapping year -> column_index
    """
    col_idx = column_index_from_string(role_column)
    
    # Iterate through all rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False), start=1):
        role_cell = row[0]
        model_role = (role_cell.value or "").strip()
        
        if not model_role:
            continue
        
        # Skip rows with "%" in the label (these are calculated, not filled)
        # Check the row label in column B
        label_cell = worksheet.cell(row=row_idx, column=2)
        label_value = str(label_cell.value or "").strip()
        if "%" in label_value:
            continue
        
        # Skip "Cash Flow Statement" title row - it should not have any values written to it
        if label_value.lower() == "cash flow statement":
            continue
        
        # Skip Cash Flow section header rows - they should not have any values written to them
        # Only year headers should be overwritten in these rows
        cash_flow_section_headers = [
            "cash flows from operating activities",
            "cash flows from investing activities",
            "cash flows from financing activities"
        ]
        if any(header in label_value.lower() for header in cash_flow_section_headers):
            continue
        
        # Look up values for this model_role
        role_values = matrix.get(model_role, {})
        
        if not role_values:
            continue
        
        # Write values to historical columns
        for year, column_idx in year_column_map.items():
            if year in role_values:
                value = role_values[year]
                target_cell = worksheet.cell(row=row_idx, column=column_idx)
                # Scale to millions
                if isinstance(value, (int, float)):
                    target_cell.value = value / MILLIONS_SCALE
                else:
                    target_cell.value = value


def _get_current_stock_price(ticker: str, use_fake_data: bool = False) -> Tuple[Optional[float], Optional[str]]:
    """
    Get current stock price and as-of date.
    
    Args:
        ticker: Stock ticker symbol (e.g., "AAPL")
        use_fake_data: If True, returns fake pricing data instead of fetching from yfinance
    
    Returns:
        Tuple of (price, as_of_date) or (None, None) if error
        as_of_date is in MM/DD/YYYY format
    """
    from datetime import date
    
    if use_fake_data:
        # Return fake pricing data for testing
        fake_price = 150.00  # Fake price
        price_date = date.today()
        as_of_date = price_date.strftime("%m/%d/%Y")
        logger.info(f"Using fake pricing data for {ticker}: ${fake_price:.2f}")
        return fake_price, as_of_date
    
    try:
        import yfinance as yf
        
        ticker_obj = yf.Ticker(ticker)
        # Get latest info - use info for current price
        info = ticker_obj.info
        
        # Try to get current price from info
        current_price = info.get("currentPrice") or info.get("regularMarketPrice") or info.get("previousClose")
        
        if current_price is None:
            # Fallback: get from history
            hist = ticker_obj.history(period="1d")
            if not hist.empty:
                current_price = float(hist["Close"].iloc[-1])
                price_date = hist.index[-1].date() if hasattr(hist.index[-1], 'date') else date.today()
            else:
                logger.warning(f"Could not fetch price for ticker {ticker}")
                return None, None
        else:
            # Use today's date for info-based price
            price_date = date.today()
        
        # Format date as MM/DD/YYYY
        as_of_date = price_date.strftime("%m/%d/%Y")
        
        return float(current_price), as_of_date
    except Exception as e:
        logger.warning(f"Error fetching stock price for {ticker}: {e}")
        return None, None


def _extract_income_statement_items(financial_json: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Extract income statement line items from financial JSON.
    
    Args:
        financial_json: JSON data structure with filings -> statements -> income_statement -> line_items
        
    Returns:
        List of income statement line items
    """
    try:
        filings = financial_json.get("filings", [])
        if not filings:
            logger.warning("No filings found in JSON")
            return []
        
        # Get first filing (most recent)
        filing = filings[0]
        statements = filing.get("statements", {})
        income_statement = statements.get("income_statement", {})
        line_items = income_statement.get("line_items", [])
        
        return line_items
    except (KeyError, TypeError) as e:
        logger.error(f"Error extracting income statement items: {e}")
        return []


def _get_period_mapping(line_items: List[Dict[str, Any]]) -> Dict[str, int]:
    """
    Extract periods from line items and map the 3 most recent to columns D, E, F.
    
    Args:
        line_items: List of line item dictionaries
        
    Returns:
        Dictionary mapping period date (YYYY-MM-DD) -> column index (D=4, E=5, F=6)
    """
    all_periods = set()
    
    for item in line_items:
        periods = item.get("periods", {})
        all_periods.update(periods.keys())
    
    if not all_periods:
        logger.warning("No periods found in line items")
        return {}
    
    # Sort periods descending (most recent first)
    sorted_periods = sorted(all_periods, reverse=True)
    
    # Take 3 most recent
    recent_periods = sorted_periods[:3]
    
    # Map to columns: D=4 (most recent), E=5 (middle), F=6 (oldest)
    period_mapping = {}
    for idx, period in enumerate(recent_periods):
        column_idx = 4 + idx  # D=4, E=5, F=6
        period_mapping[period] = column_idx
    
    logger.debug(f"Mapped periods to columns: {period_mapping}")
    return period_mapping


def _classify_line_item(item: Dict[str, Any]) -> str:
    """
    Classify a line item into a category based on model_role and label.
    
    Args:
        item: Line item dictionary
        
    Returns:
        Category string: "revenue", "cogs", "operating_expense", "operating_income", 
                        "interest", "tax", "net_income", "section_a", "section_b", "section_c"
    """
    model_role = (item.get("model_role") or "").strip().upper()
    label = (item.get("label") or "").strip().lower()
    tag = (item.get("tag") or "").strip().lower()
    
    # Primary classification by model_role
    if model_role == "IS_REVENUE":
        return "revenue"
    elif model_role == "IS_COGS":
        return "cogs"
    elif model_role == "IS_OPERATING_EXPENSE":
        return "operating_expense"
    elif model_role == "IS_OPERATING_INCOME":
        return "operating_income"
    elif model_role == "IS_INTEREST_EXPENSE" or "INTEREST" in model_role:
        return "interest"
    elif model_role == "IS_TAX_EXPENSE" or "TAX" in model_role:
        return "tax"
    elif model_role == "IS_NET_INCOME":
        return "net_income"
    
    # Fallback: fuzzy matching on label/tag
    text = f"{label} {tag}".lower()
    
    if any(word in text for word in ["revenue", "sales", "net sales", "total revenue"]):
        return "revenue"
    elif any(word in text for word in ["cogs", "cost of goods", "cost of sales", "cost of revenue"]):
        return "cogs"
    elif any(word in text for word in ["operating expense", "opex", "sg&a", "operating cost"]):
        return "operating_expense"
    elif any(word in text for word in ["operating income", "operating profit", "ebit"]):
        return "operating_income"
    elif any(word in text for word in ["interest expense", "interest cost"]):
        return "interest"
    elif any(word in text for word in ["tax expense", "income tax", "tax provision"]):
        return "tax"
    elif any(word in text for word in ["net income", "net profit", "net earnings"]):
        return "net_income"
    
    # Default classification for dynamic sections
    # Section A: items that might appear before gross income (unusual income, cost of services, etc.)
    # Section B: operating-related items
    # Section C: non-operating items (gains, losses, other income/expense)
    
    if any(word in text for word in ["gain", "loss", "other income", "other expense", "non-operating"]):
        return "section_c"
    elif any(word in text for word in ["depreciation", "amortization", "restructuring", "impairment"]):
        return "section_b"  # Operating-related
    else:
        # Default to section A if unclear
        return "section_a"


def _find_diluted_shares(line_items: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """
    Find diluted shares line item from income statement or other statements.
    
    Args:
        line_items: List of line items (may include non-income statement items)
        
    Returns:
        Line item dictionary if found, None otherwise
    """
    for item in line_items:
        tag = (item.get("tag") or "").strip().lower()
        label = (item.get("label") or "").strip().lower()
        model_role = (item.get("model_role") or "").strip().upper()
        
        text = f"{tag} {label}".lower()
        
        if any(word in text for word in ["diluted shares", "diluted share", "shares outstanding", "weighted average shares"]):
            return item
        elif model_role and "SHARES" in model_role:
            return item
    
    return None


def populate_income_statement(
    worksheet,
    financial_json: Dict[str, Any],
    start_row: int = 6
) -> int:
    """
    Populate Income Statement with fixed layout starting at B6 (or specified start_row).
    
    Fixed layout:
    - Revenue
    - Dynamic Section A items (pre-gross-profit)
    - Gross Income (formula)
    - Gross Margin % (formula)
    - Dynamic Section B items (operating expenses)
    - Operating Income
    - Operating Margin % (formula)
    - Dynamic Section C items (non-operating)
    - Income before Taxes (formula)
    - Income Tax Expense
    - Net Income
    - Net Income Margin % (formula)
    - EPS (formula, if diluted shares available)
    
    Args:
        worksheet: openpyxl Worksheet object
        financial_json: JSON data structure with filings -> statements -> income_statement -> line_items
        start_row: Starting row number (default 6, which is B6)
        
    Returns:
        Last row number used (after Income Statement ends, including EPS and diluted shares if present)
    """
    # Extract income statement items
    line_items = _extract_income_statement_items(financial_json)
    if not line_items:
        logger.warning("No income statement line items found")
        return start_row
    
    # Get period mapping (3 most recent periods -> columns C, D, E)
    period_mapping = _get_period_mapping(line_items)
    if not period_mapping:
        logger.warning("No periods found for income statement")
        return start_row
    
    # Classify all line items
    classified_items = {}
    for item in line_items:
        category = _classify_line_item(item)
        if category not in classified_items:
            classified_items[category] = []
        classified_items[category].append(item)
    
    # Track row numbers for formulas
    current_row = start_row
    revenue_row = None
    cogs_row = None
    gross_income_row = None
    operating_income_row = None
    interest_row = None
    income_before_taxes_row = None
    tax_expense_row = None
    net_income_row = None
    diluted_shares_row = None
    
    # Helper function to write values and label for a line item
    def write_line_item_values(row: int, item: Dict[str, Any], write_label: bool = True) -> None:
        # Always write label to column B if requested (assume column B is empty for line items)
        if write_label:
            label_cell = worksheet.cell(row=row, column=2)  # Column B
            label = item.get("label", "").strip()
            if label:
                label_cell.value = label
                logger.debug(f"Wrote label '{label}' to column B at row {row}")
        
        # Write values to columns D, E, F
        periods = item.get("periods", {})
        for period_date, column_idx in period_mapping.items():
            if period_date in periods:
                value = periods[period_date]
                if value is not None:
                    cell = worksheet.cell(row=row, column=column_idx)
                    cell.value = float(value)
                    cell.number_format = "General"
    
    # Helper function to write label for computed rows
    def write_computed_label(row: int, label: str) -> None:
        """Write label for computed/formula rows (Gross Income, margins, etc.)"""
        label_cell = worksheet.cell(row=row, column=2)  # Column B
        label_cell.value = label
        logger.debug(f"Wrote computed label '{label}' to column B at row {row}")
    
    # 1. Revenue (row 6 or start_row)
    revenue_items = classified_items.get("revenue", [])
    if revenue_items:
        revenue_item = revenue_items[0]  # Take first revenue item
        revenue_row = current_row
        write_line_item_values(revenue_row, revenue_item, write_label=True)  # Write label from JSON
        logger.debug(f"Wrote Revenue at row {revenue_row}")
        current_row += 1
    else:
        logger.warning("No Revenue item found")
        revenue_row = current_row
        current_row += 1
    
    # 2. Dynamic Section A: Items between Revenue and Gross Income
    section_a_items = classified_items.get("section_a", [])
    for item in section_a_items:
        write_line_item_values(current_row, item, write_label=True)  # Replace "Line Item" placeholder
        logger.debug(f"Wrote Section A item '{item.get('label')}' at row {current_row}")
        current_row += 1
    
    # 3. Gross Income (formula: Revenue - COGS)
    cogs_items = classified_items.get("cogs", [])
    if cogs_items:
        cogs_item = cogs_items[0]
        cogs_row = current_row
        write_line_item_values(cogs_row, cogs_item, write_label=True)  # Replace "Line Item" placeholder
        logger.debug(f"Wrote COGS at row {cogs_row}")
        current_row += 1
    
    gross_income_row = current_row
    if revenue_row and cogs_row:
        # Write label for Gross Income
        write_computed_label(gross_income_row, "Gross Income")
        
        # Formula: Revenue - COGS for each column
        for period_date, column_idx in period_mapping.items():
            formula = f"={get_column_letter(column_idx)}{revenue_row}-{get_column_letter(column_idx)}{cogs_row}"
            cell = worksheet.cell(row=gross_income_row, column=column_idx)
            cell.value = formula
            cell.number_format = "General"
        logger.debug(f"Wrote Gross Income formula at row {gross_income_row}")
    current_row += 1
    
    # 4. Gross Margin % (formula: Gross Income / Revenue)
    gross_margin_row = current_row
    if revenue_row and gross_income_row:
        # Write label for Gross Margin %
        write_computed_label(gross_margin_row, "Gross Margin %")
        
        for period_date, column_idx in period_mapping.items():
            formula = f"={get_column_letter(column_idx)}{gross_income_row}/{get_column_letter(column_idx)}{revenue_row}"
            cell = worksheet.cell(row=gross_margin_row, column=column_idx)
            cell.value = formula
            cell.number_format = "0.00%"  # Percentage format
        logger.debug(f"Wrote Gross Margin % formula at row {gross_margin_row}")
    current_row += 1
    
    # 5. Dynamic Section B: Operating Expenses (between Gross Margin % and Operating Income)
    section_b_items = classified_items.get("operating_expense", []) + classified_items.get("section_b", [])
    for item in section_b_items:
        write_line_item_values(current_row, item, write_label=True)  # Replace "Line Item" placeholder
        logger.debug(f"Wrote Section B item '{item.get('label')}' at row {current_row}")
        current_row += 1
    
    # 6. Operating Income
    operating_items = classified_items.get("operating_income", [])
    if operating_items:
        operating_item = operating_items[0]
        operating_income_row = current_row
        write_line_item_values(operating_income_row, operating_item, write_label=True)  # Write label from JSON
        logger.debug(f"Wrote Operating Income at row {operating_income_row}")
        current_row += 1
    else:
        logger.warning("No Operating Income item found")
        operating_income_row = current_row
        current_row += 1
    
    # 7. Operating Margin % (formula: Operating Income / Revenue)
    operating_margin_row = current_row
    if revenue_row and operating_income_row:
        # Write label for Operating Margin %
        write_computed_label(operating_margin_row, "Operating Margin %")
        
        for period_date, column_idx in period_mapping.items():
            formula = f"={get_column_letter(column_idx)}{operating_income_row}/{get_column_letter(column_idx)}{revenue_row}"
            cell = worksheet.cell(row=operating_margin_row, column=column_idx)
            cell.value = formula
            cell.number_format = "0.00%"  # Percentage format
        logger.debug(f"Wrote Operating Margin % formula at row {operating_margin_row}")
    current_row += 1
    
    # 8. Dynamic Section C: Non-operating items (between Operating Margin % and Income Before Taxes)
    section_c_items = classified_items.get("interest", []) + classified_items.get("section_c", [])
    for item in section_c_items:
        # Track interest row for Income before Taxes formula
        if item.get("model_role", "").upper() == "IS_INTEREST_EXPENSE" or "interest" in (item.get("label") or "").lower():
            if interest_row is None:
                interest_row = current_row
        write_line_item_values(current_row, item, write_label=True)  # Replace "Line Item" placeholder
        logger.debug(f"Wrote Section C item '{item.get('label')}' at row {current_row}")
        current_row += 1
    
    # 9. Income before Taxes (formula: Operating Income - Interest Expense)
    income_before_taxes_row = current_row
    if operating_income_row:
        # Write label for Income before Taxes
        write_computed_label(income_before_taxes_row, "Income before Taxes")
        
        if interest_row:
            # Formula: Operating Income - Interest Expense
            for period_date, column_idx in period_mapping.items():
                formula = f"={get_column_letter(column_idx)}{operating_income_row}-{get_column_letter(column_idx)}{interest_row}"
                cell = worksheet.cell(row=income_before_taxes_row, column=column_idx)
                cell.value = formula
                cell.number_format = "General"
        else:
            # No interest expense, Income before Taxes = Operating Income
            for period_date, column_idx in period_mapping.items():
                formula = f"={get_column_letter(column_idx)}{operating_income_row}"
                cell = worksheet.cell(row=income_before_taxes_row, column=column_idx)
                cell.value = formula
                cell.number_format = "General"
        logger.debug(f"Wrote Income before Taxes formula at row {income_before_taxes_row}")
    current_row += 1
    
    # 10. Income Tax Expense
    tax_items = classified_items.get("tax", [])
    if tax_items:
        tax_item = tax_items[0]
        tax_expense_row = current_row
        write_line_item_values(tax_expense_row, tax_item, write_label=True)  # Write label from JSON
        logger.debug(f"Wrote Income Tax Expense at row {tax_expense_row}")
        current_row += 1
    else:
        logger.warning("No Income Tax Expense item found")
        tax_expense_row = current_row
        current_row += 1
    
    # 11. Net Income
    net_income_items = classified_items.get("net_income", [])
    if net_income_items:
        net_income_item = net_income_items[0]
        net_income_row = current_row
        write_line_item_values(net_income_row, net_income_item, write_label=True)  # Write label from JSON
        logger.debug(f"Wrote Net Income at row {net_income_row}")
        current_row += 1
    else:
        logger.warning("No Net Income item found")
        net_income_row = current_row
        current_row += 1
    
    # 12. Net Income Margin % (formula: Net Income / Revenue)
    net_margin_row = current_row
    if revenue_row and net_income_row:
        # Write label for Net Income Margin %
        write_computed_label(net_margin_row, "Net Income Margin %")
        
        for period_date, column_idx in period_mapping.items():
            formula = f"={get_column_letter(column_idx)}{net_income_row}/{get_column_letter(column_idx)}{revenue_row}"
            cell = worksheet.cell(row=net_margin_row, column=column_idx)
            cell.value = formula
            cell.number_format = "0.00%"  # Percentage format
        logger.debug(f"Wrote Net Income Margin % formula at row {net_margin_row}")
    current_row += 1
    
    # 13. EPS (formula: Net Income / Diluted Shares, if available)
    # Try to find diluted shares in income statement items or all line items
    all_items = _extract_income_statement_items(financial_json)
    diluted_shares_item = _find_diluted_shares(all_items)
    
    if not diluted_shares_item:
        # Try searching in all statements
        try:
            filings = financial_json.get("filings", [])
            if filings:
                statements = filings[0].get("statements", {})
                all_statements_items = []
                for stmt_type, stmt_data in statements.items():
                    if isinstance(stmt_data, dict):
                        all_statements_items.extend(stmt_data.get("line_items", []))
                diluted_shares_item = _find_diluted_shares(all_statements_items)
        except Exception as e:
            logger.debug(f"Could not search all statements for diluted shares: {e}")
    
    eps_row = current_row
    if net_income_row and diluted_shares_item:
        # Write label for EPS
        write_computed_label(eps_row, "EPS")
        
        diluted_shares_row = current_row + 1  # Place diluted shares row after EPS for reference
        write_line_item_values(diluted_shares_row, diluted_shares_item, write_label=True)
        current_row = diluted_shares_row + 1  # Update current_row to include diluted shares row
        
        # Formula: Net Income / Diluted Shares
        for period_date, column_idx in period_mapping.items():
            formula = f"={get_column_letter(column_idx)}{net_income_row}/{get_column_letter(column_idx)}{diluted_shares_row}"
            cell = worksheet.cell(row=eps_row, column=column_idx)
            cell.value = formula
            cell.number_format = "General"
        logger.debug(f"Wrote EPS formula at row {eps_row}")
    else:
        # Still write EPS label even if diluted shares not found
        write_computed_label(eps_row, "EPS")
        current_row += 1  # Move to next row after EPS
        logger.debug("Diluted shares not found, leaving EPS blank")
    
    last_row = current_row - 1  # Last row used (current_row is already incremented)
    logger.info(f"Populated Income Statement starting at row {start_row}, ending at row {last_row}")
    return last_row


def populate_income_statement_values(
    worksheet,
    financial_data: Dict[str, Any],
    year_column_map: Dict[int, int],
    start_row: int = 6,
    use_formulas: bool = False
) -> int:
    """
    Populate Income Statement with values directly, calculating margins as values (not formulas).
    
    This function dynamically writes income statement data to Excel, calculating margins
    as actual values rather than Excel formulas. Works with both FMP format and structured JSON.
    
    Args:
        worksheet: openpyxl Worksheet object
        financial_data: Can be either:
            - FMP format: {"income_statements": [{date, revenue, costOfRevenue, ...}, ...]}
            - Structured JSON: {"filings": [{"statements": {"income_statement": {"line_items": [...]}}}]}
        year_column_map: Dictionary mapping year -> column_index (e.g., {2023: 4, 2024: 5, 2025: 6})
        start_row: Starting row number (default 6)
        use_formulas: If True, uses Excel formulas; if False, calculates values directly (default False)
        
    Returns:
        Last row number used
    """
    # Extract periods and values - handle both FMP and structured formats
    periods_data: List[Dict[str, Any]] = []
    
    # Check if it's FMP format (has income_statements array)
    if "income_statements" in financial_data:
        periods_data = financial_data["income_statements"]
    # Check if it's structured JSON format
    elif "filings" in financial_data:
        line_items = _extract_income_statement_items(financial_data)
        if line_items:
            # Convert line items to period-based format
            all_periods = set()
            for item in line_items:
                periods = item.get("periods", {})
                all_periods.update(periods.keys())
            
            # Build period dictionaries from line items
            for period_date in sorted(all_periods, reverse=True)[:3]:  # 3 most recent
                period_dict: Dict[str, Any] = {"date": period_date}
                for item in line_items:
                    periods = item.get("periods", {})
                    if period_date in periods:
                        tag = item.get("tag", "")
                        value = periods[period_date]
                        if tag:
                            period_dict[tag] = value
                periods_data.append(period_dict)
    
    if not periods_data:
        logger.warning("No income statement data found")
        return start_row
    
    # Get 3 most recent periods
    periods = periods_data[:3] if len(periods_data) >= 3 else periods_data
    
    # Extract years and map to columns
    period_year_map: Dict[str, int] = {}  # period_date -> year
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            year = int(str(date_str).split("-")[0])
            period_year_map[date_str] = year
    
    if not period_year_map:
        logger.warning("Could not extract years from income statement data")
        return start_row
    
    current_row = start_row
    
    # Helper to write value to a column for a specific period
    def write_value(row: int, value: Optional[float], period_date: str) -> None:
        if value is not None and period_date in period_year_map:
            year = period_year_map[period_date]
            if year in year_column_map:
                target_col = year_column_map[year]
                cell = worksheet.cell(row=row, column=target_col)
                cell.value = float(value)
                cell.number_format = "General"
    
    # Helper to write calculated margin value
    def write_margin_value(row: int, numerator: Optional[float], denominator: Optional[float], period_date: str) -> None:
        if numerator is not None and denominator is not None and denominator != 0:
            if period_date in period_year_map:
                year = period_year_map[period_date]
                if year in year_column_map:
                    margin_pct = numerator / denominator
                    cell = worksheet.cell(row=row, column=year_column_map[year])
                    cell.value = margin_pct
                    cell.number_format = "0.00%"
    
    # Map FMP field names to standard field names
    def get_field_value(period: Dict[str, Any], field_name: str) -> Optional[float]:
        # Try direct field name first
        value = period.get(field_name)
        if value is not None:
            return float(value)
        # Try camelCase variations
        camel_variations = {
            "revenue": ["revenue"],
            "cost_of_revenue": ["costOfRevenue", "cost_of_revenue"],
            "operating_expenses": ["operatingExpenses", "operating_expenses"],
            "operating_income": ["operatingIncome", "operating_income"],
            "interest_expense": ["interestExpense", "interest_expense"],
            "income_tax_expense": ["incomeTaxExpense", "income_tax_expense"],
            "net_income": ["netIncome", "net_income"],
        }
        if field_name in camel_variations:
            for variant in camel_variations[field_name]:
                value = period.get(variant)
                if value is not None:
                    return float(value)
        return None
    
    # 1. Revenue
    revenue_row = current_row
    label_cell = worksheet.cell(row=revenue_row, column=2)
    label_cell.value = "Revenue"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            revenue = get_field_value(period, "revenue")
            write_value(revenue_row, revenue, date_str)
    current_row += 1
    
    # 2. Cost of Revenue
    cogs_row = current_row
    label_cell = worksheet.cell(row=cogs_row, column=2)
    label_cell.value = "Cost of Revenue"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            cogs = get_field_value(period, "cost_of_revenue")
            write_value(cogs_row, cogs, date_str)
    current_row += 1
    
    # 3. Gross Income (calculated value)
    gross_income_row = current_row
    label_cell = worksheet.cell(row=gross_income_row, column=2)
    label_cell.value = "Gross Income"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            revenue = get_field_value(period, "revenue")
            cogs = get_field_value(period, "cost_of_revenue")
            if revenue is not None and cogs is not None:
                gross_income = revenue - cogs
                write_value(gross_income_row, gross_income, date_str)
    current_row += 1
    
    # 4. Gross Margin % (calculated value)
    gross_margin_row = current_row
    label_cell = worksheet.cell(row=gross_margin_row, column=2)
    label_cell.value = "Gross Margin %"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            revenue = get_field_value(period, "revenue")
            cogs = get_field_value(period, "cost_of_revenue")
            if revenue is not None and cogs is not None:
                gross_income = revenue - cogs
                write_margin_value(gross_margin_row, gross_income, revenue, date_str)
    current_row += 1
    
    # 5. Operating Expenses
    opex_row = current_row
    label_cell = worksheet.cell(row=opex_row, column=2)
    label_cell.value = "Operating Expenses"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            opex = get_field_value(period, "operating_expenses")
            write_value(opex_row, opex, date_str)
    current_row += 1
    
    # 6. Operating Income
    operating_income_row = current_row
    label_cell = worksheet.cell(row=operating_income_row, column=2)
    label_cell.value = "Operating Income"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            operating_income = get_field_value(period, "operating_income")
            write_value(operating_income_row, operating_income, date_str)
    current_row += 1
    
    # 7. Operating Margin % (calculated value)
    operating_margin_row = current_row
    label_cell = worksheet.cell(row=operating_margin_row, column=2)
    label_cell.value = "Operating Margin %"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            revenue = get_field_value(period, "revenue")
            operating_income = get_field_value(period, "operating_income")
            write_margin_value(operating_margin_row, operating_income, revenue, date_str)
    current_row += 1
    
    # 8. Interest Expense
    interest_row = current_row
    label_cell = worksheet.cell(row=interest_row, column=2)
    label_cell.value = "Interest Expense"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            interest = get_field_value(period, "interest_expense")
            write_value(interest_row, interest, date_str)
    current_row += 1
    
    # 9. Income Tax Expense
    tax_row = current_row
    label_cell = worksheet.cell(row=tax_row, column=2)
    label_cell.value = "Income Tax Expense"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            tax = get_field_value(period, "income_tax_expense")
            write_value(tax_row, tax, date_str)
    current_row += 1
    
    # 10. Net Income
    net_income_row = current_row
    label_cell = worksheet.cell(row=net_income_row, column=2)
    label_cell.value = "Net Income"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            net_income = get_field_value(period, "net_income")
            write_value(net_income_row, net_income, date_str)
    current_row += 1
    
    # 11. Net Income Margin % (calculated value)
    net_margin_row = current_row
    label_cell = worksheet.cell(row=net_margin_row, column=2)
    label_cell.value = "Net Income Margin %"
    for period in periods:
        date_str = period.get("date", "")
        if date_str:
            revenue = get_field_value(period, "revenue")
            net_income = get_field_value(period, "net_income")
            write_margin_value(net_margin_row, net_income, revenue, date_str)
    current_row += 1
    
    last_row = current_row - 1
    logger.info(f"Populated Income Statement values starting at row {start_row}, ending at row {last_row}")
    return last_row


def populate_cover_sheet(
    workbook,
    company_name: str,
    ticker: str,
    date: Optional[str] = None,
    use_fake_pricing: bool = False,
    quote_data: Optional[Dict[str, Any]] = None
) -> None:
    """
    Populate both Cover sheet and Summary sheet with company information.
    
    Cover Sheet writes:
    - Company name to B2
    - Ticker to B3
    - Date to B4 (defaults to current date if not provided)
    
    Summary Sheet writes:
    - Company name to B2
    - Ticker to C3
    - Current stock price to C4
    - As-of date for pricing data to D4
    
    Args:
        workbook: openpyxl Workbook object
        company_name: Company name to write
        ticker: Ticker symbol to write
        date: Optional date string (defaults to current date in MM/DD/YYYY format)
        use_fake_pricing: If True, uses fake pricing data instead of fetching from yfinance
    """
    # Populate Cover sheet
    cover_sheet = None
    if "Cover" in workbook.sheetnames:
        cover_sheet = workbook["Cover"]
    elif "Cover Sheet" in workbook.sheetnames:
        cover_sheet = workbook["Cover Sheet"]
    elif "Cover Page" in workbook.sheetnames:
        cover_sheet = workbook["Cover Page"]
    
    if cover_sheet:
        # Write company name to B2 (row 2, column 2) - always overwrite
        cover_sheet.cell(row=2, column=2).value = company_name
        logger.debug(f"Wrote company name '{company_name}' to Cover Sheet B2")
        
        # Write ticker to B3 (row 3, column 2)
        cover_sheet.cell(row=3, column=2).value = ticker
        
        # Write date to B4 (row 4, column 2)
        if date is None:
            date = datetime.now().strftime("%m/%d/%Y")
        cover_sheet.cell(row=4, column=2).value = date
        
        logger.info(f"Populated Cover sheet: Company={company_name}, Ticker={ticker}, Date={date}")
    else:
        logger.warning("Cover sheet not found, skipping Cover sheet population")
    
    # Populate Summary sheet
    summary_sheet = None
    if "Summary" in workbook.sheetnames:
        summary_sheet = workbook["Summary"]
    
    if summary_sheet:
        # Write company name to B2 (row 2, column 2) - always overwrite
        summary_sheet.cell(row=2, column=2).value = company_name
        logger.debug(f"Wrote company name '{company_name}' to Summary Sheet B2")
        
        # Write ticker to C3 (row 3, column 3)
        summary_sheet.cell(row=3, column=3).value = ticker
        
        # Get stock price from quote data or use fallback
        stock_price = None
        if quote_data and not use_fake_pricing:
            # Try multiple price fields from FMP quote data
            stock_price = (
                quote_data.get("price") or
                quote_data.get("lastPrice") or
                quote_data.get("close") or
                quote_data.get("previousClose")
            )
        
        # Fallback to fake/hardcoded price if needed
        if stock_price is None:
            if use_fake_pricing:
                stock_price = 150.00
                logger.info("Using fake pricing data")
            else:
                logger.warning("No quote data available, using default price")
                stock_price = 150.00  # Default fallback
        
        # Write price to C4 (row 4, column 3)
        summary_sheet.cell(row=4, column=3).value = float(stock_price)
        logger.info(f"Wrote price ${stock_price:.2f} to Summary sheet C4 (row 4, column 3)")
        
        # Write as-of date directly to the right (D4 - row 4, column 4)
        if date is None:
            date = datetime.now().strftime("%m/%d/%Y")
        as_of_text = f"As of {date}"
        summary_sheet.cell(row=4, column=4).value = as_of_text
        logger.info(f"Wrote as-of date '{as_of_text}' to Summary sheet D4 (row 4, column 4)")
        
        logger.info(f"Populated Summary sheet: Company={company_name}, Ticker={ticker}, Price=${stock_price:.2f}, As-of={as_of_text}")
    else:
        logger.warning("Summary sheet not found, skipping Summary sheet population")


def populate_three_statement_assumptions(
    worksheet,
    assumptions: Dict[str, Any],
    forecast_periods: int = 5,
    income_statement_end_row: Optional[int] = None
) -> None:
    """
    Populate 3-statement model assumptions in Excel worksheet.
    
    Supports three forecast methods:
    - "step": Step increase/decrease per period (value written to O column)
    - "constant": Constant value for all periods (value written to O, P, Q, R, S)
    - "custom": Custom values for each period (5 values written to O, P, Q, R, S)
    
    Args:
        worksheet: openpyxl Worksheet object for "3 Statement" sheet
        assumptions: Dictionary with structure:
            {
                "income_statement": {
                    "revenue": {"type": "step", "value": 0.08},
                    ...
                },
                "balance_sheet": {...},
                "cash_flow": {...}
            }
        forecast_periods: Number of forecast periods (default 5)
        income_statement_end_row: Optional row number where Income Statement ends (for finding Assets row)
    """
    # Find Income Statement assumptions header (search for "IS Assumptions" or "Income Statement Assumptions")
    is_assumptions_row = None
    for row in range(1, 200):
        cell_value = worksheet.cell(row=row, column=12).value  # Column L
        if cell_value:
            cell_str = str(cell_value).lower().strip()
            if "is assumptions" in cell_str or ("assumptions" in cell_str and "income" in cell_str):
                is_assumptions_row = row
                break
    
    # Find Balance Sheet anchor: "Assets" row (start search after Income Statement if provided)
    bs_anchor_row = None
    if income_statement_end_row:
        bs_anchor_row = find_statement_section_start(worksheet, "Assets", start_row=income_statement_end_row + 1)
    else:
        bs_anchor_row = find_statement_section_start(worksheet, "Assets")
    
    # Find Cash Flow anchor: "Cash Flows from Operating Activities" row
    cf_anchor_row = find_statement_section_start(worksheet, "Cash Flows from Operating Activities")
    
    # Use fixed rows for assumptions anchoring (matching year header positions)
    # Income Statement assumptions: anchored to row 4 (same as year headers)
    if is_assumptions_row is None:
        is_assumptions_row = 4
        logger.info("Using fixed row 4 for Income Statement assumptions")
    else:
        # Override with fixed row 4 to match year headers
        is_assumptions_row = 4
        logger.info("Using fixed row 4 for Income Statement assumptions (overriding dynamic search)")
    
    # Balance Sheet assumptions: anchored to row 24 (same as year headers)
    if bs_anchor_row is None:
        bs_anchor_row = 24
        logger.info("Using fixed row 24 for Balance Sheet assumptions")
    else:
        # Override with fixed row 24 to match year headers
        bs_anchor_row = 24
        logger.info("Using fixed row 24 for Balance Sheet assumptions (overriding dynamic search)")
    
    # Cash Flow assumptions: anchored to row 69 (same as year headers)
    if cf_anchor_row is None:
        cf_anchor_row = 69
        logger.info("Using fixed row 69 for Cash Flow assumptions")
    else:
        # Override with fixed row 69 to match year headers
        cf_anchor_row = 69
        logger.info("Using fixed row 69 for Cash Flow assumptions (overriding dynamic search)")
    
    # Mapping of assumption names to their row offsets from section headers
    # Column N = 14 (Type), Column O = 15 (Value start), P=16, Q=17, R=18, S=19
    # Supports both legacy format names and new format names
    assumption_mapping = {
        "income_statement": {
            "revenue": {"row_offset": 1},  # IS header + 1 = row 5
            "gross_margin": {"row_offset": 2},  # IS header + 2 = row 6
            "operating_margin": {"row_offset": 3},  # IS header + 3 = row 7 (legacy format)
            "operating_cost": {"row_offset": 3},  # IS header + 3 = row 7 (new format)
            "tax_rate": {"row_offset": 4},  # IS header + 4 = row 8
            "interest_rate_debt": {"row_offset": 5},  # IS header + 5 = row 9 (legacy format)
            "interest_rate_on_debt": {"row_offset": 5},  # IS header + 5 = row 9 (new format)
        },
        "balance_sheet": {
            "depreciation_ppe": {"row_offset": 1},  # BS header + 1 = row 25 (legacy format)
            "depreciation_pct_of_ppe": {"row_offset": 1},  # BS header + 1 = row 25 (new format)
            "inventory_sales": {"row_offset": 2},  # BS header + 2 = row 26 (legacy format)
            "inventory": {"row_offset": 2},  # BS header + 2 = row 26 (new format)
            "lt_debt_change": {"row_offset": 3},  # BS header + 3 = row 27 (legacy format)
            "total_debt_amount": {"row_offset": 3},  # BS header + 3 = row 27 (new format)
        },
        "cash_flow": {
            "share_repurchases": {"row_offset": 1},  # CF header (69) + 1 = row 70 (legacy format)
            "share_repurchase": {"row_offset": 1},  # CF header (69) + 1 = row 70 (new format)
            "dividends_ni": {"row_offset": 2},  # CF header (69) + 2 = row 71 (legacy format)
            "dividend_pct_of_net_income": {"row_offset": 2},  # CF header (69) + 2 = row 71 (new format)
            "capex": {"row_offset": 3},  # CF header (69) + 3 = row 72
        },
    }
    
    # Column indices: N=14 (Type), O=15, P=16, Q=17, R=18, S=19
    TYPE_COLUMN = 14
    VALUE_START_COLUMN = 15
    
    # Normalize assumptions format: handle both legacy format (assumptions key) and new format (income_statement key)
    normalized_assumptions = {}
    if "assumptions" in assumptions:
        # Legacy format: convert "assumptions" -> "income_statement"
        normalized_assumptions["income_statement"] = assumptions["assumptions"]
    if "income_statement" in assumptions:
        normalized_assumptions["income_statement"] = assumptions["income_statement"]
    if "balance_sheet" in assumptions:
        normalized_assumptions["balance_sheet"] = assumptions["balance_sheet"]
    if "cash_flow" in assumptions:
        normalized_assumptions["cash_flow"] = assumptions["cash_flow"]
    
    # Process each statement type
    for statement_type, statement_assumptions in normalized_assumptions.items():
        if statement_type not in assumption_mapping:
            logger.warning(f"Unknown statement type: {statement_type}")
            continue
        
        # Get the anchor row for this statement type
        if statement_type == "income_statement":
            section_header_row = is_assumptions_row
        elif statement_type == "balance_sheet":
            # Balance Sheet assumptions are anchored to "Assets" row
            section_header_row = bs_anchor_row
        elif statement_type == "cash_flow":
            # Cash Flow assumptions are anchored to "Cash Flows from Operating Activities" row
            section_header_row = cf_anchor_row
        else:
            logger.warning(f"Unknown statement type: {statement_type}")
            continue
        
        mapping = assumption_mapping[statement_type]
        
        # Process each assumption
        for assumption_name, assumption_data in statement_assumptions.items():
            if assumption_name not in mapping:
                logger.warning(f"Unknown assumption '{assumption_name}' for {statement_type}")
                continue
            
            # Calculate absolute row from section header + offset
            row_offset = mapping[assumption_name]["row_offset"]
            row = section_header_row + row_offset
            assumption_type = assumption_data.get("type", "").lower()
            
            # Handle both "value" (singular) and "values" (array) formats
            value = assumption_data.get("value")
            values = assumption_data.get("values")
            
            # If we have values array but no value, extract appropriate value based on type
            if value is None and values is not None and isinstance(values, list) and len(values) > 0:
                if assumption_type == "step":
                    # For step, use the increment (difference between first two values, or first value if only one)
                    if len(values) >= 2:
                        value = values[1] - values[0]
                    else:
                        value = values[0]
                elif assumption_type == "constant":
                    # For constant, use first value
                    value = values[0]
                elif assumption_type == "custom":
                    # For custom, use the values array directly
                    value = values
            
            if not assumption_type:
                logger.warning(f"Missing type for assumption '{assumption_name}'")
                continue
            
            # Write type to column N
            type_cell = worksheet.cell(row=row, column=TYPE_COLUMN)
            type_cell.value = assumption_type
            
            # Handle different types
            if assumption_type == "step":
                # Step type: write value to O column only
                if value is not None:
                    value_cell = worksheet.cell(row=row, column=VALUE_START_COLUMN)
                    value_cell.value = value
                    # Set number format to prevent date interpretation
                    # Use General format for numbers, or percentage format if value is < 1
                    if isinstance(value, (int, float)):
                        if abs(value) < 1 and value != 0:
                            value_cell.number_format = "0.00%"
                        else:
                            value_cell.number_format = "General"
                    logger.debug(f"Wrote step assumption '{assumption_name}': {value} to row {row}, column {VALUE_START_COLUMN}")
            
            elif assumption_type == "constant":
                # Constant type: write value to all period columns (O, P, Q, R, S)
                if value is not None:
                    for period_idx in range(forecast_periods):
                        col = VALUE_START_COLUMN + period_idx
                        value_cell = worksheet.cell(row=row, column=col)
                        value_cell.value = value
                        # Set number format to prevent date interpretation
                        if isinstance(value, (int, float)):
                            if abs(value) < 1 and value != 0:
                                value_cell.number_format = "0.00%"
                            else:
                                value_cell.number_format = "General"
                    logger.debug(f"Wrote constant assumption '{assumption_name}': {value} to row {row}, columns {VALUE_START_COLUMN}-{VALUE_START_COLUMN + forecast_periods - 1}")
            
            elif assumption_type == "custom":
                # Custom type: write array of values to O, P, Q, R, S
                # Handle both "value" (which might be a list) and "values" array
                values_to_write = value if isinstance(value, list) else values
                if isinstance(values_to_write, list) and len(values_to_write) >= forecast_periods:
                    for period_idx in range(forecast_periods):
                        col = VALUE_START_COLUMN + period_idx
                        value_cell = worksheet.cell(row=row, column=col)
                        period_value = values_to_write[period_idx]
                        value_cell.value = period_value
                        # Set number format to prevent date interpretation
                        if isinstance(period_value, (int, float)):
                            if abs(period_value) < 1 and period_value != 0:
                                value_cell.number_format = "0.00%"
                            else:
                                value_cell.number_format = "General"
                    logger.debug(f"Wrote custom assumption '{assumption_name}': {values_to_write[:forecast_periods]} to row {row}, columns {VALUE_START_COLUMN}-{VALUE_START_COLUMN + forecast_periods - 1}")
                else:
                    logger.warning(f"Custom assumption '{assumption_name}' must have at least {forecast_periods} values, got {len(values_to_write) if isinstance(values_to_write, list) else 'non-list'}")
            
            else:
                logger.warning(f"Unknown assumption type '{assumption_type}' for '{assumption_name}'")
    
    logger.info("Populated 3-statement assumptions")


# Statement structure definitions for dynamic line item insertion
STATEMENT_STRUCTURE = {
    "income_statement": {
        "section_name": "Income Statement",
        "anchors": [
            "gross profit",
            "gross margin %",
            "operating profit",
            "operating margin %",
            "income before taxes",
            "income tax expense",
            "net income",
            "net income margin %",
            "eps",
        ],
        "placeholder_text": "line item",
    },
    "balance_sheet": {
        "section_name": "Balance Sheet",
        "subsections": {
            "assets": {
                "header": "assets",
                "anchors": [
                    "total current assets",
                    "total assets",
                ],
                "placeholder_text": "line item",
            },
            "liabilities": {
                "header": "liabilities",
                "anchors": [
                    "total current liabilities",
                    "total liabilities",
                ],
                "placeholder_text": "line item",
            },
            "equity": {
                "header": "equity",
                "anchors": [
                    "total equity",
                    "total liabilities and equity",
                ],
                "placeholder_text": "line item",
            },
        },
    },
    "cash_flow": {
        "section_name": "Cash Flow Statement",
        "subsections": {
            "operating": {
                "header": "cash flows from operating activities",
                "anchors": [
                    "net cash provided by/(used in) operating activities",
                ],
                "placeholder_text": "line item",
            },
            "investing": {
                "header": "cash flows from investing activities",
                "anchors": [
                    "net cash provided by/(used in) investing activities",
                ],
                "placeholder_text": "line item",
            },
            "financing": {
                "header": "cash flows from financing activities",
                "anchors": [
                    "net cash provided by/(used in) financing activities",
                ],
                "placeholder_text": "line item",
            },
        },
        "final_items": [
            "effects of exchange rate",
            "net increase/(decrease) in cash, cash equivalents, and restricted cash",
            "cash, cash equivalents, and restricted cash at beginning of period",
            "net increase/(decrease) in cash, cash equivalents, and restricted cash",
            "cash, cash equivalents, and restricted cash at end of period",
        ],
    },
}


# Helper functions for dynamic line item insertion
def find_statement_section_start(worksheet, section_name: str, start_row: int = 1, max_row: int = 200) -> Optional[int]:
    """
    Find the starting row of a statement section by searching for the section name in column B.
    
    For main statement names ("Income Statement", "Balance Sheet", "Cash Flow Statement"),
    uses exact matching. For subsection headers, uses substring matching.
    
    Args:
        worksheet: openpyxl Worksheet object
        section_name: Name of the section to find (e.g., "Income Statement")
        start_row: Row to start searching from
        max_row: Maximum row to search to
        
    Returns:
        Row number if found, None otherwise
    """
    section_name_lower = section_name.lower()
    
    # Main statement names require exact matching
    main_statements = ["income statement", "balance sheet", "cash flow statement"]
    use_exact_match = section_name_lower in main_statements
    
    for row_idx in range(start_row, min(max_row + 1, worksheet.max_row + 1)):
        cell = worksheet.cell(row=row_idx, column=2)  # Column B
        cell_value = str(cell.value or "").strip().lower()
        
        if use_exact_match:
            # Exact match for main statement names
            if cell_value == section_name_lower:
                return row_idx
        else:
            # Substring match for subsection headers
            if section_name_lower in cell_value:
                return row_idx
    return None


def find_anchor_rows(worksheet, anchors: List[str], start_row: int, end_row: int) -> Dict[str, int]:
    """
    Find anchor rows within a specified range.
    
    Args:
        worksheet: openpyxl Worksheet object
        anchors: List of anchor text to search for (case-insensitive)
        start_row: Starting row to search
        end_row: Ending row to search
        
    Returns:
        Dictionary mapping anchor text (lowercase) to row number
    """
    anchor_map: Dict[str, int] = {}
    anchors_lower = [anchor.lower() for anchor in anchors]
    
    for row_idx in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
        cell = worksheet.cell(row=row_idx, column=2)  # Column B
        cell_value = str(cell.value or "").strip().lower()
        
        for anchor in anchors_lower:
            if anchor in cell_value and anchor not in anchor_map:
                anchor_map[anchor] = row_idx
    
    return anchor_map


def find_line_item_placeholders(worksheet, placeholder_text: str, start_row: int, end_row: int, exclude_anchors: Optional[List[str]] = None) -> List[int]:
    """
    Find rows containing placeholder text in column B.
    Excludes anchor rows (like "Total Current Assets", "Total Current Liabilities", etc.)
    and section header rows (rows that contain section titles like "Cash Flows from Operating Activities").
    
    Args:
        worksheet: openpyxl Worksheet object
        placeholder_text: Text to search for (e.g., "Line Item")
        start_row: Starting row to search
        end_row: Ending row to search
        exclude_anchors: Optional list of anchor text to exclude from results
        
    Returns:
        List of row numbers containing the placeholder (excluding anchor rows and section headers)
    """
    placeholder_lower = placeholder_text.lower()
    placeholder_rows: List[int] = []
    
    # Common anchor patterns to exclude (Balance Sheet terms that shouldn't appear in Cash Flow)
    default_exclude = [
        "total current assets",
        "total assets",
        "total current liabilities",
        "total liabilities",
        "total equity",
        "total liabilities and equity",
        # Section header patterns to exclude
        "cash flows from operating activities",
        "cash flows from investing activities",
        "cash flows from financing activities",
        "operating activities",
        "investing activities",
        "financing activities"
    ]
    
    exclude_patterns = (exclude_anchors or []) + default_exclude
    exclude_lower = [pattern.lower() for pattern in exclude_patterns]
    
    for row_idx in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
        cell = worksheet.cell(row=row_idx, column=2)  # Column B
        cell_value = str(cell.value or "").strip().lower()
        
        # Skip if this is an anchor row or section header row
        is_excluded = False
        for exclude_pattern in exclude_lower:
            if exclude_pattern in cell_value:
                is_excluded = True
                break
        
        # Also check if this row has values in data columns (D-F) - if so, it's likely a header row with data
        # Section headers shouldn't have data values, so skip rows that do
        has_data_values = False
        for col_idx in range(4, 7):  # Columns D, E, F
            data_cell = worksheet.cell(row=row_idx, column=col_idx)
            if data_cell.value is not None and isinstance(data_cell.value, (int, float)):
                # Check if it's a reasonable data value (not a year like 2022-2024)
                if isinstance(data_cell.value, (int, float)) and data_cell.value > 1000:
                    has_data_values = True
                    break
        
        # Only include rows that:
        # 1. Contain the placeholder text
        # 2. Are not excluded patterns (anchors/section headers)
        # 3. Don't have data values (which would indicate it's a header row with totals)
        if not is_excluded and not has_data_values and placeholder_lower in cell_value:
            placeholder_rows.append(row_idx)
    
    return placeholder_rows


def preserve_assumptions_on_row_insert(worksheet, insert_row: int, num_rows: int = 1) -> None:
    """
    When inserting rows, preserve assumptions columns (L-O) by copying from row above.
    
    Args:
        worksheet: openpyxl Worksheet object
        insert_row: Row number where rows will be inserted
        num_rows: Number of rows to insert
    """
    # Assumptions columns: L=12, M=13, N=14, O=15
    assumptions_columns = [12, 13, 14, 15]
    
    # Copy assumptions from the row above the insertion point
    source_row = insert_row - 1
    
    if source_row < 1:
        return  # No row above to copy from
    
    # Copy to each inserted row
    for row_offset in range(num_rows):
        target_row = insert_row + row_offset
        for col in assumptions_columns:
            source_cell = worksheet.cell(row=source_row, column=col)
            target_cell = worksheet.cell(row=target_row, column=col)
            
            # Copy value
            target_cell.value = source_cell.value
            
            # Copy number format (this is safe and important)
            if source_cell.has_style:
                target_cell.number_format = source_cell.number_format
                
            # Note: We don't copy font, fill, border, alignment directly as StyleProxy objects
            # are not hashable. The template formatting should handle visual styling.
            # If specific formatting is needed, it can be applied via template or
            # by copying the style ID: target_cell._style = source_cell._style.copy()


def group_line_items_by_statement(line_items: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Group line items by statement type based on model_role prefix.
    
    Args:
        line_items: List of line item dictionaries from JSON
        
    Returns:
        Dictionary with keys: "income_statement", "balance_sheet", "cash_flow"
        Each value is a list of line items for that statement type
    """
    grouped: Dict[str, List[Dict[str, Any]]] = {
        "income_statement": [],
        "balance_sheet": [],
        "cash_flow": [],
    }
    
    for item in line_items:
        model_role = (item.get("model_role") or "").strip()
        
        if model_role.startswith("IS_"):
            grouped["income_statement"].append(item)
        elif model_role.startswith("BS_"):
            grouped["balance_sheet"].append(item)
        elif model_role.startswith("CF_"):
            grouped["cash_flow"].append(item)
    
    return grouped


def replace_line_item_placeholders(
    worksheet,
    placeholder_rows: List[int],
    line_items: List[Dict[str, Any]],
    matrix: Dict[str, Dict[int, float]],
    year_column_map: Dict[int, int]
) -> None:
    """
    Replace "Line Item" placeholders with actual line item data.
    Overwrites existing placeholder rows instead of inserting new ones.
    Only inserts additional rows if there are more line items than placeholders.
    
    IMPORTANT: Never writes to row 16 (year headers), section header rows, or assumptions rows.
    
    Args:
        worksheet: openpyxl Worksheet object
        placeholder_rows: List of row numbers containing "Line Item" placeholders
        line_items: List of line items to populate
        matrix: Role/year matrix from build_role_year_matrix()
        year_column_map: Dictionary mapping year -> column_index
    """
    # Filter out row 16 (year headers) - never overwrite year headers
    placeholder_rows = [row for row in placeholder_rows if row != 16]
    
    # Filter out section header rows (rows that contain section titles like "Cash Flows from Operating Activities")
    section_header_patterns = [
        "cash flows from operating activities",
        "cash flows from investing activities",
        "cash flows from financing activities",
        "operating activities",
        "investing activities",
        "financing activities"
    ]
    
    filtered_rows = []
    for row in placeholder_rows:
        cell_value = str(worksheet.cell(row=row, column=2).value or "").strip().lower()
        is_section_header = any(pattern in cell_value for pattern in section_header_patterns)
        
        # Also check if this row has data values in columns D-F (section headers shouldn't have data)
        has_data_values = False
        for col_idx in range(4, 7):  # Columns D, E, F
            data_cell = worksheet.cell(row=row, column=col_idx)
            if data_cell.value is not None and isinstance(data_cell.value, (int, float)):
                # Check if it's a reasonable data value (not a year like 2022-2024)
                if isinstance(data_cell.value, (int, float)) and data_cell.value > 1000:
                    has_data_values = True
                    break
        
        if not is_section_header and not has_data_values:
            filtered_rows.append(row)
    
    placeholder_rows = filtered_rows
    
    if not placeholder_rows:
        logger.warning("No valid placeholder rows found (excluding row 16 and section headers)")
        return
    
    # Replace placeholders one-to-one with line items
    num_replacements = min(len(placeholder_rows), len(line_items))
    
    for idx in range(num_replacements):
        row = placeholder_rows[idx]
        item = line_items[idx]
        model_role = (item.get("model_role") or "").strip()
        label = item.get("label", "").strip()
        
        # Overwrite placeholder text with actual label (only if cell contains "Line Item")
        label_cell = worksheet.cell(row=row, column=2)
        cell_value = str(label_cell.value or "").strip()
        if "line item" in cell_value.lower() and label:
            label_cell.value = label
            logger.debug(f"Replaced 'Line Item' with '{label}' at row {row}")
        elif not label_cell.value and label:
            # If cell is empty, write the label
            label_cell.value = label
        
        # Write model_role to role column (ZZ) for later reference
        role_cell = worksheet.cell(row=row, column=column_index_from_string("ZZ"))
        role_cell.value = model_role
        
        # Write historical values (only to data columns D, E, F - not year header row)
        role_values = matrix.get(model_role, {})
        for year, column_idx in year_column_map.items():
            if year in role_values:
                value = role_values[year]
                target_cell = worksheet.cell(row=row, column=column_idx)
                target_cell.value = value
                target_cell.number_format = "General"
    
    # If there are more line items than placeholders, insert rows ONLY in data columns (A-K)
    # This prevents assumptions columns (L-O) from being shifted
    if len(line_items) > len(placeholder_rows):
        extra_items = line_items[num_replacements:]
        insert_after_row = placeholder_rows[-1] if placeholder_rows else None
        
        if insert_after_row:
            # Insert rows only affecting columns A-K (data columns)
            # We'll manually copy cells to avoid shifting assumptions
            num_extra = len(extra_items)
            
            # Insert rows (this will shift everything, but we'll fix assumptions)
            worksheet.insert_rows(insert_after_row + 1, num_extra)
            
            # Restore assumptions columns from their original positions
            # Assumptions should NOT move - they're anchored to their section headers
            # So we need to copy assumptions back from where they should be
            # For now, copy from the row above the insertion point
            preserve_assumptions_on_row_insert(worksheet, insert_after_row + 1, num_extra)
            
            # Populate the extra inserted rows
            for idx, item in enumerate(extra_items):
                row = insert_after_row + 1 + idx
                model_role = (item.get("model_role") or "").strip()
                label = item.get("label", "").strip()
                
                # Write label to column B (new rows, so always write)
                label_cell = worksheet.cell(row=row, column=2)
                if label:
                    label_cell.value = label
                
                # Write model_role to role column (ZZ)
                role_cell = worksheet.cell(row=row, column=column_index_from_string("ZZ"))
                role_cell.value = model_role
                
                # Write historical values
                role_values = matrix.get(model_role, {})
                for year, column_idx in year_column_map.items():
                    if year in role_values:
                        value = role_values[year]
                        target_cell = worksheet.cell(row=row, column=column_idx)
                        target_cell.value = value
                        target_cell.number_format = "General"


def populate_income_statement_line_items(
    worksheet,
    line_items: List[Dict[str, Any]],
    matrix: Dict[str, Dict[int, float]],
    year_column_map: Dict[int, int],
    section_start_row: int
) -> None:
    """
    Populate Income Statement with dynamic line items by replacing placeholders.
    
    Groups line items by their position relative to anchors:
    - Before Gross Income: Revenue items
    - Between Gross Income and Operating Income: COGS and other items
    - Between Operating Income and Income before Taxes: Operating expenses
    - Between Income before Taxes and Net Income: Tax-related items
    
    Args:
        worksheet: openpyxl Worksheet object
        line_items: List of Income Statement line items
        matrix: Role/year matrix
        year_column_map: Dictionary mapping year -> column_index
        section_start_row: Starting row of Income Statement section
    """
    structure = STATEMENT_STRUCTURE["income_statement"]
    anchors = structure["anchors"]
    placeholder_text = structure["placeholder_text"]
    
    # Find anchor rows
    anchor_map = find_anchor_rows(worksheet, anchors, section_start_row, section_start_row + 100)
    
    # Group line items by position relative to anchors
    gross_profit_row = anchor_map.get("gross profit")
    operating_profit_row = anchor_map.get("operating profit")
    income_before_taxes_row = anchor_map.get("income before taxes")
    net_income_row = anchor_map.get("net income")
    
    if not gross_profit_row:
        logger.warning("Could not find 'Gross Profit' anchor in Income Statement")
        return
    
    # Find placeholders before Gross Profit (Revenue items)
    if gross_profit_row:
        placeholder_rows_before_gross = find_line_item_placeholders(
            worksheet, placeholder_text, section_start_row + 1, gross_profit_row - 1
        )
        revenue_items = [item for item in line_items if item.get("model_role") == "IS_REVENUE"]
        if revenue_items and placeholder_rows_before_gross:
            # Deduplicate: only keep one Revenue item (the first one)
            unique_revenue_items = [revenue_items[0]] if revenue_items else []
            replace_line_item_placeholders(
                worksheet, placeholder_rows_before_gross[:len(unique_revenue_items)], 
                unique_revenue_items, matrix, year_column_map
            )
            logger.debug(f"Replaced {len(unique_revenue_items)} Revenue items before Gross Profit")
    
    # Find placeholders between Gross Profit and Operating Profit (COGS, etc.)
    if gross_profit_row and operating_profit_row:
        placeholder_rows_before_operating = find_line_item_placeholders(
            worksheet, placeholder_text, gross_profit_row + 1, operating_profit_row - 1
        )
        cogs_items = [item for item in line_items if item.get("model_role") == "IS_COGS"]
        if cogs_items and placeholder_rows_before_operating:
            unique_cogs_items = [cogs_items[0]] if cogs_items else []
            replace_line_item_placeholders(
                worksheet, placeholder_rows_before_operating[:len(unique_cogs_items)],
                unique_cogs_items, matrix, year_column_map
            )
    
    # Find placeholders between Operating Profit and Income before Taxes (Operating expenses)
    if operating_profit_row and income_before_taxes_row:
        placeholder_rows_before_taxes = find_line_item_placeholders(
            worksheet, placeholder_text, operating_profit_row + 1, income_before_taxes_row - 1
        )
        opex_items = [item for item in line_items if item.get("model_role") == "IS_OPERATING_EXPENSE"]
        if opex_items and placeholder_rows_before_taxes:
            unique_opex_items = [opex_items[0]] if opex_items else []
            replace_line_item_placeholders(
                worksheet, placeholder_rows_before_taxes[:len(unique_opex_items)],
                unique_opex_items, matrix, year_column_map
            )


def populate_balance_sheet_line_items(
    worksheet,
    line_items: List[Dict[str, Any]],
    matrix: Dict[str, Dict[int, float]],
    year_column_map: Dict[int, int],
    section_start_row: int
) -> None:
    """
    Populate Balance Sheet with dynamic line items for Assets, Liabilities, and Equity sections.
    
    Args:
        worksheet: openpyxl Worksheet object
        line_items: List of Balance Sheet line items
        matrix: Role/year matrix
        year_column_map: Dictionary mapping year -> column_index
        section_start_row: Starting row of Balance Sheet section
    """
    structure = STATEMENT_STRUCTURE["balance_sheet"]
    subsections = structure["subsections"]
    
    # Map model roles to subsections
    assets_roles = {"BS_CASH", "BS_INVENTORY", "BS_ACCOUNTS_RECEIVABLE", "BS_PP_AND_E"}
    liabilities_roles = {"BS_ACCOUNTS_PAYABLE", "BS_ACCRUED_LIABILITIES", "BS_DEBT_CURRENT", "BS_DEBT_NON_CURRENT"}
    equity_roles = set()  # Add equity roles if needed
    
    # Helper function to deduplicate items by model_role (keep first occurrence)
    def deduplicate_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        seen_roles = set()
        unique_items = []
        for item in items:
            role = item.get("model_role")
            if role and role not in seen_roles:
                seen_roles.add(role)
                unique_items.append(item)
        return unique_items
    
    # Process Assets section
    assets_header_row = find_statement_section_start(worksheet, subsections["assets"]["header"], section_start_row)
    if assets_header_row:
        assets_items_raw = [item for item in line_items if item.get("model_role") in assets_roles]
        assets_items = deduplicate_items(assets_items_raw)
        anchor_map = find_anchor_rows(worksheet, subsections["assets"]["anchors"], assets_header_row, assets_header_row + 50)
        first_anchor = anchor_map.get("total current assets")
        if first_anchor:
            placeholder_rows = find_line_item_placeholders(worksheet, subsections["assets"]["placeholder_text"], assets_header_row + 1, first_anchor - 1)
            if assets_items and placeholder_rows:
                replace_line_item_placeholders(
                    worksheet, placeholder_rows, assets_items, matrix, year_column_map
                )
                logger.debug(f"Replaced {len(assets_items)} Balance Sheet Assets items")
    
    # Process Liabilities section
    liabilities_header_row = find_statement_section_start(worksheet, subsections["liabilities"]["header"], section_start_row)
    if liabilities_header_row:
        liabilities_items_raw = [item for item in line_items if item.get("model_role") in liabilities_roles]
        liabilities_items = deduplicate_items(liabilities_items_raw)
        anchor_map = find_anchor_rows(worksheet, subsections["liabilities"]["anchors"], liabilities_header_row, liabilities_header_row + 50)
        first_anchor = anchor_map.get("total current liabilities")
        if first_anchor:
            placeholder_rows = find_line_item_placeholders(worksheet, subsections["liabilities"]["placeholder_text"], liabilities_header_row + 1, first_anchor - 1)
            if liabilities_items and placeholder_rows:
                replace_line_item_placeholders(
                    worksheet, placeholder_rows, liabilities_items, matrix, year_column_map
                )
                logger.debug(f"Replaced {len(liabilities_items)} Balance Sheet Liabilities items")
    
    # Process Equity section
    equity_header_row = find_statement_section_start(worksheet, subsections["equity"]["header"], section_start_row)
    if equity_header_row:
        equity_items_raw = [item for item in line_items if item.get("model_role") in equity_roles]
        equity_items = deduplicate_items(equity_items_raw)
        anchor_map = find_anchor_rows(worksheet, subsections["equity"]["anchors"], equity_header_row, equity_header_row + 50)
        first_anchor = anchor_map.get("total equity")
        if first_anchor:
            placeholder_rows = find_line_item_placeholders(worksheet, subsections["equity"]["placeholder_text"], equity_header_row + 1, first_anchor - 1)
            if equity_items and placeholder_rows:
                replace_line_item_placeholders(
                    worksheet, placeholder_rows, equity_items, matrix, year_column_map
                )
                logger.debug(f"Replaced {len(equity_items)} Balance Sheet Equity items")


def populate_cash_flow_line_items(
    worksheet,
    line_items: List[Dict[str, Any]],
    matrix: Dict[str, Dict[int, float]],
    year_column_map: Dict[int, int],
    section_start_row: int
) -> None:
    """
    Populate Cash Flow Statement with dynamic line items for Operating, Investing, and Financing sections.
    
    Args:
        worksheet: openpyxl Worksheet object
        line_items: List of Cash Flow line items
        matrix: Role/year matrix
        year_column_map: Dictionary mapping year -> column_index
        section_start_row: Starting row of Cash Flow Statement section
    """
    structure = STATEMENT_STRUCTURE["cash_flow"]
    subsections = structure["subsections"]
    
    # Map model roles to subsections
    operating_roles = {"CF_DEPRECIATION", "CF_CHANGE_IN_WORKING_CAPITAL"}
    investing_roles = {"CF_CAPEX"}
    financing_roles = {"CF_SHARE_REPURCHASES", "CF_DIVIDENDS_PAID", "CF_DEBT_ISSUANCE", "CF_DEBT_REPAYMENT"}
    
    # Helper function to deduplicate items by model_role (keep first occurrence)
    def deduplicate_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        seen_roles = set()
        unique_items = []
        for item in items:
            role = item.get("model_role")
            if role and role not in seen_roles:
                seen_roles.add(role)
                unique_items.append(item)
        return unique_items
    
    # Process Operating Activities
    operating_header_row = find_statement_section_start(worksheet, subsections["operating"]["header"], section_start_row)
    if operating_header_row:
        operating_items_raw = [item for item in line_items if item.get("model_role") in operating_roles]
        operating_items = deduplicate_items(operating_items_raw)
        logger.debug(f"Found {len(operating_items)} Operating Activities items: {[item.get('label') for item in operating_items]}")
        anchor_map = find_anchor_rows(worksheet, subsections["operating"]["anchors"], operating_header_row, operating_header_row + 50)
        first_anchor = anchor_map.get("net cash provided by/(used in) operating activities")
        if first_anchor:
            placeholder_rows = find_line_item_placeholders(worksheet, subsections["operating"]["placeholder_text"], operating_header_row + 1, first_anchor - 1)
            logger.debug(f"Found {len(placeholder_rows)} placeholder rows for Operating Activities: {placeholder_rows}")
            if operating_items and placeholder_rows:
                replace_line_item_placeholders(
                    worksheet, placeholder_rows, operating_items, matrix, year_column_map
                )
                logger.info(f"Replaced {len(operating_items)} Cash Flow Operating items")
            elif operating_items and not placeholder_rows:
                logger.warning(f"Found {len(operating_items)} Operating Activities items but no placeholder rows to replace")
            elif not operating_items and placeholder_rows:
                logger.warning(f"Found {len(placeholder_rows)} placeholder rows but no Operating Activities items to populate")
    
    # Process Investing Activities
    investing_header_row = find_statement_section_start(worksheet, subsections["investing"]["header"], section_start_row)
    if investing_header_row:
        investing_items_raw = [item for item in line_items if item.get("model_role") in investing_roles]
        investing_items = deduplicate_items(investing_items_raw)
        logger.debug(f"Found {len(investing_items)} Investing Activities items: {[item.get('label') for item in investing_items]}")
        anchor_map = find_anchor_rows(worksheet, subsections["investing"]["anchors"], investing_header_row, investing_header_row + 50)
        first_anchor = anchor_map.get("net cash provided by/(used in) investing activities")
        if first_anchor:
            placeholder_rows = find_line_item_placeholders(worksheet, subsections["investing"]["placeholder_text"], investing_header_row + 1, first_anchor - 1)
            logger.debug(f"Found {len(placeholder_rows)} placeholder rows for Investing Activities: {placeholder_rows}")
            if investing_items and placeholder_rows:
                replace_line_item_placeholders(
                    worksheet, placeholder_rows, investing_items, matrix, year_column_map
                )
                logger.info(f"Replaced {len(investing_items)} Cash Flow Investing items")
            elif investing_items and not placeholder_rows:
                logger.warning(f"Found {len(investing_items)} Investing Activities items but no placeholder rows to replace")
            elif not investing_items and placeholder_rows:
                logger.warning(f"Found {len(placeholder_rows)} placeholder rows but no Investing Activities items to populate")
    
    # Process Financing Activities
    financing_header_row = find_statement_section_start(worksheet, subsections["financing"]["header"], section_start_row)
    if financing_header_row:
        financing_items_raw = [item for item in line_items if item.get("model_role") in financing_roles]
        financing_items = deduplicate_items(financing_items_raw)
        logger.debug(f"Found {len(financing_items)} Financing Activities items: {[item.get('label') for item in financing_items]}")
        anchor_map = find_anchor_rows(worksheet, subsections["financing"]["anchors"], financing_header_row, financing_header_row + 50)
        first_anchor = anchor_map.get("net cash provided by/(used in) financing activities")
        if first_anchor:
            placeholder_rows = find_line_item_placeholders(worksheet, subsections["financing"]["placeholder_text"], financing_header_row + 1, first_anchor - 1)
            logger.debug(f"Found {len(placeholder_rows)} placeholder rows for Financing Activities: {placeholder_rows}")
            if financing_items and placeholder_rows:
                replace_line_item_placeholders(
                    worksheet, placeholder_rows, financing_items, matrix, year_column_map
                )
                logger.info(f"Replaced {len(financing_items)} Cash Flow Financing items")
            elif financing_items and not placeholder_rows:
                logger.warning(f"Found {len(financing_items)} Financing Activities items but no placeholder rows to replace")
            elif not financing_items and placeholder_rows:
                logger.warning(f"Found {len(placeholder_rows)} placeholder rows but no Financing Activities items to populate")


# Constant for scaling values to millions
MILLIONS_SCALE = 1_000_000


def populate_three_statement_historicals(
    worksheet,
    role_column: str,
    matrix: Dict[str, Dict[int, float]],
    year_column_map: Dict[int, int],
    line_items: Optional[List[Dict[str, Any]]] = None
) -> None:
    """
    Populate historical data specifically for 3-statement model worksheet.
    
    This function is tailored for the "3 Statement" sheet and handles:
    - Income Statement items (Revenue, COGS, Operating Expenses, etc.)
    - Balance Sheet items (Cash, PPE, Inventory, Debt, etc.)
    - Cash Flow items (Capex, Depreciation, Working Capital, etc.)
    
    First inserts dynamic line items to replace "Line Item" placeholders,
    then populates historical data for all items.
    
    Skips rows with "%" in the label (these are calculated, not filled).
    Ensures proper number formatting to prevent date interpretation.
    
    Args:
        worksheet: openpyxl Worksheet object for "3 Statement" sheet
        role_column: Column letter containing model_role values (typically "ZZ")
        matrix: Role/year matrix from build_role_year_matrix()
        year_column_map: Dictionary mapping year -> column_index
        line_items: Optional list of line items from JSON for dynamic insertion
    """
    col_idx = column_index_from_string(role_column)
    
    # Step 1: Insert dynamic line items if provided
    if line_items:
        logger.info("Inserting dynamic line items for 3-statement model")
        grouped_items = group_line_items_by_statement(line_items)
        
        # Find statement section starts
        income_start = find_statement_section_start(worksheet, "Income Statement")
        balance_start = find_statement_section_start(worksheet, "Balance Sheet")
        cash_flow_start = find_statement_section_start(worksheet, "Cash Flow Statement")
        
        # Populate Income Statement line items
        if income_start and grouped_items["income_statement"]:
            populate_income_statement_line_items(
                worksheet, grouped_items["income_statement"], matrix, year_column_map, income_start
            )
        
        # Populate Balance Sheet line items
        if balance_start and grouped_items["balance_sheet"]:
            populate_balance_sheet_line_items(
                worksheet, grouped_items["balance_sheet"], matrix, year_column_map, balance_start
            )
        
        # Populate Cash Flow line items
        if cash_flow_start and grouped_items["cash_flow"]:
            populate_cash_flow_line_items(
                worksheet, grouped_items["cash_flow"], matrix, year_column_map, cash_flow_start
            )
        
        # Re-add model_role column after row insertions
        add_model_role_column(worksheet, role_column=role_column)
    
    # Step 2: Populate historical data for all items (including inserted ones)
    # Key model roles for 3-statement model organized by statement type
    income_statement_roles = {
        "IS_REVENUE",
        "IS_COGS",
        "IS_GROSS_PROFIT",
        "IS_OPERATING_EXPENSE",
        "IS_NON_OPERATING_INCOME",  # For netInterestIncome
        "IS_INTEREST_EXPENSE",
        "IS_OPERATING_INCOME",  # Also used for incomeBeforeTax
        "IS_TAX_EXPENSE",
        "IS_NET_INCOME",  # Also used for EPS
    }
    
    balance_sheet_roles = {
        "BS_CASH",
        "BS_PP_AND_E",
        "BS_INVENTORY",
        "BS_ACCOUNTS_RECEIVABLE",
        "BS_ACCOUNTS_PAYABLE",
        "BS_ACCRUED_LIABILITIES",
        "BS_DEBT_CURRENT",
        "BS_DEBT_NON_CURRENT",
    }
    
    cash_flow_roles = {
        "CF_CAPEX",
        "CF_DEPRECIATION",
        "CF_CHANGE_IN_WORKING_CAPITAL",
        "CF_SHARE_REPURCHASES",
        "CF_DIVIDENDS_PAID",
        "CF_DEBT_ISSUANCE",
        "CF_DEBT_REPAYMENT",
    }
    
    # Track which roles we've populated for logging
    populated_roles = {
        "income_statement": [],
        "balance_sheet": [],
        "cash_flow": [],
    }
    
    # Iterate through all rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False), start=1):
        role_cell = row[0]
        model_role = (role_cell.value or "").strip()
        
        if not model_role:
            continue
        
        # Skip rows with "%" in the label (these are calculated, not filled)
        # Check the row label in column B
        label_cell = worksheet.cell(row=row_idx, column=2)
        label_value = str(label_cell.value or "").strip()
        if "%" in label_value:
            continue
        
        # Skip "Cash Flow Statement" title row - it should not have any values written to it
        if label_value.lower() == "cash flow statement":
            continue
        
        # Skip Cash Flow section header rows - they should not have any values written to them
        # Only year headers should be overwritten in these rows
        cash_flow_section_headers = [
            "cash flows from operating activities",
            "cash flows from investing activities",
            "cash flows from financing activities"
        ]
        if any(header in label_value.lower() for header in cash_flow_section_headers):
            continue
        
        # Determine which statement type this role belongs to
        statement_type = None
        if model_role in income_statement_roles:
            statement_type = "income_statement"
        elif model_role in balance_sheet_roles:
            statement_type = "balance_sheet"
        elif model_role in cash_flow_roles:
            statement_type = "cash_flow"
        else:
            # Unknown role, skip it (or could log a warning)
            continue

        # Look up values for this model_role
        role_values = matrix.get(model_role, {})
        
        if not role_values:
            continue
        
        # Write values to historical columns with proper formatting
        # Include zero values as they're meaningful in financial statements
        for year, column_idx in year_column_map.items():
            if year in role_values:
                value = role_values[year]
                target_cell = worksheet.cell(row=row_idx, column=column_idx)
                # Write value even if it's 0 (zero values are meaningful)
                # Scale to millions
                if isinstance(value, (int, float)):
                    target_cell.value = value / MILLIONS_SCALE
                else:
                    target_cell.value = value
                
                # Set number format to prevent date interpretation
                # Use General format for all numeric values
                if isinstance(value, (int, float)):
                    target_cell.number_format = "General"
        
        # Track populated role
        if model_role not in populated_roles[statement_type]:
            populated_roles[statement_type].append(model_role)
    
    # Step 3: Write income statement fields to fixed rows (bypassing model_role matching)
    # These rows have fixed positions in the Excel template
    income_start = find_statement_section_start(worksheet, "Income Statement", 1)
    if income_start and line_items:
        # Fixed row mappings for income statement fields
        # Row numbers are absolute Excel row numbers
        # Field name (from JSON) -> row number
        fixed_income_rows = {
            "otherExpenses": 11,  # Other Expenses row
            "operatingIncome": 12,  # Operating Profit row
            "nonOperatingIncomeExcludingInterest": 39,  # Other Non Operating Adjustments row
            "netInterestIncome": 14,  # Net Interest Income row
            "incomeBeforeTax": 16,  # Income Before Taxes row
            "incomeTaxExpense": 17,  # Income Tax Expense row
            "eps": 20,  # EPS row (use epsDiluted if eps not available)
        }
        
        # Group line items by statement type
        grouped_items = group_line_items_by_statement(line_items)
        income_items = grouped_items.get("income_statement", [])
        
        # Write values to fixed rows
        for field_name, target_row in fixed_income_rows.items():
            # Find the line item by tag (field name)
            matching_item = None
            for item in income_items:
                if item.get("tag") == field_name:
                    matching_item = item
                    break
            
            # For EPS, try epsDiluted if eps not found
            if not matching_item and field_name == "eps":
                for item in income_items:
                    if item.get("tag") == "epsDiluted":
                        matching_item = item
                        break
            
            # Get periods from the line item
            periods = matching_item.get("periods", {}) if matching_item else {}
            
            # Write values to columns D, E, F (columns 4, 5, 6)
            # Include zero values as they're meaningful in financial statements
            # For otherExpenses specifically, ensure we write even if value is 0
            if periods:
                for period_date, value in periods.items():
                    try:
                        # Extract year from period date
                        year = int(str(period_date).split("-")[0])
                        if year in year_column_map:
                            column_idx = year_column_map[year]
                            target_cell = worksheet.cell(row=target_row, column=column_idx)
                            # Write value even if it's 0 (zero values are meaningful)
                            # Scale to millions
                            # Explicitly handle 0 values - they should be written
                            if value == 0 or value == 0.0:
                                target_cell.value = 0.0
                            else:
                                target_cell.value = float(value) / MILLIONS_SCALE
                            
                            # Set number format
                            target_cell.number_format = "General"
                            logger.debug(f"Wrote {field_name} = {value} to row {target_row}, column {column_idx}")
                    except (ValueError, IndexError, TypeError) as e:
                        logger.debug(f"Error processing {field_name} period {period_date}: {e}")
                        continue
                
                logger.debug(f"Wrote {field_name} to row {target_row}")
            elif matching_item:
                # If we found the item but no periods, log a warning
                # This shouldn't happen if the conversion is working correctly
                logger.warning(f"Found {field_name} line item but no periods data - this may indicate a conversion issue")
            else:
                # For otherExpenses, log more detail since it's required to be written even if 0
                if field_name == "otherExpenses":
                    logger.warning(f"otherExpenses not found in income statement line items. Available tags: {[item.get('tag') for item in income_items[:10]]}")
                logger.debug(f"No line item found for {field_name}")
    
    # Step 4: Write balance sheet fields to fixed rows (bypassing model_role matching)
    # These rows have fixed positions in the Excel template
    balance_start = find_statement_section_start(worksheet, "Balance Sheet", 1)
    if balance_start and line_items:
        # Fixed row mappings for balance sheet fields
        # Row numbers are absolute Excel row numbers
        # Field name (from JSON) -> row number
        fixed_balance_rows = {
            # Assets section
            "cashAndCashEquivalents": 25,  # Cash & Cash Equivalents
            "shortTermInvestments": 26,  # Short-Term Investments
            "netReceivables": 27,  # Net Receivables
            "inventory": 28,  # Inventory
            "prepaids": 29,  # Prepaid Expenses
            "otherCurrentAssets": 30,  # Other Current Assets
            "totalCurrentAssets": 31,  # Total Current Assets
            "propertyPlantEquipmentNet": 32,  # Property, Plant & Equipment (Net)
            "goodwill": 33,  # Goodwill
            "intangibleAssets": 34,  # Intangible Assets
            "longTermInvestments": 35,  # Long-Term Investments
            "taxAssets": 36,  # Tax Assets
            "otherNonCurrentAssets": 37,  # Other Non-Current Assets
            "totalAssets": 40,  # Total Assets
            # Liabilities section
            "accountsPayable": 42,  # Accounts Payable
            "accountPayables": 42,  # Accounts Payable (alternative field name)
            "otherPayables": 43,  # Other Payables
            "accruedExpenses": 44,  # Accrued Expenses
            "shortTermDebt": 45,  # Short-Term Debt
            "capitalLeaseObligationsCurrent": 46,  # Current Capital Lease Obligations
            "currentCapitalLeaseObligations": 46,  # Current Capital Lease Obligations (alternative)
            "taxPayables": 47,  # Tax Payables
            "deferredRevenue": 48,  # Deferred Revenue (Current)
            "otherCurrentLiabilities": 49,  # Other Current Liabilities
            "totalCurrentLiabilities": 50,  # Total Current Liabilities
            "longTermDebt": 51,  # Long-Term Debt
            "capitalLeaseObligationsNonCurrent": 52,  # Non-Current Capital Lease Obligations
            "nonCurrentCapitalLeaseObligations": 52,  # Non-Current Capital Lease Obligations (alternative)
            "deferredRevenueNonCurrent": 53,  # Deferred Revenue (Non-Current)
            "deferredTaxLiabilitiesNonCurrent": 54,  # Deferred Tax Liabilities (Non-Current)
            "otherNonCurrentLiabilities": 55,  # Other Non-Current Liabilities
            "totalNonCurrentLiabilities": 56,  # Total Non-Current Liabilities
            "totalLiabilities": 57,  # Total Liabilities
            # Equity section
            "preferredStock": 59,  # Preferred Stock
            "treasuryStock": 60,  # Treasury Stock
            "commonStock": 61,  # Common Stock
            "retainedEarnings": 62,  # Retained Earnings
            "additionalPaidInCapital": 63,  # Additional Paid-In Capital
            "accumulatedOtherComprehensiveIncomeLoss": 64,  # Accumulated Other Comprehensive Income (Loss)
            "otherStockholdersEquity": 65,  # Other Stockholders' Equity
            "otherTotalStockholdersEquity": 65,  # Other Stockholders' Equity (alternative field name)
            "totalEquity": 66,  # Total Equity
            "totalStockholdersEquity": 66,  # Total Equity (alternative field name)
            # Note: Row 67 is "Total Liabilities & Equity" which is computed, so we don't map it
        }
        
        # Group line items by statement type
        grouped_items = group_line_items_by_statement(line_items)
        balance_items = grouped_items.get("balance_sheet", [])
        
        # Write values to fixed rows
        for field_name, target_row in fixed_balance_rows.items():
            # Find the line item by tag (field name)
            matching_item = None
            for item in balance_items:
                if item.get("tag") == field_name:
                    matching_item = item
                    break
            
            # Get periods from the line item (or empty dict if not found)
            periods = matching_item.get("periods", {}) if matching_item else {}
            
            # Write values to columns D, E, F (columns 4, 5, 6)
            # Include zero values as they're meaningful in financial statements
            # For balance sheet fields, ensure we write even if value is 0
            if periods:
                for period_date, value in periods.items():
                    try:
                        # Extract year from period date
                        year = int(str(period_date).split("-")[0])
                        if year in year_column_map:
                            column_idx = year_column_map[year]
                            target_cell = worksheet.cell(row=target_row, column=column_idx)
                            # Write value even if it's 0 (zero values are meaningful)
                            # Scale to millions
                            # Explicitly handle 0 values - they should be written
                            if value == 0 or value == 0.0:
                                target_cell.value = 0.0
                            else:
                                target_cell.value = float(value) / MILLIONS_SCALE
                            
                            # Set number format
                            target_cell.number_format = "General"
                            logger.debug(f"Wrote balance sheet {field_name} = {value} to row {target_row}, column {column_idx}")
                    except (ValueError, IndexError, TypeError) as e:
                        logger.debug(f"Error processing balance sheet {field_name} period {period_date}: {e}")
                        continue
                
                logger.debug(f"Wrote balance sheet {field_name} to row {target_row}")
            elif matching_item:
                # If we found the item but no periods, log a warning
                # This shouldn't happen if the conversion is working correctly
                logger.warning(f"Found balance sheet {field_name} line item but no periods data - this may indicate a conversion issue")
            else:
                logger.debug(f"No line item found for balance sheet field {field_name}")
    
    # Step 5: Write cash flow statement fields to fixed rows (bypassing model_role matching)
    # These rows have fixed positions in the Excel template
    cash_flow_start = find_statement_section_start(worksheet, "Cash Flow Statement", 1)
    if cash_flow_start and line_items:
        # Fixed row mappings for cash flow statement fields
        # Row numbers are absolute Excel row numbers
        # Field name (from JSON) -> row number
        fixed_cash_flow_rows = {
            # Operating Activities section
            "netIncome": 70,  # Net Income
            "depreciationAndAmortization": 71,  # Depreciation & Amortization
            "deferredIncomeTax": 72,  # Deferred Income Tax
            "stockBasedCompensation": 73,  # Stock-Based Compensation
            "changeInWorkingCapital": 74,  # Change in Working Capital
            "otherNonCashItems": 75,  # Other Non-Cash Items
            "netCashProvidedByOperatingActivities": 76,  # Net Cash Provided by Operating Activities
            # Investing Activities section
            "investmentsInPropertyPlantAndEquipment": 78,  # Investments in Property, Plant & Equipment (CapEx)
            "salesMaturitiesOfInvestments": 81,  # Sales / Maturities of Investments
            # Financing Activities section
            "netDebtIssuance": 85,  # Net Debt Issuance
            "netStockIssuance": 86,  # Net Stock Issuance
            "netCommonStockIssuance": 87,  # Net Common Stock Issuance
            "commonStockIssuance": 88,  # Common Stock Issued
            "commonStockRepurchased": 89,  # Common Stock Repurchased
            "netPreferredStockIssuance": 90,  # Net Preferred Stock Issuance
            "dividendsPaid": 91,  # Dividends Paid (Net)
            "commonDividendsPaid": 92,  # Common Dividends Paid
            "preferredDividendsPaid": 93,  # Preferred Dividends Paid
            "otherFinancingActivities": 94,  # Other Financing Activities
            "netCashProvidedByFinancingActivities": 95,  # Net Cash Provided by Financing Activities
            # Other items
            "effectOfForexChangesOnCash": 97,  # Effect of Foreign Exchange on Cash
            "cashAtBeginningOfPeriod": 99,  # Cash at Beginning of Period
            "netChangeInCash": 100,  # Net Change in Cash
            "cashAtEndOfPeriod": 101,  # Cash at End of Period
            "freeCashFlow": 103,  # Free Cash Flow (renamed from Unlevered Free Cashflow)
        }
        
        # Group line items by statement type
        grouped_items = group_line_items_by_statement(line_items)
        cash_flow_items = grouped_items.get("cash_flow", [])
        
        # Write values to fixed rows
        for field_name, target_row in fixed_cash_flow_rows.items():
            # Find the line item by tag (field name)
            matching_item = None
            for item in cash_flow_items:
                if item.get("tag") == field_name:
                    matching_item = item
                    break
            
            if not matching_item:
                logger.debug(f"No line item found for cash flow field {field_name}")
                continue
            
            # Get periods from the line item
            periods = matching_item.get("periods", {})
            if not periods:
                continue
            
            # Write values to columns D, E, F (columns 4, 5, 6)
            for period_date, value in periods.items():
                try:
                    # Extract year from period date
                    year = int(str(period_date).split("-")[0])
                    if year in year_column_map:
                        column_idx = year_column_map[year]
                        target_cell = worksheet.cell(row=target_row, column=column_idx)
                        # Scale to millions
                        target_cell.value = float(value) / MILLIONS_SCALE
                        
                        # Set number format
                        target_cell.number_format = "General"
                except (ValueError, IndexError, TypeError) as e:
                    logger.debug(f"Error processing cash flow {field_name} period {period_date}: {e}")
                    continue
            
            logger.debug(f"Wrote cash flow {field_name} to row {target_row}")
    
    # Log summary
    total_populated = sum(len(roles) for roles in populated_roles.values())
    logger.info(f"Populated 3-statement historicals: {total_populated} roles")
    logger.debug(f"  Income Statement: {len(populated_roles['income_statement'])} roles")
    logger.debug(f"  Balance Sheet: {len(populated_roles['balance_sheet'])} roles")
    logger.debug(f"  Cash Flow: {len(populated_roles['cash_flow'])} roles")


def extract_years_from_fmp_json(json_data: Dict[str, Any]) -> List[int]:
    """
    Extract fiscal years from FMP JSON data structure.
    
    FMP data can have years in multiple formats:
    - income_statements/balance_sheets/cash_flow_statements arrays with 'date', 'calendarYear', 'fiscalYear' fields
    - Each statement record may have different year fields
    
    Args:
        json_data: FMP JSON data dictionary
        
    Returns:
        List of unique fiscal years (integers), sorted ascending
    """
    years_set = set()
    
    # Check for FMP format arrays
    for statement_type in ["income_statements", "balance_sheets", "cash_flow_statements"]:
        statements = json_data.get(statement_type, [])
        if not isinstance(statements, list):
            continue
            
        for stmt in statements:
            if not isinstance(stmt, dict):
                continue
                
            # Try multiple year fields in order of preference
            year = None
            
            # 1. Try fiscalYear (most reliable for FMP)
            if "fiscalYear" in stmt and stmt["fiscalYear"]:
                try:
                    year = int(stmt["fiscalYear"])
                except (ValueError, TypeError):
                    pass
            
            # 2. Try calendarYear
            if year is None and "calendarYear" in stmt and stmt["calendarYear"]:
                try:
                    year = int(stmt["calendarYear"])
                except (ValueError, TypeError):
                    pass
            
            # 3. Try extracting from date field (YYYY-MM-DD format)
            if year is None and "date" in stmt and stmt["date"]:
                date_str = str(stmt["date"])
                if len(date_str) >= 4:
                    try:
                        year = int(date_str[:4])
                    except (ValueError, TypeError):
                        pass
            
            if year and 1900 <= year <= 2100:  # Sanity check
                years_set.add(year)
    
    years_list = sorted(list(years_set))
    logger.info(f"Extracted {len(years_list)} years from FMP JSON: {years_list}")
    return years_list


def populate_template_from_json(
    json_path: str,
    template_path: str,
    output_path: str,
    target_sheets: Optional[List[str]] = None,
    assumptions_path: Optional[str] = None
) -> None:
    """
    Main orchestration function to populate Excel template with historical data from JSON.
    
    Supports both structured JSON format and FMP raw JSON format.
    
    Args:
        json_path: Path to structured JSON file or FMP raw JSON file
        template_path: Path to Excel template file
        output_path: Path where populated template will be saved
        target_sheets: List of sheet names to populate (default: ["DCF Base", "3 Statement"])
        assumptions_path: Optional path to assumptions JSON file for 3-statement model
    """
    if target_sheets is None:
        target_sheets = ["DCF Base"]
    
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for reading Excel templates. Install with: pip install openpyxl")
    
    # Step 1: Load JSON
    logger.info(f"Loading JSON from {json_path}")
    json_data = load_structured_json(json_path)
    
    # Step 1.5: Detect FMP format and extract years if needed
    is_fmp_format = (
        "income_statements" in json_data or 
        "balance_sheets" in json_data or 
        "cash_flow_statements" in json_data
    )
    
    if is_fmp_format:
        logger.info("Detected FMP JSON format, extracting years from FMP data")
        years_list = extract_years_from_fmp_json(json_data)
        
        # For FMP format, we still need to build a matrix for compatibility
        # But we'll use the years extracted directly from FMP
        logger.info("Extracting line items from FMP JSON")
        line_items = extract_line_items_from_json(json_data)
        logger.info(f"Extracted {len(line_items)} line items")
        
        logger.info("Building role/year matrix")
        matrix = build_role_year_matrix(line_items)
        logger.info(f"Built matrix with {len(matrix)} model roles")
        
        # Use years from FMP extraction, but merge with any years found in matrix
        matrix_years = set()
        for role_values in matrix.values():
            matrix_years.update(role_values.keys())
        
        # Merge years: prefer FMP extracted years, but include matrix years as fallback
        if years_list:
            all_years = set(years_list) | matrix_years
        else:
            all_years = matrix_years
            years_list = sorted(list(all_years))
            logger.warning("No years found in FMP data, using years from matrix")
    else:
        # Standard structured JSON format
        logger.info("Detected structured JSON format")
    logger.info("Extracting line items from JSON")
    line_items = extract_line_items_from_json(json_data)
    logger.info(f"Extracted {len(line_items)} line items")
    
    logger.info("Building role/year matrix")
    matrix = build_role_year_matrix(line_items)
    logger.info(f"Built matrix with {len(matrix)} model roles")
    
    # Get all years from matrix
    all_years = set()
    for role_values in matrix.values():
        all_years.update(role_values.keys())
    years_list = sorted(all_years)
    
    logger.info(f"Found years: {years_list}")
    
    # Step 2: Copy template to output path first (preserves original template as source of truth)
    # This also avoids permission errors if the template is open in Excel
    if template_path != output_path:
        logger.info(f"Copying template from {template_path} to {output_path}")
        shutil.copy2(template_path, output_path)
        template_path = output_path  # Work on the copy
    
    # Step 3: Load Excel template (now working on copy)
    logger.info(f"Loading Excel template from {template_path}")
    workbook = load_excel_template(template_path)
    
    # Step 3.5: Populate cover sheet and summary sheet
    # Extract company info - handle both structured JSON and FMP formats
    company_info = json_data.get("company", {})
    company_name = company_info.get("company_name", "") or company_info.get("name", "")
    ticker = company_info.get("ticker", "") or company_info.get("symbol", "")
    
    # For FMP format, try to get company name from first income statement if available
    if is_fmp_format and not company_name:
        income_statements = json_data.get("income_statements", [])
        if income_statements and isinstance(income_statements, list) and len(income_statements) > 0:
            company_name = income_statements[0].get("companyName", "")
    
    # Clean up company name and ticker
    company_name = str(company_name).strip() if company_name else ""
    ticker = str(ticker).strip().upper() if ticker else ""
    
    # Extract quote data if available (for FMP format)
    quote_data = None
    if is_fmp_format:
        quote_data = json_data.get("quote") or json_data.get("quote_data")
        # If quote is a file path string, try to load it
        if isinstance(quote_data, str) and Path(quote_data).exists():
            try:
                with open(quote_data, "r", encoding="utf-8") as f:
                    quote_data = json.load(f)
            except Exception as e:
                logger.warning(f"Failed to load quote data from {quote_data}: {e}")
                quote_data = None
    
    if company_name and ticker:
        logger.info(f"Populating cover sheet and summary sheet: {company_name} ({ticker})")
        # Use fake pricing for dummy/test data
        use_fake_pricing = ticker.upper() in ["DUMMY", "TEST", "FAKE"]
        populate_cover_sheet(
            workbook, 
            company_name, 
            ticker, 
            use_fake_pricing=use_fake_pricing,
            quote_data=quote_data
        )
    else:
        logger.warning(f"Missing company info (name={company_name}, ticker={ticker}), skipping cover sheet")
    
    # Step 3.6: Populate DCF year headers for all DCF sheets
    logger.info("Populating DCF year headers")
    populate_dcf_year_headers(workbook, years_list, forecast_periods=5)
    
    # Step 3.7: Load assumptions if provided
    assumptions_data = None
    if assumptions_path:
        logger.info(f"Loading assumptions from {assumptions_path}")
        try:
            raw_assumptions = load_structured_json(assumptions_path)
            assumptions_data = convert_assumptions_to_legacy_format(raw_assumptions)
            logger.info("Assumptions loaded and converted successfully")
        except Exception as e:
            logger.warning(f"Failed to load assumptions: {e}")
    
    # Step 4: Process each target sheet
    for sheet_name in target_sheets:
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template. Skipping.")
            continue
        
        logger.info(f"Processing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        
        # Add model_role column
        add_model_role_column(worksheet, role_column="ZZ")
        
        # For 3 Statement sheet, detect Year"A" placeholders like DCF sheets
        if sheet_name == "3 Statement":
            logger.info("Processing 3 Statement sheet")
            
            # Set sheet title: "Company Name 3 Statement Model (UNITS)"
            if company_name:
                title_cell = worksheet.cell(row=2, column=2)  # Cell B2
                title_cell.value = f"{company_name} 3 Statement Model (UNITS)"
                title_cell.font = Font(bold=True, size=14)
                logger.info(f"Set sheet title to: {company_name} 3 Statement Model (UNITS)")
            
            # Detect historical columns from Income Statement row 4 (where Year"A" placeholders are)
            # Check row 4 for Year"A" placeholders (Income Statement header row)
            hist_columns = []
            for col_idx in range(4, 12):  # Columns D through K
                cell = worksheet.cell(row=4, column=col_idx)
                cell_value = str(cell.value or "").strip().lower()
                # Check for Year"A" patterns
                if "year" in cell_value and ("\"a\"" in cell_value or "'a'" in cell_value or " a" in cell_value):
                    hist_columns.append(col_idx)
            
            # If not found in row 4, try row 16 as fallback
            if not hist_columns:
                hist_columns = detect_historical_columns(worksheet)
            
            # If still not found, use default columns D, E, F
            if not hist_columns:
                logger.info("No Year\"A\" placeholders found in 3 Statement sheet, using default columns D, E, F")
                hist_columns = [4, 5, 6]
            
            logger.info(f"Using historical columns for 3 Statement: {hist_columns}")
            
            # Map years to columns
            year_column_map = map_years_to_columns(years_list, hist_columns)
            logger.info(f"Mapped years to columns: {year_column_map}")
            
            # Populate year headers for all three statements
            # Income Statement (row 4), Balance Sheet (row 24), Cash Flow (row 69)
            # This function handles Year"A" and Year"E" placeholders
            logger.info("Populating 3-statement year headers")
            populate_three_statement_year_headers(worksheet, years_list, forecast_periods=5)
        else:
            # For other sheets (e.g., DCF Base), detect historical columns from row 16
            hist_columns = detect_historical_columns(worksheet)
            if not hist_columns:
                logger.warning(f"No 'Year\"A\"' columns found in sheet '{sheet_name}'. Skipping.")
                continue
        
        logger.info(f"Found {len(hist_columns)} historical columns: {hist_columns}")
        
        # Map years to columns
        year_column_map = map_years_to_columns(years_list, hist_columns)
        logger.info(f"Mapped years to columns: {year_column_map}")
        
        # For 3 Statement sheet, year headers are already populated by populate_three_statement_year_headers
        # For other sheets (DCF Base), write year headers to row 16
        if sheet_name != "3 Statement":
            write_year_headers(worksheet, year_column_map, header_row=16)
            logger.info(f"Wrote year headers to row 16: {sorted(year_column_map.keys())}")
        
        # Populate historicals - use dedicated function for 3 Statement sheet
        if sheet_name == "3 Statement":
            # Ensure line item replacement doesn't affect row 16 (year headers)
            populate_three_statement_historicals(
                worksheet, 
                role_column="ZZ", 
                matrix=matrix, 
                year_column_map=year_column_map,
                line_items=line_items
            )
            logger.info(f"Populated 3-statement historicals for sheet '{sheet_name}'")
            
            # Populate assumptions for 3 Statement sheet if assumptions provided
            # This happens AFTER line items to ensure assumptions are anchored correctly
            if assumptions_data:
                logger.info("Populating 3-statement assumptions")
                populate_three_statement_assumptions(worksheet, assumptions_data, forecast_periods=5)
        elif sheet_name in ["DCF Base", "DCF Bear", "DCF Bull"]:
            # Set DCF sheet title (B2): "Company Name DCF Base" (or Bear/Bull)
            if company_name:
                title_cell = worksheet.cell(row=2, column=2)  # Cell B2
                scenario_name = sheet_name.replace("DCF ", "")  # "Base", "Bear", or "Bull"
                title_cell.value = f"{company_name} DCF {scenario_name}"
                title_cell.font = Font(bold=True, size=14)
                logger.info(f"Set DCF sheet title to: {company_name} DCF {scenario_name}")
            
            # DCF sheets should NOT have historical data populated directly
            # They reference the "3 Statement" sheet via formulas (e.g., D17: "='3 Statement'!D5")
            # Populating historicals here would overwrite those formulas with raw values
            logger.info(f"Skipping historical population for DCF sheet '{sheet_name}' - it uses formulas referencing '3 Statement' sheet")
            
            # Then populate DCF assumptions (WACC, MRP, Risk Free Rate, etc.) if assumptions provided
            if assumptions_data:
                logger.info(f"Populating DCF assumptions for {sheet_name} sheet")
                try:
                    # Build CompanyModelInput from json_data for populate_wacc_and_assumptions
                    model_input = build_company_model_input(json_data)
                    
                    # Create a minimal DcfOutput (populate_wacc_and_assumptions needs it but may not use all fields)
                    from app.services.modeling.types import DcfOutput
                    dcf_output = DcfOutput(
                        enterprise_value=0.0,
                        equity_value=0.0,
                        implied_share_price=0.0,
                        wacc=float(assumptions_data.get("wacc", 0.10)),
                        terminal_growth_rate=float(assumptions_data.get("terminal_growth_rate", assumptions_data.get("dcf", {}).get("terminalGrowthRate", 0.025))),
                        yearly_results=[],
                        terminal_value=0.0,
                        pv_terminal_value=0.0
                    )
                    
                    populate_wacc_and_assumptions(
                        worksheet,
                        model_input,
                        dcf_output,
                        assumptions_data,
                        quote_data=quote_data,
                        json_data=json_data
                    )
                    logger.info(f"Populated DCF assumptions for {sheet_name} sheet")
                    
                    # Populate DCF Base forecast formulas (Revenue Growth, GP Margin, EBIT Margin, etc.)
                    if sheet_name == "DCF Base":
                        logger.info("Populating DCF Base forecast formulas")
                        try:
                            from app.services.modeling.excel_dcf_scenarios import (
                                build_scenario_formula_dict,
                                apply_scenario_to_sheet
                            )
                            
                            # Extract base assumptions from assumptions_data
                            # Base assumptions are in the main assumptions dict, not in a scenario sub-dict
                            base_assumptions = {}
                            
                            # Map income statement assumptions
                            if "assumptions" in assumptions_data:
                                is_assumptions = assumptions_data["assumptions"]
                                
                                # Revenue growth
                                if "revenue" in is_assumptions:
                                    base_assumptions["revenue_growth"] = is_assumptions["revenue"]
                                
                                # GP margin (gross_margin)
                                if "gross_margin" in is_assumptions:
                                    base_assumptions["gp_margin"] = is_assumptions["gross_margin"]
                                
                                # EBIT margin (operating_margin)
                                if "operating_margin" in is_assumptions:
                                    base_assumptions["ebit_margin"] = is_assumptions["operating_margin"]
                            
                            # Map balance sheet assumptions
                            if "balance_sheet" in assumptions_data:
                                bs_assumptions = assumptions_data["balance_sheet"]
                                
                                # D&A (% of Revenue) - depreciation_ppe
                                if "depreciation_ppe" in bs_assumptions:
                                    base_assumptions["da_pct_rev"] = bs_assumptions["depreciation_ppe"]
                            
                            # Map cash flow assumptions
                            if "cash_flow" in assumptions_data:
                                cf_assumptions = assumptions_data["cash_flow"]
                                
                                # CapEx (% of Revenue or % of D&A)
                                if "capex" in cf_assumptions:
                                    capex_assumption = cf_assumptions["capex"]
                                    if isinstance(capex_assumption, dict):
                                        # Check if it's % of revenue or % of D&A based on method
                                        method = capex_assumption.get("method", "").lower()
                                        if method == "revenue":
                                            base_assumptions["capex_pct_rev"] = capex_assumption
                                        elif method == "depreciation":
                                            base_assumptions["capex_pct_da"] = capex_assumption
                                
                                # Changes in WC
                                if "change_in_working_capital" in cf_assumptions:
                                    base_assumptions["wc_change"] = cf_assumptions["change_in_working_capital"]
                            
                            # Apply formulas if we have any base assumptions
                            if base_assumptions:
                                formula_dict = build_scenario_formula_dict()
                                apply_scenario_to_sheet(worksheet, base_assumptions, formula_dict)
                                logger.info(f"Applied DCF Base forecast formulas for {len(base_assumptions)} line items")
                            else:
                                logger.warning("No base assumptions found for DCF Base forecast formulas")
                        except Exception as e:
                            logger.warning(f"Failed to populate DCF Base forecast formulas: {e}")
                            import traceback
                            traceback.print_exc()
                except Exception as e:
                    logger.warning(f"Failed to populate DCF assumptions for {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
        else:
            # Use generic historical population for other sheets
            populate_historicals(worksheet, role_column="ZZ", matrix=matrix, year_column_map=year_column_map)
            logger.info(f"Populated historicals for sheet '{sheet_name}'")
    
    # Step 4.5: Populate RV sheet if it exists
    if "RV" in workbook.sheetnames:
        logger.info("Processing RV sheet")
        worksheet = workbook["RV"]
        populate_rv_sheet(worksheet, json_data, company_name, is_fmp_format, quote_data, assumptions=assumptions_data)
        logger.info("Populated RV sheet")
    
    # Step 4.6: Apply DCF Bear/Bull scenario formulas (if scenario assumptions are present)
    if assumptions_data:
        bear_scenario = assumptions_data.get("bearScenario")
        if bear_scenario and "DCF Bear" in workbook.sheetnames:
            logger.info("Applying DCF Bear scenario formulas")
            populate_dcf_scenario(workbook, bear_scenario, "DCF Bear")
        
        bull_scenario = assumptions_data.get("bullScenario")
        if bull_scenario and "DCF Bull" in workbook.sheetnames:
            logger.info("Applying DCF Bull scenario formulas")
            populate_dcf_scenario(workbook, bull_scenario, "DCF Bull")
    
    # Step 5: Save workbook
    logger.info(f"Saving populated template to {output_path}")
    workbook.save(output_path)
    logger.info("Template population complete")


def export_full_model_to_excel(
    json_path: str,
    template_path: str,
    output_path: str,
    comps_input: Optional[List[Dict[str, Any]]] = None,
    assumptions_path: Optional[str] = None,
) -> None:
    """
    Main orchestration function to populate Excel template with historical data from JSON.
    
    Loads structured JSON, extracts historical financial data, and writes to Excel template.
    Excel formulas handle all forecasting/calculations - this function only provides historical inputs.
    
    Args:
        json_path: Path to structured JSON file
        template_path: Path to Excel template file
        output_path: Path where populated template will be saved
        comps_input: Optional list of comparable company data for RV sheet (deprecated, use assumptions)
            Each dict should have: name, ev_ebitda, pe, ev_sales (or None)
        assumptions_path: Optional path to assumptions JSON file containing competitors and other assumptions
            Expected structure: {"competitors": ["TICKER1", "TICKER2", "TICKER3", "TICKER4"], ...}
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for reading Excel templates. Install with: pip install openpyxl")
    
    # Step 1: Load JSON and build CompanyModelInput
    logger.info(f"Loading JSON from {json_path}")
    json_data = load_structured_json(json_path)
    
    logger.info("Building CompanyModelInput")
    model_input = build_company_model_input(json_data)
    logger.info(f"Built model input for {model_input.ticker} ({model_input.name})")
    
    # Step 1.5: Run modeling modules (for programmatic access, but Excel handles forecasting)
    # Default assumptions - can be made configurable later
    default_assumptions = {
        "revenue_growth": 0.05,
        "operating_margin_target": 0.15,
        "tax_rate": 0.21,
        "capex_as_pct_revenue": 0.05,
        "depreciation_as_pct_revenue": 0.03,
        "wacc": 0.10,
        "terminal_growth_rate": 0.025,
        # WACC components
        "beta": 1.2,
        "market_risk_premium": 0.06,
        "risk_free_rate": 0.04,
        "cost_of_debt": 0.05,
        # Other assumptions
        "shares_outstanding": None,  # Can be calculated or provided
        "debt": None,  # Can be extracted from historicals or provided
        "current_price": None,  # Market data, optional
    }
    
    logger.info("Running 3-statement model")
    three_stmt_output = run_three_statement(
        model_input=model_input,
        assumptions=default_assumptions,
        forecast_periods=3,
    )
    
    logger.info("Running DCF model")
    dcf_output = run_dcf(
        projections=three_stmt_output,
        assumptions=default_assumptions,
    )
    
    logger.info("Running comps analysis")
    comps_output = run_comps(
        model_input=model_input,
        comparables=comps_input,
    )
    
    logger.info(f"DCF Enterprise Value: ${dcf_output.enterprise_value:,.0f}")
    logger.info(f"Comps implied values: {comps_output.implied_values}")
    
    # Step 2: Build role/year matrix from historicals
    logger.info("Building role/year matrix")
    matrix = model_input.historicals.by_role
    logger.info(f"Built matrix with {len(matrix)} model roles")
    
    # Get all years from matrix
    all_years = set()
    for role_values in matrix.values():
        all_years.update(role_values.keys())
    years_list = sorted(all_years)
    logger.info(f"Found years: {years_list}")
    
    # Step 3: Copy template to output path first (preserves original template as source of truth)
    # This also avoids permission errors if the template is open in Excel
    if template_path != output_path:
        logger.info(f"Copying template from {template_path} to {output_path}")
        shutil.copy2(template_path, output_path)
        template_path = output_path  # Work on the copy
    
    # Step 3.5: Load assumptions if provided
    assumptions_data = None
    if assumptions_path:
        logger.info(f"Loading assumptions from {assumptions_path}")
        try:
            raw_assumptions = load_structured_json(assumptions_path)
            assumptions_data = convert_assumptions_to_legacy_format(raw_assumptions)
            logger.info("Assumptions loaded and converted successfully")
            competitors = assumptions_data.get("competitors")
            if competitors:
                logger.info(f"Found {len(competitors)} competitors in assumptions: {competitors}")
        except Exception as e:
            logger.warning(f"Failed to load assumptions: {e}")
    
    # Step 4: Load Excel template (now working on copy)
    logger.info(f"Loading Excel template from {template_path}")
    workbook = load_excel_template(template_path)
    
    # Step 4.5: Populate cover sheet and summary sheet
    logger.info("Populating cover sheet and summary sheet")
    # Use fake pricing for dummy/test data
    use_fake_pricing = model_input.ticker.upper() in ["DUMMY", "TEST", "FAKE"]
    # Extract quote data if available
    quote_data = json_data.get("quote")
    populate_cover_sheet(
        workbook, 
        model_input.name, 
        model_input.ticker, 
        use_fake_pricing=use_fake_pricing,
        quote_data=quote_data
    )
    
    # Step 4.5: Populate DCF year headers for all DCF sheets
    logger.info("Populating DCF year headers")
    populate_dcf_year_headers(workbook, years_list, forecast_periods=3)
    
    # Step 4.6: Populate DCF sheet titles (B2) for all DCF sheets
    logger.info("Populating DCF sheet titles")
    dcf_sheets = ["DCF Base", "DCF Bear", "DCF Bull"]
    for sheet_name in dcf_sheets:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            title_cell = worksheet.cell(row=2, column=2)  # Cell B2
            title_cell.value = f"{model_input.name} DCF Analysis"
            logger.debug(f"Wrote '{model_input.name} DCF Analysis' to {sheet_name} B2")
        else:
            logger.debug(f"DCF sheet '{sheet_name}' not found, skipping B2 population")
    
    # Step 4.7: Extract line items from JSON for 3-statement sheet
    logger.info("Extracting line items from JSON")
    line_items = extract_line_items_from_json(json_data)
    logger.info(f"Extracted {len(line_items)} line items")
    
    # Step 5: Process DCF Base, DCF Bear, DCF Bull, and 3 Statement sheets
    target_sheets = ["DCF Base", "DCF Bear", "DCF Bull", "3 Statement"]
    for sheet_name in target_sheets:
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template. Skipping.")
            continue
        
        logger.info(f"Processing sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        
        # Handle 3 Statement sheet differently from DCF Base sheet
        if sheet_name == "3 Statement":
            # Add model_role column for 3 Statement sheet
            add_model_role_column(worksheet, role_column="ZZ")
            logger.info("Processing 3 Statement sheet")
            
            # Set sheet title: "Company Name 3 Statement Model (millions)"
            if model_input.name:
                title_cell = worksheet.cell(row=2, column=2)  # Cell B2
                title_cell.value = f"{model_input.name} 3 Statement Model (millions)"
                if Font:
                    title_cell.font = Font(bold=True, size=14)
                logger.info(f"Set sheet title to: {model_input.name} 3 Statement Model (millions)")
            
            # Detect historical columns from Income Statement row 4 (where Year"A" placeholders are)
            # Check row 4 for Year"A" placeholders (Income Statement header row)
            hist_columns = []
            for col_idx in range(4, 12):  # Columns D through K
                cell = worksheet.cell(row=4, column=col_idx)
                cell_value = str(cell.value or "").strip().lower()
                # Check for Year"A" patterns
                if "year" in cell_value and ("\"a\"" in cell_value or "'a'" in cell_value or " a" in cell_value):
                    hist_columns.append(col_idx)
            
            # If not found in row 4, try row 16 as fallback
            if not hist_columns:
                hist_columns = detect_historical_columns(worksheet)
            
            # If still not found, use default columns D, E, F
            if not hist_columns:
                logger.info("No Year\"A\" placeholders found in 3 Statement sheet, using default columns D, E, F")
                hist_columns = [4, 5, 6]
            
            logger.info(f"Using historical columns for 3 Statement: {hist_columns}")
            
            # Map years to columns
            year_column_map = map_years_to_columns(years_list, hist_columns)
            logger.info(f"Mapped years to columns: {year_column_map}")
            
            # Populate year headers for all three statements
            # Income Statement (row 4), Balance Sheet (row 24), Cash Flow (row 69)
            # This function handles Year"A" and Year"E" placeholders
            logger.info("Populating 3-statement year headers")
            populate_three_statement_year_headers(worksheet, years_list, forecast_periods=5)
            
            # Populate historicals using dedicated function for 3 Statement sheet
            populate_three_statement_historicals(
                worksheet, 
                role_column="ZZ", 
                matrix=matrix, 
                year_column_map=year_column_map,
                line_items=line_items
            )
            logger.info(f"Populated 3-statement historicals for sheet '{sheet_name}'")
            
            # Populate assumptions table if assumptions provided
            if assumptions_data and assumptions_data.get("assumptions"):
                logger.info("Populating 3-statement assumptions table")
                populate_three_statement_assumptions_table(worksheet, assumptions_data)
        elif sheet_name in ["DCF Base", "DCF Bear", "DCF Bull"]:
            # For DCF sheets (Base, Bear, Bull), populate WACC and assumptions if available
            if assumptions_data:
                logger.info(f"Populating WACC and assumptions for {sheet_name} sheet")
                logger.info(f"Assumptions data type: {type(assumptions_data)}, keys: {list(assumptions_data.keys()) if isinstance(assumptions_data, dict) else 'not a dict'}")
                # Extract quote_data if available
                quote_data_for_dcf = json_data.get("quote") or json_data.get("quote_data")
                if isinstance(quote_data_for_dcf, str) and Path(quote_data_for_dcf).exists():
                    try:
                        with open(quote_data_for_dcf, "r", encoding="utf-8") as f:
                            quote_data_for_dcf = json.load(f)
                    except Exception as e:
                        logger.warning(f"Failed to load quote data from file: {e}")
                        quote_data_for_dcf = None
                
                populate_wacc_and_assumptions(
                    worksheet,
                    model_input,
                    dcf_output,
                    assumptions_data,
                    quote_data=quote_data_for_dcf,
                    json_data=json_data
                )
            else:
                logger.warning(f"{sheet_name} sheet: No assumptions provided, skipping WACC/assumptions population")
        else:
            # For other sheets (if any), use standard DCF logic
            # Detect historical columns (looks for "Year"A" headers)
            hist_columns = detect_historical_columns(worksheet)
            if not hist_columns:
                logger.warning(f"No 'Year\"A\"' columns found in sheet '{sheet_name}'. Skipping.")
                continue
            
            logger.info(f"Found {len(hist_columns)} historical columns: {hist_columns}")
            
            # Map years to columns
            year_column_map = map_years_to_columns(years_list, hist_columns)
            logger.info(f"Mapped years to columns: {year_column_map}")
            
            # Write year headers to row 16
            write_year_headers(worksheet, year_column_map, header_row=16)
            logger.info(f"Wrote year headers to row 16")
            
            # Populate historicals
            populate_historicals(worksheet, role_column="ZZ", matrix=matrix, year_column_map=year_column_map)
            logger.info(f"Populated historicals for sheet '{sheet_name}'")
    
    # Step 6: Process RV sheet (subject company only)
    if "RV" in workbook.sheetnames:
        logger.info("Processing RV sheet (subject company)")
        worksheet = workbook["RV"]
        # Extract quote data if available
        quote_data = json_data.get("quote")
        # Auto-detect FMP format
        is_fmp_format = (
            "income_statements" in json_data or 
            "balance_sheets" in json_data or 
            "cash_flow_statements" in json_data
        )
        populate_rv_sheet(
            worksheet, 
            json_data, 
            model_input.name, 
            is_fmp_format=is_fmp_format,
            quote_data=quote_data,
            assumptions=assumptions_data
        )
        logger.info("Populated RV sheet with subject company data")
    
    # Step 7: Apply DCF Bear/Bull scenario formulas (if scenario assumptions are present)
    if assumptions_data:
        bear_scenario = assumptions_data.get("bearScenario")
        if bear_scenario:
            logger.info("Applying DCF Bear scenario formulas")
            populate_dcf_scenario(workbook, bear_scenario, "DCF Bear")

        bull_scenario = assumptions_data.get("bullScenario")
        if bull_scenario:
            logger.info("Applying DCF Bull scenario formulas")
            populate_dcf_scenario(workbook, bull_scenario, "DCF Bull")

    # Step 8: Save workbook
    logger.info(f"Saving populated template to {output_path}")
    workbook.save(output_path)
    logger.info("Template population complete")


def populate_wacc_and_assumptions(
    worksheet,
    model_input: CompanyModelInput,
    dcf_output: DcfOutput,
    assumptions: Dict[str, Any],
    quote_data: Optional[Dict[str, Any]] = None,
    json_data: Optional[Dict[str, Any]] = None
) -> None:
    """
    Populate WACC calculation inputs and Other Assumptions in DCF sheets (Base, Bear, Bull).
    
    Writes directly to specific cells (assumption-driven only):
    Column E:
    - Market Risk Premium: E6
    - Risk Free Rate: E7
    - Cost of Debt: E10 (formula-driven in template - not overwritten here)
    - Tax Rate: E13 (formula-driven in template - not overwritten here)
    
    Column J:
    - Long Term Growth Rate (LTGR): J6 (scenario-specific: base/bear/bull)
    - Cash and Cash Equivalents: J7 (from historicals)
    - Shares Outstanding: J8 (from assumptions, quote_data, or JSON)
    - Current Price: J9 (from assumptions, quote_data, or JSON)
    - Market Cap: J10 (from quote_data or calculated as shares * price)
    
    Also searches for labels in column B (WACC section) and column G (Other Assumptions),
    then writes values to columns D/E and I/J respectively for other fields.
    
    Args:
        worksheet: openpyxl Worksheet object for DCF Base/Bear/Bull sheet
        model_input: CompanyModelInput with historical data
        dcf_output: DcfOutput from run_dcf()
        assumptions: Dict with assumptions (wacc, tax_rate, terminal_growth_rate, etc.)
        quote_data: Optional quote data dict with market cap, price, shares, etc.
        json_data: Optional full JSON data for extracting additional fields
    """
    # Verify we're working with a DCF worksheet
    if worksheet.title not in ["DCF Base", "DCF Bear", "DCF Bull"]:
        logger.warning(f"populate_wacc_and_assumptions called on unexpected sheet: {worksheet.title}, expected one of: DCF Base, DCF Bear, DCF Bull")
    
    logger.info(f"Populating DCF assumptions for {worksheet.title} sheet. Assumptions keys: {list(assumptions.keys())}")
    logger.info(f"Worksheet title: {worksheet.title}, Max row: {worksheet.max_row}, Max col: {worksheet.max_column}")
    
    # Check current values in target cells for debugging
    cell_e6_before = worksheet.cell(row=6, column=5).value
    cell_e7_before = worksheet.cell(row=7, column=5).value
    cell_e10_before = worksheet.cell(row=10, column=5).value
    cell_e13_before = worksheet.cell(row=13, column=5).value
    logger.info(f"Current cell values - E6: {cell_e6_before}, E7: {cell_e7_before}, E10: {cell_e10_before}, E13: {cell_e13_before}")
    
    # Write directly to specific cells for DCF assumptions (ALWAYS overwrite)
    # Market Risk Premium: E6
    market_risk_premium = assumptions.get("market_risk_premium")
    if market_risk_premium is None:
        # Try to get from dcf section
        dcf_data = assumptions.get("dcf", {})
        market_risk_premium = dcf_data.get("market_risk_premium")
    
    if market_risk_premium is not None:
        cell_e6 = worksheet.cell(row=6, column=5)  # Column E, Row 6
        old_value = cell_e6.value
        cell_e6.value = float(market_risk_premium)
        logger.info(f"Wrote Market Risk Premium = {market_risk_premium} to E6 (was: {old_value})")
    else:
        logger.warning("Market Risk Premium not found in assumptions")
    
    # Risk Free Rate: E7
    risk_free_rate = assumptions.get("risk_free_rate")
    if risk_free_rate is None:
        # Try to get from dcf section
        dcf_data = assumptions.get("dcf", {})
        risk_free_rate = dcf_data.get("risk_free_rate")
    
    if risk_free_rate is not None:
        cell_e7 = worksheet.cell(row=7, column=5)  # Column E, Row 7
        old_value = cell_e7.value
        cell_e7.value = float(risk_free_rate)
        logger.info(f"Wrote Risk Free Rate = {risk_free_rate} (was: {old_value}) to E7")
    else:
        logger.warning("Risk Free Rate not found in assumptions")
    
    # Cost of Debt: E10 and Tax Rate: E13 are now formula-driven in the template.
    # Intentionally do NOT overwrite these cells here so Excel formulas/cell
    # references control their values based on other inputs (e.g., WACC table).
    
    # WACC section labels (column B) -> values go in column E (for other fields)
    wacc_labels = {
        "beta": ("Beta", assumptions.get("beta", 1.2)),
        "cost of equity": ("Cost of Equity", None),  # Calculated, skip
        "weight of equity": ("Weight of Equity", None),  # Calculated, skip
        "after tax cod": ("After Tax CoD", None),  # Calculated, skip
        "weight of debt": ("Weight of Debt", None),  # Calculated, skip
        "wacc": ("WACC", assumptions.get("wacc", 0.10)),
    }
    
    # Other Assumptions labels (column G) -> values go in column J
    # Net Debt: derive from JSON balance sheet (latest period) when available
    net_debt_from_json = None
    if json_data:
        balance_sheets = json_data.get("balance_sheets", [])
        if balance_sheets and isinstance(balance_sheets, list):
            latest_balance = balance_sheets[0]
            net_debt_from_json = latest_balance.get("netDebt")
            if net_debt_from_json is None:
                total_debt_bs = (
                    latest_balance.get("totalDebt")
                    or latest_balance.get("TotalDebt")
                )
                cash_bs = (
                    latest_balance.get("cashAndCashEquivalents")
                    or latest_balance.get("cash")
                    or latest_balance.get("CashAndCashEquivalents")
                )
                if total_debt_bs is not None and cash_bs is not None:
                    net_debt_from_json = total_debt_bs - cash_bs

    other_assumptions_labels = {
        "fiscal year end": ("Fiscal Year End", "12/31/2025"),
        "ltgr": ("LTGR", f"{assumptions.get('terminal_growth_rate', 0.025) * 100:.2f}%"),
        "cash & equivalents": ("Cash & Equivalents", None),  # From historicals
        "cash and equivalents": ("Cash & Equivalents", None),  # Alternative label
        "diluted s/o": ("Diluted S/O", None),  # Will write to row 8 explicitly
        "diluted shares": ("Diluted S/O", None),  # Will write to row 8 explicitly
        "current price": ("Current Price", None),  # Will write to row 9 explicitly
        "mkt cap": ("Mkt Cap", None),  # Calculated, skip
        "market cap": ("Mkt Cap", None),  # Alternative label
        # Template label may still say BV debt, but we now feed it Net Debt from JSON
        "bv debt": ("BV debt", net_debt_from_json),
        "book value debt": ("BV debt", net_debt_from_json),
        "net debt": ("Net Debt", net_debt_from_json),
        # Always force Financial Units label to reflect scaled units
        "financial units": ("Financial Units", "USD in millions"),
    }
    
    # Search for WACC labels in column B (rows 1-50) for other fields
    for row_idx in range(1, min(51, worksheet.max_row + 1)):
        label_cell = worksheet.cell(row=row_idx, column=2)  # Column B
        label_value = str(label_cell.value or "").strip().lower()
        
        # Check if this matches any WACC label
        for key, (display_name, value) in wacc_labels.items():
            if key in label_value:
                # Special-case Financial Units: always overwrite with correct text
                if key == "financial units":
                    value_cell = worksheet.cell(row=row_idx, column=5)  # Column E
                    value_cell.value = value
                    logger.debug(f"Wrote {display_name} = {value} to row {row_idx}, column E (forced units label)")
                    break
                # Other WACC fields only written when we actually have a value
                if value is not None:
                    value_cell = worksheet.cell(row=row_idx, column=5)  # Column E
                    if value_cell.value is None or str(value_cell.value).strip() == "#":
                        value_cell.value = value
                        logger.debug(f"Wrote {display_name} = {value} to row {row_idx}, column E")
                    break
    
    # Get latest cash and debt from historicals
    historicals = model_input.historicals.by_role
    all_years = set()
    for role_values in historicals.values():
        all_years.update(role_values.keys())
    latest_year = max(all_years) if all_years else None
    latest_cash = None
    latest_debt = None
    if latest_year:
        latest_cash = historicals.get("BS_CASH", {}).get(latest_year)
        # Try to get debt from various possible model roles
        latest_debt = (
            historicals.get("BS_DEBT_CURRENT", {}).get(latest_year) or
            historicals.get("BS_DEBT_NON_CURRENT", {}).get(latest_year) or
            assumptions.get("debt")
        )
    
    # Search for Other Assumptions labels in column G (rows 1-50)
    for row_idx in range(1, min(51, worksheet.max_row + 1)):
        label_cell = worksheet.cell(row=row_idx, column=7)  # Column G
        label_value = str(label_cell.value or "").strip().lower()
        
        # Check if this matches any Other Assumptions label
        for key, (display_name, value) in other_assumptions_labels.items():
            if key in label_value:
                # Special handling for Cash & Equivalents and Net/BV debt
                if "cash" in key and latest_cash is not None:
                    value = latest_cash
                elif "bv debt" in key or "book value debt" in key or "net debt" in key:
                    # Prefer Net Debt from JSON (latest period); fall back to latest_debt
                    MILLIONS_SCALE_LOCAL = 1_000_000
                    if net_debt_from_json is not None:
                        value = float(net_debt_from_json) / MILLIONS_SCALE_LOCAL
                    elif latest_debt is not None:
                        value = float(latest_debt) / MILLIONS_SCALE_LOCAL
                    else:
                        continue  # Skip if no debt value available
                # Skip Diluted S/O and Current Price here - will write to specific rows below
                elif "diluted" in key or "current price" in key:
                    continue
                
                if value is not None:
                    # Write to column J
                    value_cell = worksheet.cell(row=row_idx, column=10)  # Column J
                    if value_cell.value is None or str(value_cell.value).strip() == "#":
                        value_cell.value = value
                        logger.debug(f"Wrote {display_name} = {value} to row {row_idx}, column J")
                break
    
    # Write directly to specific cells in column J (ALWAYS overwrite)
    sheet_name = worksheet.title
    
    # Long Term Growth Rate (LTGR): J6 - manually written to ALL sheets (Base, Bear, Bull)
    terminal_growth_rate = None
    
    # Try to get scenario-specific LTGR
    if sheet_name == "DCF Base":
        terminal_growth_rate = assumptions.get("terminal_growth_rate")
        if terminal_growth_rate is None:
            dcf_data = assumptions.get("dcf", {})
            terminal_growth_rate = dcf_data.get("terminalGrowthRate") or dcf_data.get("terminal_growth_rate")
    elif sheet_name == "DCF Bear":
        bear_scenario = assumptions.get("bearScenario", {})
        terminal_growth_rate = bear_scenario.get("terminalGrowthRate") or bear_scenario.get("terminal_growth_rate")
    elif sheet_name == "DCF Bull":
        bull_scenario = assumptions.get("bullScenario", {})
        terminal_growth_rate = bull_scenario.get("terminalGrowthRate") or bull_scenario.get("terminal_growth_rate")
    
    # Default to base scenario if not found
    if terminal_growth_rate is None:
        terminal_growth_rate = assumptions.get("terminal_growth_rate", 0.025)
        dcf_data = assumptions.get("dcf", {})
        if terminal_growth_rate == 0.025:
            terminal_growth_rate = dcf_data.get("terminalGrowthRate") or dcf_data.get("terminal_growth_rate") or 0.025
    
    # Convert to decimal for Excel percentage format (Excel expects 0.02 for 2%, not 2)
    if terminal_growth_rate is not None:
        try:
            tgr_val = float(terminal_growth_rate)
            # If value is > 1, assume it's already a percentage (e.g., 2 means 2%), convert to decimal (0.02)
            # If value is <= 1, assume it's already a decimal (e.g., 0.02 means 2%), use as-is
            if tgr_val > 1:
                ltgr_decimal = tgr_val / 100.0  # Convert percentage to decimal for Excel
            else:
                ltgr_decimal = tgr_val  # Already a decimal
            cell_j6 = worksheet.cell(row=6, column=10)  # Column J, Row 6
            old_value = cell_j6.value
            cell_j6.value = ltgr_decimal
            # Set percentage format
            cell_j6.number_format = '0.00%'
            logger.info(f"Wrote LTGR = {ltgr_decimal} ({tgr_val}%) to J6 (was: {old_value}) for {sheet_name}")
        except (ValueError, TypeError) as e:
            logger.warning(f"Could not convert terminal_growth_rate '{terminal_growth_rate}' to float: {e}")
    else:
        logger.warning(f"LTGR not found for {sheet_name}")
    
    # Cash, Shares Outstanding, Current Price, Market Cap: Only write to DCF Base
    # Bear and Bull reference Base via cell references in the template
    if sheet_name == "DCF Base":
        # Cash and Cash Equivalents: J7 - from historicals (scale to millions)
        MILLIONS_SCALE = 1_000_000
        if latest_cash is not None:
            cell_j7 = worksheet.cell(row=7, column=10)  # Column J, Row 7
            old_value = cell_j7.value
            cash_scaled = float(latest_cash) / MILLIONS_SCALE
            cell_j7.value = cash_scaled
            logger.info(f"Wrote Cash & Equivalents = {cash_scaled}M (was: {old_value}, original: {latest_cash}) to J7")
        else:
            logger.warning("Cash & Equivalents not found in historicals")
        
        # Shares Outstanding: J8 - from assumptions, quote_data, or JSON (scale to millions)
        shares_outstanding = assumptions.get("shares_outstanding")
        if shares_outstanding is None and quote_data:
            shares_outstanding = (
                quote_data.get("sharesOutstanding") or
                quote_data.get("shares_outstanding") or
                quote_data.get("weightedAverageShsOutDil") or
                quote_data.get("weightedAverageShsOut")
            )
        if shares_outstanding is None and json_data:
            # Try to get from income statements
            income_statements = json_data.get("income_statements", [])
            if income_statements and isinstance(income_statements, list) and len(income_statements) > 0:
                latest_income = income_statements[0]
                shares_outstanding = (
                    latest_income.get("weightedAverageShsOutDil") or
                    latest_income.get("weightedAverageShsOut")
                )
        
        if shares_outstanding is not None:
            cell_j8 = worksheet.cell(row=8, column=10)  # Column J, Row 8
            old_value = cell_j8.value
            shares_scaled = float(shares_outstanding) / MILLIONS_SCALE
            cell_j8.value = shares_scaled
            logger.info(f"Wrote Shares Outstanding = {shares_scaled}M (was: {old_value}, original: {shares_outstanding}) to J8")
        else:
            logger.warning("Shares Outstanding not found")
        
        # Current Price: J9 - from assumptions, quote_data, or JSON
        current_price = assumptions.get("current_price")
        if current_price is None and quote_data:
            current_price = (
                quote_data.get("price") or
                quote_data.get("sharePrice") or
                quote_data.get("currentPrice") or
                quote_data.get("regularMarketPrice")
            )
        if current_price is None and json_data:
            quote_data_from_json = json_data.get("quote") or json_data.get("quote_data")
            if isinstance(quote_data_from_json, dict):
                current_price = (
                    quote_data_from_json.get("price") or
                    quote_data_from_json.get("sharePrice") or
                    quote_data_from_json.get("currentPrice")
                )
        
        if current_price is not None:
            cell_j9 = worksheet.cell(row=9, column=10)  # Column J, Row 9
            old_value = cell_j9.value
            cell_j9.value = float(current_price)
            logger.info(f"Wrote Current Price = {current_price} to J9 (was: {old_value})")
        else:
            logger.warning("Current Price not found")
        
        # Market Cap: J10 - calculate from shares * price, or get from quote_data
        market_cap = None
        if quote_data:
            market_cap = (
                quote_data.get("marketCap") or
                quote_data.get("market_cap") or
                quote_data.get("MarketCap")
            )
        if market_cap is None and json_data:
            quote_data_from_json = json_data.get("quote") or json_data.get("quote_data")
            if isinstance(quote_data_from_json, dict):
                market_cap = (
                    quote_data_from_json.get("marketCap") or
                    quote_data_from_json.get("market_cap") or
                    quote_data_from_json.get("MarketCap")
                )
        
        # Calculate market cap if we have shares and price but not market cap
        if market_cap is None and shares_outstanding is not None and current_price is not None:
            market_cap = float(shares_outstanding) * float(current_price)
            logger.info(f"Calculated Market Cap = {shares_outstanding} * {current_price} = {market_cap}")
        
        if market_cap is not None:
            cell_j10 = worksheet.cell(row=10, column=10)  # Column J, Row 10
            old_value = cell_j10.value
            market_cap_scaled = float(market_cap) / MILLIONS_SCALE
            cell_j10.value = market_cap_scaled
            logger.info(f"Wrote Market Cap = {market_cap_scaled}M (was: {old_value}, original: {market_cap}) to J10")
        else:
            logger.warning("Market Cap not found and could not be calculated")
        
        # Total Debt: assign to J11 from most recent period balance sheet in JSON
        total_debt_value = None
        if json_data:
            balance_sheets = json_data.get("balance_sheets", [])
            if balance_sheets and isinstance(balance_sheets, list):
                # Determine most recent period by 'date' or 'fiscalYear' if available
                def _balance_sort_key(bs_row: Dict[str, Any]) -> str:
                    # Prefer explicit date string; fall back to fiscalYear; otherwise empty
                    return str(bs_row.get("date") or bs_row.get("fiscalYear") or "")

                latest_balance = max(balance_sheets, key=_balance_sort_key)
                total_debt_value = (
                    latest_balance.get("totalDebt")
                    or latest_balance.get("TotalDebt")
                    or latest_balance.get("total_debt")
                )
        if total_debt_value is not None:
            cell_j11 = worksheet.cell(row=11, column=10)  # Column J, Row 11
            old_value = cell_j11.value
            total_debt_scaled = float(total_debt_value) / MILLIONS_SCALE
            cell_j11.value = total_debt_scaled
            logger.info(
                f"Wrote Total Debt = {total_debt_scaled}M "
                f"(was: {old_value}, original: {total_debt_value}) to J11"
            )
        else:
            logger.warning("Total Debt for J11 not found in JSON balance sheet data")
    else:
        # For Bear and Bull, these values are cell-referenced to Base, so skip writing
        logger.debug(f"Skipping J7-J11 for {sheet_name} (cell-referenced to DCF Base)")

def populate_dcf_scenario(
    workbook,
    scenario_assumptions: Dict[str, Any],
    sheet_name: str,
) -> None:
    """
    Populate DCF Bear/Bull scenario formulas based on scenario assumptions.

    Args:
        workbook: openpyxl Workbook object
        scenario_assumptions: legacy-format scenario dict, e.g. assumptions["bearScenario"].
        sheet_name: Name of the DCF scenario sheet (e.g., "DCF Bear" or "DCF Bull").
    """
    if not scenario_assumptions:
        logger.info(f"No scenario assumptions for {sheet_name}, skipping scenario formulas.")
        return

    if sheet_name not in workbook.sheetnames:
        logger.warning(f"Sheet {sheet_name} not found in workbook, skipping scenario formulas.")
        return

    sheet = workbook[sheet_name]
    formula_dict = build_scenario_formula_dict()
    apply_scenario_to_sheet(sheet, scenario_assumptions, formula_dict)
    logger.info(f"Applied DCF scenario formulas to sheet {sheet_name}")

def populate_three_statement_assumptions_table(
    worksheet,
    assumptions: Dict[str, Any]
) -> None:
    """
    Populate assumptions tables in 3 Statement sheet.
    
    Income Statement Assumptions (M4-S9):
    - Title "Assumptions" to M4
    - Row labels to M5-M9: Revenue, Gross Margin, Operating Margin, Tax Rate, Interest Rate of Debt
    - Types to N5-N9: step, constant, or custom
    - Values to O5-S9: 5 period values for each assumption
    
<｜tool▁sep｜>new_string
    Balance Sheet Assumptions (M25-S27):
    - Row labels to M25-M27: Depreciation % of PPE, Inventory % Sales, Long Term Debt Change
    - Types to N25-N27: step, constant, or custom
    - Values to O25-S27: 5 period values for each assumption
    
    Cash Flow Assumptions (M70-S72):
    - Row labels to M70-M72: Share Repurchases, Dividends as a % of Net Income, CAPEX
    - Types to N70-N72: step, constant, or custom
    - Values to O70-S72: 5 period values for each assumption
    
    Args:
        worksheet: openpyxl Worksheet object for "3 Statement" sheet
        assumptions: Dict containing assumptions data from JSON:
            {
                "assumptions": {
                    "revenue": {"type": "step", "values": [0.05, 0.06, ...]},
                    "gross_margin": {"type": "constant", "values": [0.40, ...]},
                    ...
                },
                "balance_sheet": {
                    "depreciation_ppe": {"type": "constant", "values": [0.10, ...]},
                    "inventory_sales": {"type": "step", "values": [0.15, ...]},
                    "lt_debt_change": {"type": "custom", "values": [0.05, ...]}
                },
                "cash_flow": {
                    "share_repurchases": {"type": "step", "values": [1000000, ...]},
                    "dividends_ni": {"type": "constant", "values": [0.30, ...]},
                    "capex": {"type": "custom", "values": [5000000, ...]}
                }
            }
    """
    assumptions_data = assumptions.get("assumptions")
    if not assumptions_data:
        logger.warning("No assumptions data found, skipping assumptions table population")
        return
    
    # Column mappings
    COL_M = 13  # Row labels
    COL_N = 14  # Types
    COL_O = 15  # Period 1
    COL_P = 16  # Period 2
    COL_Q = 17  # Period 3
    COL_R = 18  # Period 4
    COL_S = 19  # Period 5
    
    # Row mappings
    ROW_TITLE = 4
    ROW_REVENUE = 5
    ROW_GROSS_MARGIN = 6
    ROW_OPERATING_MARGIN = 7
    ROW_TAX_RATE = 8
    ROW_INTEREST_RATE_DEBT = 9
    
    # Assumption mapping: key name -> (row, display_name)
    assumption_map = {
        "revenue": (ROW_REVENUE, "Revenue"),
        "gross_margin": (ROW_GROSS_MARGIN, "Gross Margin"),
        "operating_margin": (ROW_OPERATING_MARGIN, "Operating Margin"),
        "tax_rate": (ROW_TAX_RATE, "Tax Rate"),
        "interest_rate_debt": (ROW_INTEREST_RATE_DEBT, "Interest Rate of Debt"),
    }
    
    # Write title to M4
    title_cell = worksheet.cell(row=ROW_TITLE, column=COL_M)
    if title_cell.value is None or str(title_cell.value).strip() == "":
        title_cell.value = "Assumptions"
        logger.debug(f"Wrote title 'Assumptions' to M{ROW_TITLE}")
    
    # Write row labels, types, and values
    for key, (row_num, display_name) in assumption_map.items():
        assumption_info = assumptions_data.get(key)
        
        if not assumption_info:
            logger.debug(f"Assumption '{key}' not found in assumptions data, skipping row {row_num}")
            continue
        
        # Write row label to column M
        label_cell = worksheet.cell(row=row_num, column=COL_M)
        if label_cell.value is None or str(label_cell.value).strip() == "":
            label_cell.value = display_name
            logger.debug(f"Wrote label '{display_name}' to M{row_num}")
        
        # Write type to column N
        assumption_type = assumption_info.get("type")
        if assumption_type:
            type_cell = worksheet.cell(row=row_num, column=COL_N)
            if type_cell.value is None or str(type_cell.value).strip() == "":
                type_cell.value = assumption_type
                logger.debug(f"Wrote type '{assumption_type}' to N{row_num}")
        
        # Write values to columns O-S (5 periods)
        values = assumption_info.get("values", [])
        if not values:
            logger.warning(f"No values found for assumption '{key}', skipping values")
            continue
        
        # Ensure we have exactly 5 values (pad with last value or None if needed)
        if len(values) < 5:
            # Pad with last value if available, otherwise None
            last_value = values[-1] if values else None
            values = values + [last_value] * (5 - len(values))
            logger.debug(f"Padded values for '{key}' to 5 periods")
        elif len(values) > 5:
            # Use first 5 values
            values = values[:5]
            logger.debug(f"Truncated values for '{key}' to 5 periods")
        
        # Write values to columns O-S
        value_columns = [COL_O, COL_P, COL_Q, COL_R, COL_S]
        for col_idx, value in zip(value_columns, values):
            if value is not None:
                value_cell = worksheet.cell(row=row_num, column=col_idx)
                # Write as number (float)
                if isinstance(value, (int, float)):
                    value_cell.value = float(value)
                else:
                    # Try to convert to float
                    try:
                        value_cell.value = float(value)
                    except (ValueError, TypeError):
                        logger.warning(f"Could not convert value '{value}' to float for '{key}' at column {col_idx}")
                        continue
                logger.debug(f"Wrote value {value} to {chr(64 + col_idx)}{row_num} for '{key}'")
        
        logger.info(f"Populated assumption '{display_name}' (type: {assumption_type}) with {len([v for v in values if v is not None])} values")
    
    logger.info("Populated 3-statement assumptions table (M4-S9)")
    
    # Populate Balance Sheet assumptions table (M25-S27)
    # Note: balance_sheet is at the root level of assumptions dict, not under assumptions_data
    bs_assumptions_data = assumptions.get("balance_sheet")
    if bs_assumptions_data:
        logger.info("Populating Balance Sheet assumptions table")
        
        # Column mappings (same as Income Statement)
        COL_M = 13  # Row labels
        COL_N = 14  # Types
        COL_O = 15  # Period 1
        COL_P = 16  # Period 2
        COL_Q = 17  # Period 3
        COL_R = 18  # Period 4
        COL_S = 19  # Period 5
        
        # Row mappings for Balance Sheet assumptions (starting at row 25)
        ROW_DEPRECIATION_PPE = 25
        ROW_INVENTORY_SALES = 26
        ROW_LT_DEBT_CHANGE = 27
        
        # Balance Sheet assumption mapping: key name -> (row, display_name)
        bs_assumption_map = {
            "depreciation_ppe": (ROW_DEPRECIATION_PPE, "Depreciation % of PPE"),
            "inventory_sales": (ROW_INVENTORY_SALES, "Inventory % Sales"),
            "lt_debt_change": (ROW_LT_DEBT_CHANGE, "Long Term Debt Change"),
        }
        
        # Write row labels, types, and values for Balance Sheet assumptions
        for key, (row_num, display_name) in bs_assumption_map.items():
            assumption_info = bs_assumptions_data.get(key)
            
            if not assumption_info:
                logger.debug(f"Balance Sheet assumption '{key}' not found in assumptions data, skipping row {row_num}")
                continue
            
            # Write row label to column M
            label_cell = worksheet.cell(row=row_num, column=COL_M)
            label_cell.value = display_name
            logger.debug(f"Wrote label '{display_name}' to M{row_num}")
            
            # Write type to column N
            assumption_type = assumption_info.get("type")
            if assumption_type:
                type_cell = worksheet.cell(row=row_num, column=COL_N)
                type_cell.value = assumption_type
                logger.debug(f"Wrote type '{assumption_type}' to N{row_num}")
            
            # Write values to columns O-S (5 periods)
            values = assumption_info.get("values", [])
            if not values:
                logger.warning(f"No values found for Balance Sheet assumption '{key}', skipping values")
                continue
            
            # Ensure we have exactly 5 values (pad with last value or None if needed)
            if len(values) < 5:
                # Pad with last value if available, otherwise None
                last_value = values[-1] if values else None
                values = values + [last_value] * (5 - len(values))
                logger.debug(f"Padded values for '{key}' to 5 periods")
            elif len(values) > 5:
                # Use first 5 values
                values = values[:5]
                logger.debug(f"Truncated values for '{key}' to 5 periods")
            
            # Write values to columns O-S
            value_columns = [COL_O, COL_P, COL_Q, COL_R, COL_S]
            for col_idx, value in zip(value_columns, values):
                if value is not None:
                    value_cell = worksheet.cell(row=row_num, column=col_idx)
                    # Write as number (float)
                    if isinstance(value, (int, float)):
                        value_cell.value = float(value)
                    else:
                        # Try to convert to float
                        try:
                            value_cell.value = float(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert value '{value}' to float for '{key}' at column {col_idx}")
                            continue
                    logger.debug(f"Wrote value {value} to {chr(64 + col_idx)}{row_num} for '{key}'")
            
            logger.info(f"Populated Balance Sheet assumption '{display_name}' (type: {assumption_type}) with {len([v for v in values if v is not None])} values")
        
        logger.info("Populated Balance Sheet assumptions table (M25-S27)")
    else:
        logger.debug("No Balance Sheet assumptions data found, skipping Balance Sheet assumptions table")
    
    # Populate Cash Flow assumptions table (M72-S74)
    # Note: cash_flow is at the root level of assumptions dict
    cf_assumptions_data = assumptions.get("cash_flow")
    if cf_assumptions_data:
        logger.info("Populating Cash Flow assumptions table")
        
        # Column mappings (same as Income Statement and Balance Sheet)
        COL_M = 13  # Row labels
        COL_N = 14  # Types
        COL_O = 15  # Period 1
        COL_P = 16  # Period 2
        COL_Q = 17  # Period 3
        COL_R = 18  # Period 4
        COL_S = 19  # Period 5
        
        # Row mappings for Cash Flow assumptions
        ROW_SHARE_REPURCHASES = 70
        ROW_DIVIDENDS_NI = 71
        ROW_CAPEX = 72
        
        # Cash Flow assumption mapping: key name -> (row, display_name)
        cf_assumption_map = {
            "share_repurchases": (ROW_SHARE_REPURCHASES, "Share Repurchases"),
            "dividends_ni": (ROW_DIVIDENDS_NI, "Dividends as a % of Net Income"),
            "capex": (ROW_CAPEX, "CAPEX"),
        }
        
        # Write row labels, types, and values for Cash Flow assumptions
        for key, (row_num, display_name) in cf_assumption_map.items():
            assumption_info = cf_assumptions_data.get(key)
            
            if not assumption_info:
                logger.debug(f"Cash Flow assumption '{key}' not found in assumptions data, skipping row {row_num}")
                continue
            
            # Write row label to column M
            label_cell = worksheet.cell(row=row_num, column=COL_M)
            label_cell.value = display_name
            logger.debug(f"Wrote label '{display_name}' to M{row_num}")
            
            # Write type to column N
            assumption_type = assumption_info.get("type")
            if assumption_type:
                type_cell = worksheet.cell(row=row_num, column=COL_N)
                type_cell.value = assumption_type
                logger.debug(f"Wrote type '{assumption_type}' to N{row_num}")
            
            # Write values to columns O-S (5 periods)
            values = assumption_info.get("values", [])
            if not values:
                logger.warning(f"No values found for Cash Flow assumption '{key}', skipping values")
                continue
            
            # Ensure we have exactly 5 values (pad with last value or None if needed)
            if len(values) < 5:
                # Pad with last value if available, otherwise None
                last_value = values[-1] if values else None
                values = values + [last_value] * (5 - len(values))
                logger.debug(f"Padded values for '{key}' to 5 periods")
            elif len(values) > 5:
                # Use first 5 values
                values = values[:5]
                logger.debug(f"Truncated values for '{key}' to 5 periods")
            
            # Write values to columns O-S
            value_columns = [COL_O, COL_P, COL_Q, COL_R, COL_S]
            for col_idx, value in zip(value_columns, values):
                if value is not None:
                    value_cell = worksheet.cell(row=row_num, column=col_idx)
                    # Write as number (float)
                    if isinstance(value, (int, float)):
                        value_cell.value = float(value)
                    else:
                        # Try to convert to float
                        try:
                            value_cell.value = float(value)
                        except (ValueError, TypeError):
                            logger.warning(f"Could not convert value '{value}' to float for '{key}' at column {col_idx}")
                            continue
                    logger.debug(f"Wrote value {value} to {chr(64 + col_idx)}{row_num} for '{key}'")
            
            logger.info(f"Populated Cash Flow assumption '{display_name}' (type: {assumption_type}) with {len([v for v in values if v is not None])} values")
        
        logger.info("Populated Cash Flow assumptions table (M70-S72)")
    else:
        logger.debug("No Cash Flow assumptions data found, skipping Cash Flow assumptions table")


def fetch_competitor_metrics(ticker: str) -> Optional[Dict[str, Any]]:
    """
    Fetch and extract competitor metrics from FMP API.
    
    Optimized to fetch only the minimum required data:
    - Quote data (Market Cap, Share Price) - single API call
    - Income Statement (Sales, EPS, Shares Outstanding) - limit=1 (latest period only)
    - Balance Sheet (Book Value, Net Debt) - limit=1 (latest period only)
    
    Does NOT fetch Cash Flow Statement (not needed for competitor metrics).
    
    Extracts the same metrics as subject company:
    - Market Cap (from quote)
    - Share Price (from quote)
    - Sales/Revenue (from income statement)
    - Enterprise Value (calculated: Market Cap + Net Debt)
    - Net Debt (from balance sheet or calculated)
    - EPS (from income statement)
    - Book Value (from balance sheet)
    - Shares Outstanding (from income statement)
    
    Args:
        ticker: Competitor ticker symbol (e.g., "AAPL")
        
    Returns:
        Dictionary with metrics or None if data unavailable
        {
            "market_cap": float,
            "share_price": float,
            "sales": float,
            "enterprise_value": float,
            "net_debt": float,
            "eps": float,
            "book_value": float,
            "shares_outstanding": float,
            "company_name": str
        }
    """
    from app.data.fmp_client import fetch_quote, fetch_income_statement, fetch_balance_sheet
    
    try:
        logger.info(f"Fetching competitor data for {ticker} (optimized: quote + IS + BS only)")
        metrics = {}
        company_name = ticker
        
        # Fetch quote data (Market Cap, Share Price)
        try:
            quote_data = fetch_quote(ticker.upper())
            metrics["market_cap"] = quote_data.get("marketCap")
            metrics["share_price"] = quote_data.get("price")
            logger.debug(f"Fetched quote data for {ticker}: marketCap={metrics.get('market_cap')}, price={metrics.get('share_price')}")
        except Exception as e:
            logger.warning(f"Error fetching quote data for {ticker}: {e}")
        
        # Fetch income statement (Sales, EPS, Shares Outstanding) - only latest period
        try:
            income_data = fetch_income_statement(ticker.upper(), limit=1)
            if income_data and isinstance(income_data, list) and len(income_data) > 0:
                latest_income = income_data[0]
                company_name = latest_income.get("companyName", ticker)
                
                # Extract income statement metrics
                metrics["sales"] = latest_income.get("revenue")
                metrics["eps"] = latest_income.get("epsDiluted") or latest_income.get("eps")
                metrics["shares_outstanding"] = (
                    latest_income.get("weightedAverageShsOutDil") or 
                    latest_income.get("weightedAverageShsOut")
                )
                logger.debug(f"Fetched income statement for {ticker}: revenue={metrics.get('sales')}, eps={metrics.get('eps')}")
        except Exception as e:
            logger.warning(f"Error fetching income statement for {ticker}: {e}")
        
        # Fetch balance sheet (Book Value, Net Debt) - only latest period
        try:
            balance_data = fetch_balance_sheet(ticker.upper(), limit=1)
            if balance_data and isinstance(balance_data, list) and len(balance_data) > 0:
                latest_balance = balance_data[0]
                
                # Book Value
                metrics["book_value"] = (
                    latest_balance.get("totalStockholdersEquity") or 
                    latest_balance.get("totalEquity")
                )
                
                # Net Debt
                metrics["net_debt"] = latest_balance.get("netDebt")
                if metrics["net_debt"] is None:
                    total_debt = latest_balance.get("totalDebt")
                    cash = latest_balance.get("cashAndCashEquivalents")
                    if total_debt is not None and cash is not None:
                        metrics["net_debt"] = total_debt - cash
                logger.debug(f"Fetched balance sheet for {ticker}: bookValue={metrics.get('book_value')}, netDebt={metrics.get('net_debt')}")
        except Exception as e:
            logger.warning(f"Error fetching balance sheet for {ticker}: {e}")
        
        # Calculate Enterprise Value
        if metrics.get("market_cap") is not None and metrics.get("net_debt") is not None:
            metrics["enterprise_value"] = metrics["market_cap"] + metrics["net_debt"]
        
        metrics["company_name"] = company_name
        
        logger.info(f"Successfully extracted metrics for {ticker}: {list(metrics.keys())}")
        return metrics
        
    except Exception as e:
        logger.warning(f"Error fetching competitor metrics for {ticker}: {e}")
        return None


def populate_competitor_data(
    worksheet,
    competitors: List[str]
) -> None:
    """
    Populate competitor data in RV sheet.
    
    Writes competitor names to paired rows (B6/B17, B7/B18, B8/B19, B9/B20)
    and metrics to rows 17-20, columns D-K.
    
    Args:
        worksheet: openpyxl Worksheet object for RV sheet
        competitors: List of 4 competitor ticker symbols
    """
    MILLIONS_SCALE = 1_000_000
    
    # Competitor name rows (paired)
    name_rows_upper = [6, 7, 8, 9]  # B6-B9
    name_rows_lower = [17, 18, 19, 20]  # B17-B20
    
    # Metrics rows (same as lower name rows)
    metrics_rows = [17, 18, 19, 20]
    
    metric_order = [
        ("market_cap", 4, True),      # D - Market Cap - scale to millions
        ("share_price", 5, False),    # E - Share Price - keep as-is (dollars)
        ("sales", 6, True),           # F - Sales (Revenue) - scale to millions
        ("enterprise_value", 7, True), # G - Enterprise Value - scale to millions
        ("net_debt", 8, True),        # H - Net Debt - scale to millions
        ("eps", 9, False),            # I - EPS - keep as-is (dollars per share)
        ("book_value", 10, True),      # J - Book Value - scale to millions
        ("shares_outstanding", 11, False), # K - Shares Outstanding - keep as-is (shares)
    ]
    
    for idx, ticker in enumerate(competitors):
        if idx >= 4:
            break  # Only handle first 4 competitors
        
        # Fetch competitor metrics
        competitor_metrics = fetch_competitor_metrics(ticker)
        
        if not competitor_metrics:
            logger.warning(f"Skipping competitor {ticker} - failed to fetch metrics")
            continue
        
        company_name = competitor_metrics.get("company_name", ticker)
        
        # Write competitor name to paired rows
        if idx < len(name_rows_upper):
            # Upper row (B6-B9)
            worksheet.cell(row=name_rows_upper[idx], column=2).value = company_name
            logger.debug(f"Wrote competitor {idx+1} name '{company_name}' to B{name_rows_upper[idx]}")
        
        if idx < len(name_rows_lower):
            # Lower row (B17-B20)
            worksheet.cell(row=name_rows_lower[idx], column=2).value = company_name
            logger.debug(f"Wrote competitor {idx+1} name '{company_name}' to B{name_rows_lower[idx]}")
        
        # Write metrics to row 17-20, columns D-K
        if idx < len(metrics_rows):
            target_row = metrics_rows[idx]
            for metric_key, column, scale_to_millions in metric_order:
                value = competitor_metrics.get(metric_key)
                if value is not None:
                    cell = worksheet.cell(row=target_row, column=column)
                    if scale_to_millions:
                        cell.value = float(value) / MILLIONS_SCALE
                    else:
                        cell.value = float(value)
                    cell.number_format = "General"
                    logger.debug(f"Wrote competitor {idx+1} {metric_key} = {value} {'(scaled to millions)' if scale_to_millions else ''} to row {target_row}, column {column}")
                else:
                    logger.debug(f"No value found for competitor {idx+1} {metric_key}, skipping")
            
            logger.info(f"Populated competitor {idx+1} ({ticker}) metrics to row {target_row}")


def populate_rv_sheet(
    worksheet,
    json_data: Dict[str, Any],
    company_name: str,
    is_fmp_format: bool = False,
    quote_data: Optional[Dict[str, Any]] = None,
    assumptions: Optional[Dict[str, Any]] = None
) -> None:
    """
    Populate RV (Relative Valuation) sheet with company name and key metrics.
    
    Writes:
    - Company name to cells B5, B13, B16
    - Metrics to columns D-K in row 16 only: Market Cap | Share Price | Sales (Revenue) | Enterprise Value | Net Debt | EPS | Book Value | Shares Outstanding
    - Competitor names to B6-B9 and B17-B20 (paired rows)
    - Competitor metrics to rows 17-20, columns D-K
    
    Args:
        worksheet: openpyxl Worksheet object for RV sheet
        json_data: JSON data containing financial statements
        company_name: Company name string
        is_fmp_format: Whether JSON is in FMP format
        quote_data: Optional quote/price data dictionary
        assumptions: Optional assumptions dictionary containing competitors list: {"competitors": ["TICKER1", "TICKER2", "TICKER3", "TICKER4"]}
    """
    if not company_name:
        logger.warning("No company name provided, skipping RV sheet population")
        return
    
    # Write company name to B2: "Company Name Relative Valuation"
    title_cell = worksheet.cell(row=2, column=2)  # Cell B2
    title_cell.value = f"{company_name} Relative Valuation"
    logger.debug(f"Wrote '{company_name} Relative Valuation' to RV Sheet B2")
    
    # Write company name to B5, B13, B16
    worksheet.cell(row=5, column=2).value = company_name  # B5
    worksheet.cell(row=13, column=2).value = company_name  # B13
    worksheet.cell(row=16, column=2).value = company_name  # B16
    logger.info(f"Wrote company name '{company_name}' to B5, B13, B16")
    
    # Extract metrics from JSON data
    metrics = {}
    
    # Get quote data if not provided (try regardless of format)
    if not quote_data:
        quote_data = json_data.get("quote") or json_data.get("quote_data")
        # If quote is a file path string, try to load it
        if isinstance(quote_data, str) and Path(quote_data).exists():
            try:
                with open(quote_data, "r", encoding="utf-8") as f:
                    quote_data = json.load(f)
                logger.debug(f"Loaded quote data from file")
            except Exception as e:
                logger.warning(f"Failed to load quote data from file: {e}")
                quote_data = None
        elif quote_data:
            logger.debug(f"Found quote data in JSON")
    
    # Extract from quote data
    if quote_data:
        # Try multiple field name variations for market cap
        metrics["market_cap"] = (
            quote_data.get("marketCap") or
            quote_data.get("market_cap") or
            quote_data.get("MarketCap") or
            quote_data.get("mktCap")
        )
        # Try multiple field name variations for share price
        metrics["share_price"] = (
            quote_data.get("price") or
            quote_data.get("sharePrice") or
            quote_data.get("SharePrice") or
            quote_data.get("currentPrice")
        )
        if metrics["market_cap"]:
            logger.debug(f"Found market cap: {metrics['market_cap']}")
        if metrics["share_price"]:
            logger.debug(f"Found share price: {metrics['share_price']}")
    
    # Auto-detect FMP format if JSON has FMP structure
    has_fmp_structure = (
        "income_statements" in json_data or 
        "balance_sheets" in json_data or 
        "cash_flow_statements" in json_data
    )
    if has_fmp_structure:
        is_fmp_format = True
        logger.debug("Auto-detected FMP format from JSON structure")
    
    # Extract from income statement (most recent period)
    income_statements = []
    if is_fmp_format:
        income_statements = json_data.get("income_statements", [])
        logger.debug(f"Found {len(income_statements)} income statements in FMP format")
    else:
        # Structured format - use extract_line_items_from_json to handle all formats
        line_items = extract_line_items_from_json(json_data)
        # Group by statement type and find latest period
        income_items = [item for item in line_items if item.get("statement") == "income_statement"]
        
        # Get latest period from income statement items
        if income_items:
            # Find all periods across income statement items
            all_periods = set()
            for item in income_items:
                periods = item.get("periods", {})
                all_periods.update(periods.keys())
            
            if all_periods:
                # Get latest period (most recent date)
                latest_period = max(all_periods)
                
                # Build latest income statement from line items
                latest_income = {}
                for item in income_items:
                    periods = item.get("periods", {})
                    if latest_period in periods:
                        tag = item.get("tag", "")
                        value = periods[latest_period]
                        latest_income[tag] = value
                
                income_statements = [latest_income]
    
    if income_statements and len(income_statements) > 0:
        latest_income = income_statements[0]  # Most recent is first
        logger.debug(f"Extracting metrics from income statement dated: {latest_income.get('date', 'unknown')}")
        
        # Sales (Revenue) - try multiple field names
        metrics["sales"] = (
            latest_income.get("revenue") or 
            latest_income.get("Revenue") or
            latest_income.get("totalRevenue")
        )
        if metrics["sales"]:
            logger.debug(f"Found revenue: {metrics['sales']}")
        
        # EPS (prefer diluted)
        metrics["eps"] = latest_income.get("epsDiluted") or latest_income.get("eps") or latest_income.get("EPS")
        if metrics["eps"]:
            logger.debug(f"Found EPS: {metrics['eps']}")
        
        # Shares Outstanding (prefer diluted)
        metrics["shares_outstanding"] = (
            latest_income.get("weightedAverageShsOutDil") or 
            latest_income.get("weightedAverageShsOut") or
            latest_income.get("sharesOutstanding")
        )
        if metrics["shares_outstanding"]:
            logger.debug(f"Found shares outstanding: {metrics['shares_outstanding']}")
    
    # Extract from balance sheet (most recent period)
    balance_sheets = []
    if is_fmp_format:
        balance_sheets = json_data.get("balance_sheets", [])
        logger.debug(f"Found {len(balance_sheets)} balance sheets in FMP format")
    else:
        # Structured format - use extract_line_items_from_json to handle all formats
        line_items = extract_line_items_from_json(json_data)
        # Group by statement type and find latest period
        balance_items = [item for item in line_items if item.get("statement") == "balance_sheet"]
        
        # Get latest period from balance sheet items
        if balance_items:
            # Find all periods across balance sheet items
            all_periods = set()
            for item in balance_items:
                periods = item.get("periods", {})
                all_periods.update(periods.keys())
            
            if all_periods:
                # Get latest period (most recent date)
                latest_period = max(all_periods)
                
                # Build latest balance sheet from line items
                latest_balance = {}
                for item in balance_items:
                    periods = item.get("periods", {})
                    if latest_period in periods:
                        tag = item.get("tag", "")
                        value = periods[latest_period]
                        latest_balance[tag] = value
                
                balance_sheets = [latest_balance]
    
    if balance_sheets and len(balance_sheets) > 0:
        latest_balance = balance_sheets[0]  # Most recent is first
        logger.debug(f"Extracting metrics from balance sheet dated: {latest_balance.get('date', 'unknown')}")
        
        # Book Value (Total Equity) - try multiple field names
        metrics["book_value"] = (
            latest_balance.get("totalStockholdersEquity") or 
            latest_balance.get("totalEquity") or
            latest_balance.get("TotalStockholdersEquity") or
            latest_balance.get("TotalEquity")
        )
        if metrics["book_value"]:
            logger.debug(f"Found book value: {metrics['book_value']}")
        
        # Net Debt - check if already calculated, otherwise calculate from Total Debt - Cash
        metrics["net_debt"] = latest_balance.get("netDebt")  # FMP format may have this pre-calculated
        
        if metrics["net_debt"] is None:
            # Calculate Net Debt = Total Debt - Cash
            total_debt = (
                latest_balance.get("totalDebt") or 
                latest_balance.get("TotalDebt") or
                latest_balance.get("totalLiabilities")
            )
            cash = (
                latest_balance.get("cashAndCashEquivalents") or 
                latest_balance.get("CashAndCashEquivalents") or
                latest_balance.get("cash") or
                latest_balance.get("Cash")
            )
            if total_debt is not None and cash is not None:
                metrics["net_debt"] = total_debt - cash
                logger.debug(f"Calculated net debt: {total_debt} - {cash} = {metrics['net_debt']}")
        else:
            logger.debug(f"Found net debt (pre-calculated): {metrics['net_debt']}")
        
        # Enterprise Value = Market Cap + Net Debt
        # (or Market Cap + Total Debt - Cash if net_debt not available)
        if metrics.get("market_cap") is not None:
            if metrics.get("net_debt") is not None:
                metrics["enterprise_value"] = metrics["market_cap"] + metrics["net_debt"]
                logger.debug(f"Calculated enterprise value: {metrics['market_cap']} + {metrics['net_debt']} = {metrics['enterprise_value']}")
            else:
                total_debt = latest_balance.get("totalDebt")
                cash = latest_balance.get("cashAndCashEquivalents")
                if total_debt is not None and cash is not None:
                    metrics["enterprise_value"] = metrics["market_cap"] + total_debt - cash
                    logger.debug(f"Calculated enterprise value: {metrics['market_cap']} + {total_debt} - {cash} = {metrics['enterprise_value']}")
    
    # Write metrics to columns D-K (columns 4-11) for row 16 only
    # Order: Market Cap | Share Price | Sales (Revenue) | Enterprise Value | Net Debt | EPS | Book Value | Shares Outstanding
    # Scale to millions where appropriate (Market Cap, Sales, Enterprise Value, Net Debt, Book Value)
    # Keep as-is: Share Price, EPS, Shares Outstanding
    MILLIONS_SCALE = 1_000_000
    
    metric_order = [
        ("market_cap", 4, True),      # D - Market Cap - scale to millions
        ("share_price", 5, False),    # E - Share Price - keep as-is (dollars)
        ("sales", 6, True),           # F - Sales (Revenue) - scale to millions
        ("enterprise_value", 7, True), # G - Enterprise Value - scale to millions
        ("net_debt", 8, True),        # H - Net Debt - scale to millions
        ("eps", 9, False),            # I - EPS - keep as-is (dollars per share)
        ("book_value", 10, True),      # J - Book Value - scale to millions
        ("shares_outstanding", 11, False), # K - Shares Outstanding - keep as-is (shares)
    ]
    
    # Write metrics to row 16 only
    target_row = 16
    
    for metric_key, column, scale_to_millions in metric_order:
        value = metrics.get(metric_key)
        if value is not None:
            cell = worksheet.cell(row=target_row, column=column)
            if scale_to_millions:
                cell.value = float(value) / MILLIONS_SCALE
            else:
                cell.value = float(value)
            cell.number_format = "General"
            logger.debug(f"Wrote {metric_key} = {value} {'(scaled to millions)' if scale_to_millions else ''} to row {target_row}, column {column}")
        else:
            logger.debug(f"No value found for {metric_key}, skipping")
    
    logger.info(f"Populated RV sheet metrics to row {target_row}: {list(metrics.keys())}")
    
    # Step 2: Populate competitor data if provided in assumptions
    if assumptions:
        competitors = assumptions.get("competitors")
        if competitors and isinstance(competitors, list):
            if len(competitors) > 4:
                logger.warning(f"More than 4 competitors provided ({len(competitors)}), using first 4")
                competitors = competitors[:4]
            elif len(competitors) < 4:
                logger.warning(f"Less than 4 competitors provided ({len(competitors)}), will populate available competitors")
            
            logger.info(f"Populating competitor data for {len(competitors)} competitors")
            populate_competitor_data(worksheet, competitors)
        elif competitors:
            logger.warning(f"Competitors in assumptions is not a list: {type(competitors)}")


def _write_rv_sheet(
    worksheet,
    model_input: CompanyModelInput,
    comps_input: List[Dict[str, Any]]
) -> None:
    """
    Write relative valuation data to RV sheet.
    
    Args:
        worksheet: openpyxl Worksheet object for RV sheet
        model_input: CompanyModelInput with historical data
        comps_input: List of comparable company dictionaries
    """
    # Get latest year's data for subject company
    historicals = model_input.historicals.by_role
    latest_year = max(
        (year for role_data in historicals.values() for year in role_data.keys()),
        default=None
    )
    
    if not latest_year:
        logger.warning("No historical data found for RV sheet")
        return
    
    # Extract subject company metrics
    revenue = historicals.get("IS_REVENUE", {}).get(latest_year, 0.0)
    net_income = historicals.get("IS_NET_INCOME", {}).get(latest_year, 0.0)
    
    # Calculate EBITDA if possible (EBIT + D&A)
    ebit = historicals.get("IS_EBIT", {}).get(latest_year)
    da = historicals.get("CF_DEPRECIATION", {}).get(latest_year, 0.0)
    if ebit is not None:
        ebitda = ebit + da
    else:
        ebitda = None
    
    # Write comps table (assuming a standard layout)
    # This is a simplified implementation - adjust based on actual RV sheet structure
    # Typically comps are in rows starting around row 5-10, with columns for Name, EV/EBITDA, P/E, EV/Sales
    
    # Find where to write comps (look for header row)
    # Try multiple search patterns for comps table
    comps_start_row = None
    search_patterns = ["comparable", "comp", "peer", "company", "name"]
    
    for row_idx in range(1, min(30, worksheet.max_row + 1)):
        row = worksheet[row_idx]
        for cell in row:
            if cell.value:
                cell_value_lower = str(cell.value).lower()
                # Check if cell contains any of the search patterns
                if any(pattern in cell_value_lower for pattern in search_patterns):
                    # Check if this looks like a header row (has multiple headers or is near other headers)
                    # Look for common comps headers: Name, EV/EBITDA, P/E, EV/Sales
                    header_indicators = ["ev", "ebitda", "p/e", "pe", "sales", "multiple"]
                    if any(indicator in cell_value_lower for indicator in header_indicators):
                        comps_start_row = row_idx + 1
                        break
                    # Or if it's just "comparable" or "comp", use the next row
                    elif "comparable" in cell_value_lower or ("comp" in cell_value_lower and len(cell_value_lower) < 15):
                        comps_start_row = row_idx + 1
                        break
        if comps_start_row:
            break
    
    # If still not found, try a default location (row 5 is common for comps tables)
    if not comps_start_row:
        logger.debug("Could not find comps table header in RV sheet, using default row 5")
        comps_start_row = 5
    
    # Write comps data
    for idx, comp in enumerate(comps_input):
        row_num = comps_start_row + idx
        # Column B: Name
        worksheet.cell(row=row_num, column=2).value = comp.get("name", "")
        # Column C: EV/EBITDA
        if comp.get("ev_ebitda") is not None:
            worksheet.cell(row=row_num, column=3).value = comp["ev_ebitda"]
        # Column D: P/E
        if comp.get("pe") is not None:
            worksheet.cell(row=row_num, column=4).value = comp["pe"]
        # Column E: EV/Sales
        if comp.get("ev_sales") is not None:
            worksheet.cell(row=row_num, column=5).value = comp["ev_sales"]
    
    # Write subject company metrics (typically in a separate section)
    # Look for "Subject" or company name row
    subject_row = None
    for row_idx in range(1, min(30, worksheet.max_row + 1)):
        row = worksheet[row_idx]
        for cell in row:
            if cell.value and ("subject" in str(cell.value).lower() or model_input.ticker.lower() in str(cell.value).lower()):
                subject_row = row_idx
                break
        if subject_row:
            break
    
    if subject_row:
        # Write subject metrics (adjust columns based on actual layout)
        # Revenue
        worksheet.cell(row=subject_row, column=6).value = revenue
        # EBITDA
        if ebitda is not None:
            worksheet.cell(row=subject_row, column=7).value = ebitda
        # Net Income
        worksheet.cell(row=subject_row, column=8).value = net_income


# Formula template dictionary for Summary sheet Excel output
# Format: {cell_address: formula_string}
# All formulas reference other sheets in the workbook
SUMMARY_FORMULAS: Dict[str, str] = {
    "C3": "='Cover Sheet'!B3",
    "C5": "=(C9*D9)+(C10*D10)",
    "C6": "=(C4/C5-1)",
    "D9": "=(C13*D13)+(C14*D14)+(C15*D15)",
    "D10": "=(C18*D18)+(C19*D19)+(C20*D20)+(C21*D21)",
    "D13": "='DCF Base'!K48",
    "D14": "='DCF Bear'!K48",
    "D15": "='DCF Bull'!K48",
    "C18": "=RV!J5",
    "D18": "=RV!D13",
    "C19": "=RV!J6",
    "D19": "=RV!E13",
    "C20": "=RV!J7",
    "D20": "=RV!F13",
    "C21": "=RV!J8",
    "D21": "=RV!G13",
}


def _col_to_letter(col_num: int) -> str:
    """
    Convert 0-indexed column number to Excel column letter(s).
    
    Args:
        col_num: 0-indexed column number (0 = A, 1 = B, ..., 25 = Z, 26 = AA, etc.)
    
    Returns:
        Excel column letter(s) (e.g., "A", "B", "Z", "AA", "AB")
    """
    result = ""
    col_num += 1  # Convert to 1-indexed
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result


def write_summary_sheet(
    workbook,
    sheet_name: str = "Summary",
    cover_sheet_name: str = "Cover Sheet",
    dcf_base_sheet_name: str = "DCF Base",
    dcf_bear_sheet_name: str = "DCF Bear",
    dcf_bull_sheet_name: str = "DCF Bull",
    rv_sheet_name: str = "RV",
    start_row: int = 0,
    start_col: int = 0,
    custom_formulas: Optional[Dict[str, str]] = None,
) -> None:
    """
    Write summary/valuation summary sheet to Excel worksheet with formulas.
    
    Helper for excel_export.py to write the summary sheet that consolidates
    valuation results from DCF and Comps analyses.
    
    Expected usage:
        workbook = xlsxwriter.Workbook(...)
        write_summary_sheet(workbook)
    
    Args:
        workbook: xlsxwriter Workbook object
        sheet_name: Name of the Excel worksheet (default: "Summary")
        cover_sheet_name: Name of the cover sheet (default: "Cover Sheet")
        dcf_base_sheet_name: Name of the DCF Base sheet (default: "DCF Base")
        dcf_bear_sheet_name: Name of the DCF Bear sheet (default: "DCF Bear")
        dcf_bull_sheet_name: Name of the DCF Bull sheet (default: "DCF Bull")
        rv_sheet_name: Name of the relative valuation/comps sheet (default: "RV")
        start_row: Starting row for the table (0-indexed, default: 0)
        start_col: Starting column for the table (0-indexed, default: 0)
        custom_formulas: Optional dict mapping cell addresses to formula strings.
                        If provided, these formulas will be used instead of default formulas.
                        Example: {"C3": "='Cover Sheet'!B3", "C5": "=(C9*D9)+(C10*D10)"}
    
    The sheet contains:
        - Company name from Cover Sheet
        - Weighted average valuation calculations
        - DCF scenario valuations (Base, Bear, Bull)
        - Relative valuation multiples
        - Excel formulas for calculations (from SUMMARY_FORMULAS or custom_formulas)
    """
    import xlsxwriter
    
    worksheet = workbook.add_worksheet(sheet_name)
    
    # Create formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#D3D3D3',
        'align': 'center',
        'valign': 'vcenter',
    })
    
    label_format = workbook.add_format({
        'bold': True,
        'align': 'left',
    })
    
    number_format = workbook.add_format({
        'num_format': '#,##0',
    })
    
    currency_format = workbook.add_format({
        'num_format': '$#,##0',
    })
    
    percent_format = workbook.add_format({
        'num_format': '0.00%',
    })
    
    formula_format = workbook.add_format({
        'num_format': '#,##0',
        'italic': True,
    })
    
    # Process formulas and replace sheet name placeholders if needed
    # For now, formulas use hardcoded sheet names, but we could make them dynamic
    
    # Write formulas from SUMMARY_FORMULAS dictionary
    # Convert formulas to use actual sheet names if they differ from defaults
    formulas_to_write = {}
    if custom_formulas:
        formulas_to_write = custom_formulas.copy()
    else:
        formulas_to_write = SUMMARY_FORMULAS.copy()
        
        # Replace sheet names in formulas if they differ from defaults
        if cover_sheet_name != "Cover Sheet":
            for cell, formula in formulas_to_write.items():
                formulas_to_write[cell] = formula.replace("'Cover Sheet'", f"'{cover_sheet_name}'")
        
        if dcf_base_sheet_name != "DCF Base":
            for cell, formula in formulas_to_write.items():
                formulas_to_write[cell] = formula.replace("'DCF Base'", f"'{dcf_base_sheet_name}'")
        
        if dcf_bear_sheet_name != "DCF Bear":
            for cell, formula in formulas_to_write.items():
                formulas_to_write[cell] = formula.replace("'DCF Bear'", f"'{dcf_bear_sheet_name}'")
        
        if dcf_bull_sheet_name != "DCF Bull":
            for cell, formula in formulas_to_write.items():
                formulas_to_write[cell] = formula.replace("'DCF Bull'", f"'{dcf_bull_sheet_name}'")
        
        if rv_sheet_name != "RV":
            for cell, formula in formulas_to_write.items():
                formulas_to_write[cell] = formula.replace("RV!", f"'{rv_sheet_name}'!")
    
    # Write formulas to their respective cells
    for cell_address, formula in formulas_to_write.items():
        # Parse cell address (e.g., "C3" -> row 2, col 2)
        col_letter = ''.join([c for c in cell_address if c.isalpha()])
        row_num = int(''.join([c for c in cell_address if c.isdigit()]))
        
        # Convert column letter to 0-indexed number
        col_num = 0
        for char in col_letter:
            col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)
        col_num -= 1  # Convert to 0-indexed
        
        # Convert row number to 0-indexed
        row_num -= 1
        
        # Adjust for start_row and start_col
        actual_row = start_row + row_num
        actual_col = start_col + col_num
        
        # Write formula
        worksheet.write(actual_row, actual_col, formula, formula_format)
    
    # Set column widths for readability
    worksheet.set_column(start_col, start_col + 3, 20)  # Columns A-D


def build_excel_workbook(company_id: int,
                         model_version: Optional[str],
                         db: Session) -> bytes:
    """
    Generate and return an Excel workbook for download.

    Workflow:
        1. Run three-statement model → get projections.
        2. Run DCF → get valuation outputs.
        3. Run comps → get peer comparison outputs.
        4. Create workbook (memory-only).
        5. Write:
            - Historical tables
            - Projection tables
            - DCF valuation summary
            - Comps valuation summary
            - Inputs / assumptions for transparency
        6. Return workbook as bytes.

    TODO:
    - Design tab structure (MVP: 'Inputs', 'Projections', 'DCF', 'Comps')
    - Implement clean table formatting (headers bold, numbers comma-formatted)
    - Add company & model_version labeling to header

    Returns:
        bytes: Excel file content suitable for StreamingResponse
    """
    if not XLSXWRITER_AVAILABLE:
        raise ImportError("xlsxwriter is required for creating new Excel workbooks. Install with: pip install xlsxwriter")
    
    # Create an in-memory workbook buffer
    output = BytesIO()

    # Workbook object
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    # TODO: Create sheets and write data
    # sheet = workbook.add_worksheet("Inputs")
    # sheet.write(0, 0, "TODO: Populate with assumptions")

    workbook.close()
    output.seek(0)
    return output.read()